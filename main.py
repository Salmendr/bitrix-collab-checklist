import os
import threading
import uuid
from collections import defaultdict
from pathlib import Path
from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
import requests
import json
import html
import sqlite3
import re
import mimetypes
from io import BytesIO
from datetime import datetime
from urllib.parse import quote, urlparse
import openpyxl

app = FastAPI()

BASE_DIR = Path(__file__).resolve().parent

VOLUME_DIR = BASE_DIR / "Volume"
VOLUME_DIR.mkdir(parents=True, exist_ok=True)

DB_DIR = VOLUME_DIR / "db"
DB_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = str(DB_DIR / "app.db")

APP_PORTAL_PATH = (os.getenv("APP_PORTAL_PATH", "/marketplace/app/80/") or "/marketplace/app/80/").strip()
APP_BASE_PATH = (os.getenv("APP_BASE_PATH", "") or "").strip()
PUBLIC_APP_BASE_URL = (os.getenv("PUBLIC_APP_BASE_URL", "") or "").strip()
TECH_USER_ID = int(os.getenv("TECH_USER_ID", "138"))
BITRIX_TECH_WEBHOOK_URL = os.getenv("BITRIX_TECH_WEBHOOK_URL", "").strip()
N8N_SHARED_TOKEN = os.getenv("N8N_SHARED_TOKEN", "").strip()
YANDEX_DISK_OAUTH_TOKEN = os.getenv("YANDEX_DISK_OAUTH_TOKEN", "").strip()
YANDEX_DISK_API_BASE = "https://cloud-api.yandex.net/v1/disk"

UPLOAD_ROOT = VOLUME_DIR / "uploads"
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)

CHECKLIST_UPLOAD_ROOT = UPLOAD_ROOT / "checklists"
CHECKLIST_UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)

DEBUG_DIR = BASE_DIR / "debug"
DEBUG_DIR.mkdir(parents=True, exist_ok=True)

DEBUG_LOG_PATH = DEBUG_DIR / "close_popup.log"
EDIT_LOCK_TTL_SECONDS = 45
EDIT_LOCK_HEARTBEAT_SECONDS = 15
ACTIVE_CHECKLIST_LOCKS = {}
ACTIVE_CHECKLIST_LOCKS_GUARD = threading.Lock()

app.mount("/uploads", StaticFiles(directory=str(UPLOAD_ROOT)), name="uploads")

PROJECT_CHECKLISTS = [
    {"key": "id", "title": "Чек-лист ИД"},
    {"key": "opr", "title": "Чек-лист ОПР"},
    {"key": "concept", "title": "Чек-лист Концепция"},
]

CHECKLIST_GROUPS = {
    1: {
        "title": "ИД",
        "items": [
            "ППТ",
            "Выписка ЕГРН",
            "ГПЗУ",
            "Тех задание",
            "ИГДИ",
            "ИГИ",
            "ИЭИ",
            "ИГМИ",
            "ИГИ СМР",
        ],
    },
    2: {
        "title": "ТУ",
        "items": [
            "ТУ Тепловые сети",
            "ТУ Водоснабжение",
            "ТУ Бытовая канализация",
            "ТУ Электроснабжение",
            "ТУ Сети связи",
            "ТУ Наружное освещение",
            "ТУ Ливневая канализация",
            "ТУ Газоснабжение",
        ],
    },
    3: {
        "title": "Прочее",
        "items": [
            "Согласование с Аэропортом",
            "Примыкание к УДС",
            "Порубочный лист",
            "Справка вывоза мусора",
            "СТУ",
            "Сокращение ОКН",
            "Расположение пожарных гидрантов",
        ],
    },
    4: {
        "title": "Не требуется",
        "items": [],
    },
}

STANDARD_ID_YANDEX_FOLDER_SPECS = {
    "ППТ": {
        "alias": "ppt",
        "folderName": "06_ППТ",
        "relativePath": "06_ППТ",
    },
    "Выписка ЕГРН": {
        "alias": "egrn_extract",
        "folderName": "01_Выписка ЕГРН",
        "relativePath": "02_Градплан/01_Выписка ЕГРН",
    },
    "ГПЗУ": {
        "alias": "gpzu",
        "folderName": "02_ГПЗУ",
        "relativePath": "02_Градплан/02_ГПЗУ",
    },
    "Тех задание": {
        "alias": "tz_design",
        "folderName": "01_ТЗ на проектирование",
        "relativePath": "01_ТЗ на проектирование",
    },
    "ИГДИ": {
        "alias": "survey_igdi",
        "folderName": "01_ИГДИ",
        "relativePath": "04_Изыскания/01_ИГДИ",
    },
    "ИГИ": {
        "alias": "survey_igi",
        "folderName": "02_ИГИ",
        "relativePath": "04_Изыскания/02_ИГИ",
    },
    "ИЭИ": {
        "alias": "survey_iei",
        "folderName": "03_ИЭИ",
        "relativePath": "04_Изыскания/03_ИЭИ",
    },
    "ИГМИ": {
        "alias": "survey_igmi",
        "folderName": "04_ИГМИ",
        "relativePath": "04_Изыскания/04_ИГМИ",
    },
    "ИГИ СМР": {
        "alias": "survey_igi_smr",
        "folderName": "05_ИГИ СМР",
        "relativePath": "04_Изыскания/05_ИГИ СМР",
    },
    "ТУ Тепловые сети": {
        "alias": "tu_heat",
        "folderName": "01_ТУ Тепловые сети",
        "relativePath": "03_ТУ/01_ТУ Тепловые сети",
    },
    "ТУ Водоснабжение": {
        "alias": "tu_water",
        "folderName": "02_ТУ Водоснабжение",
        "relativePath": "03_ТУ/02_ТУ Водоснабжение",
    },
    "ТУ Бытовая канализация": {
        "alias": "tu_sewer",
        "folderName": "03_ТУ Бытовая канализация",
        "relativePath": "03_ТУ/03_ТУ Бытовая канализация",
    },
    "ТУ Электроснабжение": {
        "alias": "tu_power",
        "folderName": "04_ТУ Электроснабжение",
        "relativePath": "03_ТУ/04_ТУ Электроснабжение",
    },
    "ТУ Сети связи": {
        "alias": "tu_comm",
        "folderName": "05_ТУ Сети связи",
        "relativePath": "03_ТУ/05_ТУ Сети связи",
    },
    "ТУ Наружное освещение": {
        "alias": "tu_light",
        "folderName": "06_ТУ Наружное освещение",
        "relativePath": "03_ТУ/06_ТУ Наружное освещение",
    },
    "ТУ Ливневая канализация": {
        "alias": "tu_storm",
        "folderName": "07_ТУ Ливневая канализация",
        "relativePath": "03_ТУ/07_ТУ Ливневая канализация",
    },
    "ТУ Газоснабжение": {
        "alias": "tu_gas",
        "folderName": "08_ТУ Газоснабжение",
        "relativePath": "03_ТУ/08_ТУ Газоснабжение",
    },
    "Согласование с Аэропортом": {
        "alias": "ref_airport",
        "folderName": "01_Согласование с аэропортом",
        "relativePath": "05_Справки/01_Согласование с аэропортом",
    },
    "Примыкание к УДС": {
        "alias": "ref_uds",
        "folderName": "02_Примыкание к УДС",
        "relativePath": "05_Справки/02_Примыкание к УДС",
    },
    "Порубочный лист": {
        "alias": "ref_cutting",
        "folderName": "03_Порубочный лист",
        "relativePath": "05_Справки/03_Порубочный лист",
    },
    "Справка вывоза мусора": {
        "alias": "ref_waste",
        "folderName": "04_Справка вывоза мусора",
        "relativePath": "05_Справки/04_Справка вывоза мусора",
    },
    "Расположение пожарных гидрантов": {
        "alias": "ref_hydrants",
        "folderName": "05_Расположение пожарных гидрантов",
        "relativePath": "05_Справки/05_Расположение пожарных гидрантов",
    },
    "Сокращение ОКН": {
        "alias": "okn",
        "folderName": "07_ОКН",
        "relativePath": "07_ОКН",
    },
    "СТУ": {
        "alias": "stu",
        "folderName": "08_СТУ",
        "relativePath": "08_СТУ",
    },
}

STANDARD_OPR_YANDEX_FOLDER_SPECS = {
    "ОПР.ГП": {
        "alias": "opr_gp",
        "folderName": "01_ОПР.ГП",
        "relativePath": "01_ОПР.ГП",
    },
    "ОПР.АР": {
        "alias": "opr_ar",
        "folderName": "02_ОПР.АР",
        "relativePath": "02_ОПР.АР",
    },
    "ИОС 1": {
        "alias": "opr_ios_1",
        "folderName": "ИОС_1",
        "relativePath": "03_ИОС/ИОС_1",
    },
    "ИОС 2": {
        "alias": "opr_ios_2",
        "folderName": "ИОС_2",
        "relativePath": "03_ИОС/ИОС_2",
    },
    "ИОС 3": {
        "alias": "opr_ios_3",
        "folderName": "ИОС_3",
        "relativePath": "03_ИОС/ИОС_3",
    },
    "ИОС 4": {
        "alias": "opr_ios_4",
        "folderName": "ИОС_4",
        "relativePath": "03_ИОС/ИОС_4",
    },
    "ИОС 5": {
        "alias": "opr_ios_5",
        "folderName": "ИОС_5",
        "relativePath": "03_ИОС/ИОС_5",
    },
}

CONCEPT_GROUPS = [
    {
        "id": 1,
        "title": "Общее",
        "items": [
            {"name": "Участок сформирован", "source": "ИД", "statusKind": "bool", "extraPlaceholder": "Например: 2%"},
            {"name": "Наличие ГПЗУ", "source": "ИД", "statusKind": "bool", "extraPlaceholder": "№"},
            {"name": "Наличие проекта планировки", "source": "ИД", "statusKind": "bool", "extraPlaceholder": "Например: проект межевания от 03.07.2024 №3135"},
            {"name": "Наличие ТЗ", "source": "ИД", "statusKind": "bool", "extraPlaceholder": ""},
        ],
    },
    {
        "id": 2,
        "title": "Градостроительные ограничения",
        "items": [
            {"name": "Предельная высота", "source": "ГПЗУ, ПЗЗ", "statusKind": "text", "statusPlaceholder": "___м", "extraPlaceholder": "По проекту: ___"},
            {"name": "Предельная этажность", "source": "ГПЗУ, ПЗЗ", "statusKind": "text", "statusPlaceholder": "___эт.", "extraPlaceholder": "По проекту: ___"},
            {"name": "Процент застройки", "source": "ГПЗУ, ПЗЗ", "statusKind": "text", "statusPlaceholder": "__%", "extraPlaceholder": "По проекту: ___"},
            {"name": "Плотность застройки", "source": "ГПЗУ, ПЗЗ", "statusKind": "text", "statusPlaceholder": "______м2/ГА", "extraPlaceholder": "По проекту: ___"},
            {"name": "Процент озеленения", "source": "ГПЗУ, ПЗЗ", "statusKind": "text", "statusPlaceholder": "__%", "extraPlaceholder": "По проекту: ___"},
            {"name": "Отступы от границ участка", "source": "ГПЗУ, ПЗЗ", "statusKind": "text", "statusPlaceholder": "_м", "extraPlaceholder": "Выполнено / не выполнено"},
            {
                "name": "Вид разрешенного использования",
                "source": "ГПЗУ, ПЗЗ",
                "statusKind": "select",
                "statusOptions": ["Основные", "Условно разрешенные", "Вспомогательные"],
                "extraPlaceholder": "Разрешенный"
            },
        ],
    },
    {
        "id": 3,
        "title": "Приаэродромные территории",
        "items": [
            {"name": "Макс. высота приаэродромных терр.:", "source": "ГПЗУ, Акт ПАТ аэропорта", "statusKind": "text", "statusPlaceholder": "___м", "extraPlaceholder": "По концепции: ___м"},
            {"name": "3 подзона", "source": "ГПЗУ, Акт ПАТ аэропорта", "statusKind": "text", "statusPlaceholder": "___м", "extraPlaceholder": "По концепции: ___м"},
            {"name": "4 подзона", "source": "ГПЗУ, Акт ПАТ аэропорта", "statusKind": "text", "statusPlaceholder": "___м", "extraPlaceholder": "По концепции: ___м"},
            {"name": "Согласование от военного аэропорта", "source": "", "statusKind": "bool", "extraPlaceholder": "Требуется: да/нет"},
        ],
    },
    {
        "id": 4,
        "title": "Зоны затопления и подтопления",
        "items": [
            {"name": "Участок в зоне затопления/подтопления", "source": "ГПЗУ, ПЗЗ", "statusKind": "bool", "extraPlaceholder": "Минимальная отметка земли: ___м"},
        ],
    },
    {
        "id": 5,
        "title": "Дорожно-транспортная сеть",
        "items": [
            {"name": "Требуется согласование примыканий", "source": "", "statusKind": "bool", "extraPlaceholder": ""},
        ],
    },
    {
        "id": 6,
        "title": "Охранные и санитарно-защитные зоны",
        "items": [
            {"name": "Требуется согласование размещения объекта в охранных зонах сетей", "source": "ГПЗУ / СП 124.13330 / Приказ N 197", "statusKind": "bool", "extraPlaceholder": "В т.ч.: ____________"},
            {"name": "Требуется вынос сетей", "source": "ГПЗУ / СП 124.13330 / Приказ N 197", "statusKind": "bool", "extraPlaceholder": "В т.ч.: ____________"},
            {"name": "Учтены отступы от наружных ТС до зданий, парковок, площадок", "source": "ГПЗУ / СП 124.13330 / Приказ N 197", "statusKind": "bool", "extraPlaceholder": ""},
            {"name": "Размещение в зонах ОКН", "source": "ГПЗУ / СП 124.13330 / Приказ N 197", "statusKind": "bool", "extraPlaceholder": "В т.ч.: ____________"},
            {"name": "Учтены санитарные разрывы от парковочных мест (кроме гостевых жилых домов)", "source": "СанПиН 2.2.1/2.1.1.1200-03", "statusKind": "bool", "extraPlaceholder": "До жилых домов, площадок благоустройства, школ, ДОУ, больниц и т.п."},
            {"name": "Учтены санитарно-защитные зоны существующих объектов", "source": "СанПиН 2.2.1/2.1.1.1200-03", "statusKind": "bool", "extraPlaceholder": "В т.ч.: ____________"},
            {"name": "Учтены санитарно-защитные зоны проектируемых объектов", "source": "СанПиН 2.2.1/2.1.1.1200-03", "statusKind": "bool", "extraPlaceholder": "В т.ч.: площадки ТКО, парковки, ТП, КНС, ________"},
        ],
    },
    {
        "id": 7,
        "title": "Пожарные требования",
        "items": [
            {"name": "Требуется разработка СТУ", "source": "ФЗ 123 / СП 4.13130", "statusKind": "bool", "extraPlaceholder": "При отступлениях от требований СП"},
            {"name": "Пожарные проезды в границах участка", "source": "ФЗ 123 / СП 4.13130", "statusKind": "bool", "extraPlaceholder": ""},
            {"name": "Требуется разработка ПТП по проездам", "source": "ФЗ 123 / СП 4.13130", "statusKind": "bool", "extraPlaceholder": "При отступлениях от требований СП"},
            {"name": "Пожарные разрывы от зданий и сооружений", "source": "ФЗ 123 / СП 4.13130", "statusKind": "bool", "extraPlaceholder": "Выполнено: да/нет"},
        ],
    },
    {
        "id": 8,
        "title": "Инсоляция",
        "items": [
            {"name": "Обеспечена инсоляция проектируемого объекта", "source": "СанПиН 1.2.3685-21 пп.165-168", "statusKind": "bool", "extraPlaceholder": "в т.ч. площадок"},
            {"name": "Обеспечена инсоляция окружающей застройки", "source": "СанПиН 1.2.3685-21 пп.165-168", "statusKind": "bool", "extraPlaceholder": "Запрошены тех. паспорта зданий по адресам: 1. __________"},
        ],
    },
    {
        "id": 9,
        "title": "Прочие требования",
        "items": [
            {"name": "Размещена площадка ТКО", "source": "СанПиН", "statusKind": "bool", "extraPlaceholder": "min 20м до домов и площадок (8м при раздельном сборе мусора)"},
            {"name": "К ТП, ТКО и проч. обеспечен проезд", "source": "", "statusKind": "bool", "extraPlaceholder": ""},
            {"name": "Объект обеспечен требуемым количеством машиномест в границах ЗУ*", "source": "РНГП", "statusKind": "bool", "extraPlaceholder": ""},
            {"name": "Объект обеспечен требуемым количеством площадок благоустройства в границах ЗУ*", "source": "РНГП", "statusKind": "bool", "extraPlaceholder": ""},
        ],
    },
    {
        "id": 10,
        "title": "Не требуется",
        "items": [],
    },
]

OPR_GROUPS = [
    {
        "id": 1,
        "title": "ОПР",
        "items": [
            "ОПР.ГП",
            "ОПР.АР",
            "ИОС 1",
            "ИОС 2",
            "ИОС 3",
            "ИОС 4",
            "ИОС 5",
        ],
    },
    {
        "id": 2,
        "title": "Не требуется",
        "items": [],
    },
]

STATUS_OPTIONS = [
    "",
    "Есть",
    "Нет",
    "Не требуется",
]

PRIORITY_OPTIONS = ["white", "green", "gray"]


def build_default_groups(checklist_key: str = "id"):
    if checklist_key == "concept":
        return [{"id": group["id"], "title": group["title"]} for group in CONCEPT_GROUPS]
    if checklist_key == "opr":
        return [{"id": group["id"], "title": group["title"]} for group in OPR_GROUPS]

    return [
        {"id": group_id, "title": group_data["title"]}
        for group_id, group_data in CHECKLIST_GROUPS.items()
    ]


def resolve_group_id(name: str) -> int:
    name = str(name or "").strip()

    for group_id, group_data in CHECKLIST_GROUPS.items():
        if group_id == 4:
            continue
        if name in group_data["items"]:
            return group_id

    return 3


def resolve_concept_group_id_by_item_id_or_name(item: dict) -> int:
    item_id = str(item.get("id") or "")
    name = str(item.get("name") or "").strip()

    if item_id.startswith("concept_g"):
        try:
            group_part = item_id.split("_", 2)[1]
            if group_part.startswith("g"):
                group_id = int(group_part[1:])
                if group_id != 10:
                    return group_id
        except Exception:
            pass

    for group in CONCEPT_GROUPS:
        if group["id"] == 10:
            continue

        for order, spec in enumerate(group["items"], start=1):
            expected_id = f"concept_g{group['id']}_{order}"
            if item_id == expected_id or name == spec["name"]:
                return group["id"]

    current_group = int(item.get("group") or 1)
    return 1 if current_group == 10 else current_group


def resolve_opr_group_id_by_item_id_or_name(item: dict) -> int:
    item_id = str(item.get("id") or "")
    name = str(item.get("name") or "").strip()

    if item_id.startswith("opr_g"):
        parts = item_id.split("_", 2)
        if len(parts) > 1 and parts[1].startswith("g"):
            try:
                group_id = int(parts[1][1:])
                if group_id == 2:
                    return 2
            except Exception:
                pass

    for group in OPR_GROUPS:
        if group["id"] == 2:
            continue

        for order, item_name in enumerate(group["items"], start=1):
            expected_id = f"opr_g{group['id']}_{order}"
            if item_id == expected_id or name == item_name:
                return group["id"]

    current_group = int(item.get("group") or 1)
    return 1 if current_group == 2 else current_group


def clean_cell_value(value):
    value = str(value or "").strip()
    if value == "—":
        return ""
    return value


def normalize_priority(value: str) -> str:
    value = str(value or "").strip().lower()
    if value in PRIORITY_OPTIONS:
        return value
    return "white"


def normalize_status(value: str) -> str:
    value = clean_cell_value(value)

    if not value:
        return ""

    normalized_map = {
        "есть": "Есть",
        "нет": "Нет",
        "не требуется": "Не требуется",
        "подписан": "Есть",
        "запрос опросного листа": "",
        "договор тех прис": "",
    }

    key = value.strip().lower()
    if key in normalized_map:
        return normalized_map[key]

    if value in STATUS_OPTIONS:
        return value

    return ""


def normalize_date_string(value: str) -> str:
    value = clean_cell_value(value)
    if not value:
        return ""

    # Оставляем как есть, если это уже формат ДД.ММ.ГГГГ
    parts = value.split(".")
    if len(parts) == 3 and len(parts[0]) == 2 and len(parts[1]) == 2 and len(parts[2]) == 4:
        return value

    return value


def build_item_id(group_id: int, order: int) -> str:
    return f"item_g{group_id}_{order}"


def build_project_checklists():
    return [dict(x) for x in PROJECT_CHECKLISTS]


def derive_indicator_from_status(status: str) -> str:
    status = normalize_status(status)
    if status == "Есть":
        return "green"
    if status == "Нет" or status == "Не требуется":
        return "gray"
    return "white"


def move_item_to_required_group(item: dict) -> int:
    status = normalize_status(item.get("status"))
    if status == "Не требуется":
        return 4

    current_group = item.get("group")
    if isinstance(current_group, int) and current_group in [1, 2, 3]:
        return current_group

    return resolve_group_id(item.get("name"))


def build_upload_rel_path(dialog_id: str, item_id: str, filename: str) -> str:
    dialog_id = normalize_dialog_id(dialog_id)
    safe_name = Path(filename or "file.bin").name
    unique_name = f"{uuid.uuid4().hex}_{safe_name}"
    return f"checklists/{dialog_id}/{item_id}/{unique_name}"


def remove_item_document_file(item: dict):
    doc_url = str(item.get("documentUrl") or "").strip()
    if not doc_url.startswith("/uploads/"):
        return

    rel_path = doc_url.replace("/uploads/", "", 1)
    file_path = UPLOAD_ROOT / rel_path

    try:
        if file_path.exists():
            file_path.unlink()
    except Exception:
        pass

def format_file_size(size_bytes: int) -> str:
    try:
        size = int(size_bytes or 0)
    except Exception:
        size = 0

    if size <= 0:
        return ""

    units = ["Б", "КБ", "МБ", "ГБ"]
    value = float(size)
    unit_index = 0

    while value >= 1024 and unit_index < len(units) - 1:
        value /= 1024.0
        unit_index += 1

    if unit_index == 0:
        return f"{int(value)} {units[unit_index]}"

    if value >= 100:
        return f"{value:.0f} {units[unit_index]}"
    if value >= 10:
        return f"{value:.1f} {units[unit_index]}"
    return f"{value:.2f} {units[unit_index]}"


def build_folder_view_url(dialog_id: str, checklist_key: str, item_id: str) -> str:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    item_id = str(item_id or "").strip()
    base_path = normalize_base_path(APP_BASE_PATH)

    return (
        f"{base_path}/api/checklist/folder"
        f"?dialogId={quote(dialog_id)}"
        f"&checklistKey={quote(checklist_key)}"
        f"&itemId={quote(item_id)}"
    )


def build_document_view_url(dialog_id: str, checklist_key: str, item_id: str, document_id: str) -> str:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    item_id = str(item_id or "").strip()
    document_id = str(document_id or "").strip()
    base_path = normalize_base_path(APP_BASE_PATH)

    return (
        f"{base_path}/api/checklist/file"
        f"?dialogId={quote(dialog_id)}"
        f"&checklistKey={quote(checklist_key)}"
        f"&itemId={quote(item_id)}"
        f"&documentId={quote(document_id)}"
    )


def get_upload_file_path_from_url(document_url: str):
    document_url = clean_cell_value(document_url)
    if not document_url.startswith("/uploads/"):
        return None

    rel_path = document_url.replace("/uploads/", "", 1)
    return UPLOAD_ROOT / rel_path


def can_preview_in_browser(filename: str, media_type: str = "") -> bool:
    filename = str(filename or "").strip().lower()
    media_type = str(media_type or "").strip().lower()

    if media_type.startswith("image/"):
        return True
    if media_type.startswith("text/"):
        return True

    previewable_media_types = {
        "application/pdf",
        "application/json",
        "application/xml",
        "text/csv",
    }

    previewable_extensions = {
        ".pdf", ".txt", ".md", ".csv", ".json", ".xml", ".html", ".htm"
    }

    if media_type in previewable_media_types:
        return True

    suffix = Path(filename).suffix.lower()
    return suffix in previewable_extensions

def slugify_folder_part(value: str) -> str:
    value = str(value or "").strip().lower()

    translit_map = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d",
        "е": "e", "ё": "e", "ж": "zh", "з": "z", "и": "i",
        "й": "y", "к": "k", "л": "l", "м": "m", "н": "n",
        "о": "o", "п": "p", "р": "r", "с": "s", "т": "t",
        "у": "u", "ф": "f", "х": "h", "ц": "ts", "ч": "ch",
        "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "",
        "э": "e", "ю": "yu", "я": "ya"
    }

    chars = []
    for ch in value:
        if ch in translit_map:
            chars.append(translit_map[ch])
        else:
            chars.append(ch)

    value = "".join(chars)
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value or "item"


def build_folder_key(checklist_key: str, item_name: str, item_id: str = "") -> str:
    checklist_key = normalize_checklist_key(checklist_key)
    raw_name = clean_cell_value(item_name)
    raw_item_id = str(item_id or "").strip()

    special_map = {
        ("id", "Тех задание"): "id_tech_task",
    }

    special_key = (checklist_key, raw_name)
    if special_key in special_map:
        return special_map[special_key]

    if raw_item_id:
        safe_item_id = re.sub(r"[^a-zA-Z0-9_]+", "_", raw_item_id).strip("_").lower()
        if safe_item_id:
            return f"{checklist_key}_{safe_item_id}"

    return f"{checklist_key}_{slugify_folder_part(raw_name)}"

def normalize_id_builtin_name(name: str) -> str:
    value = clean_cell_value(name)

    rename_map = {
        "ТУ ТС": "ТУ Тепловые сети",
        "ТУ Свет": "ТУ Электроснабжение",
        "ТУ СС": "ТУ Сети связи",
        "ТУ Ливневка": "ТУ Ливневая канализация",
        "ТУ Газ": "ТУ Газоснабжение",
        "Аэропорт": "Согласование с Аэропортом",
        "Примыкание ОДД": "Примыкание к УДС",
        "Расположение пож гидрантов": "Расположение пожарных гидрантов",
    }

    return rename_map.get(value, value)


def build_current_id_default_name_set() -> set[str]:
    result = set()

    for group_id, group_data in CHECKLIST_GROUPS.items():
        if group_id == 4:
            continue

        for item_name in group_data["items"]:
            normalized_name = clean_cell_value(item_name).lower()
            if normalized_name:
                result.add(normalized_name)

    return result

def normalize_document_record(doc: dict) -> dict:
    doc = dict(doc or {})

    file_url = clean_cell_value(doc.get("fileUrl") or doc.get("url") or doc.get("documentUrl"))
    preview_url = clean_cell_value(doc.get("previewUrl") or file_url)
    path = clean_cell_value(doc.get("path") or file_url)
    name = clean_cell_value(doc.get("name") or doc.get("documentName"))

    if not name and file_url:
        try:
            name = Path(urlparse(file_url).path).name
        except Exception:
            name = ""

    return {
        "id": clean_cell_value(doc.get("id")) or uuid.uuid4().hex,
        "name": name,
        "path": path,
        "fileUrl": file_url,
        "previewUrl": preview_url,
        "size": int(doc.get("size") or 0),
        "modifiedAt": clean_cell_value(doc.get("modifiedAt")),
        "source": clean_cell_value(doc.get("source")) or "local",

        "mirrorStatus": clean_cell_value(doc.get("mirrorStatus")) or "",
        "mirrorError": clean_cell_value(doc.get("mirrorError")) or "",
        "yandexPath": clean_cell_value(doc.get("yandexPath")) or "",
        "yandexFileUrl": clean_cell_value(doc.get("yandexFileUrl")) or "",
        "yandexFolderAlias": clean_cell_value(doc.get("yandexFolderAlias")) or "",
    }


def normalize_documents_list(value) -> list[dict]:
    if not isinstance(value, list):
        return []

    result = []
    for raw_doc in value:
        normalized = normalize_document_record(raw_doc)
        if not normalized.get("name") and not normalized.get("fileUrl"):
            continue
        result.append(normalized)

    result.sort(key=lambda x: (x.get("name") or "").lower())
    return result


def migrate_legacy_document_fields(item: dict) -> dict:
    item = dict(item or {})

    documents = item.get("documents")
    if isinstance(documents, list):
        item["documents"] = normalize_documents_list(documents)
        return item

    legacy_url = clean_cell_value(item.get("documentUrl"))
    legacy_name = clean_cell_value(item.get("documentName"))

    if legacy_url or legacy_name:
        item["documents"] = normalize_documents_list([
            {
                "id": uuid.uuid4().hex,
                "name": legacy_name,
                "path": legacy_url,
                "fileUrl": legacy_url,
                "previewUrl": legacy_url,
                "size": 0,
                "modifiedAt": "",
                "source": "local",
            }
        ])
    else:
        item["documents"] = []

    return item

def build_default_checklist_template(dialog_id: str = "", checklist_key: str = "id"):
    items = []

    if checklist_key == "concept":
        for group in CONCEPT_GROUPS:
            if group["id"] == 10:
                continue

            for order, spec in enumerate(group["items"], start=1):
                item_id = f"concept_g{group['id']}_{order}"
                item_name = spec["name"]

                items.append({
                    "id": item_id,
                    "group": group["id"],
                    "order": order,
                    "name": item_name,
                    "source": spec.get("source", ""),
                    "statusKind": spec.get("statusKind", "bool"),
                    "statusOptions": spec.get("statusOptions", []),
                    "statusPlaceholder": spec.get("statusPlaceholder", ""),
                    "status": "",
                    "extraInfo": "",
                    "extraInfoPlaceholder": spec.get("extraPlaceholder", ""),
                    "folderKey": build_folder_key("concept", item_name, item_id),
                    "folderPath": "",
                    "folderUrl": "",
                    "documents": [],
                    "isCustom": False,
                })

        return normalize_checklist_data({
            "title": "Чек-лист Концепция",
            "checklistKey": "concept",
            "collabTitle": "",
            "contractDeadline": "",
            "startDate": "",
            "groups": build_default_groups("concept"),
            "items": items,
            "notice": "",
        }, "concept")

    if checklist_key == "opr":
        for group in OPR_GROUPS:
            if group["id"] == 2:
                continue

            for order, name in enumerate(group["items"], start=1):
                item_id = f"opr_g{group['id']}_{order}"

                items.append({
                    "id": item_id,
                    "group": group["id"],
                    "order": order,
                    "name": name,
                    "priority": "white",
                    "status": "",
                    "plan": "",
                    "fact": "",
                    "folderKey": build_folder_key("opr", name, item_id),
                    "folderPath": "",
                    "folderUrl": "",
                    "documents": [],
                    "documentUrl": "",
                    "documentName": "",
                    "isCustom": False,
                })

        return normalize_checklist_data({
            "title": "Чек-лист ОПР",
            "checklistKey": "opr",
            "collabTitle": "",
            "contractDeadline": "",
            "startDate": "",
            "groups": build_default_groups("opr"),
            "items": items,
            "notice": "",
        }, "opr")
    
    for group_id, group_data in CHECKLIST_GROUPS.items():
        if group_id == 4:
            continue

        for order, name in enumerate(group_data["items"], start=1):
            item_id = build_item_id(group_id, order)

            items.append({
                "id": item_id,
                "group": group_id,
                "order": order,
                "name": name,
                "priority": "white",
                "status": "",
                "plan": "",
                "fact": "",
                "folderKey": build_folder_key("id", name, item_id),
                "folderPath": "",
                "folderUrl": "",
                "documents": [],
                "isCustom": False,
            })

    return normalize_checklist_data({
        "title": "Чек-лист ИД",
        "checklistKey": "id",
        "collabTitle": "",
        "contractDeadline": "",
        "startDate": "",
        "groups": build_default_groups("id"),
        "items": items,
        "notice": "",
    }, "id")


def calculate_progress(items: list) -> dict:
    items = items or []

    active_items = [x for x in items if normalize_status(x.get("status")) != "Не требуется"]
    completed_items = [x for x in active_items if normalize_status(x.get("status")) == "Есть"]

    active_count = len(active_items)
    completed_count = len(completed_items)

    progress_percent = 0
    if active_count > 0:
        progress_percent = round((completed_count / active_count) * 100)

    return {
        "activeCount": active_count,
        "completedCount": completed_count,
        "progressPercent": progress_percent,
    }


def calculate_concept_progress(items: list) -> dict:
    items = items or []

    def is_active(item: dict) -> bool:
        return clean_cell_value(item.get("status")) != "Не требуется"

    def is_completed(item: dict) -> bool:
        status = clean_cell_value(item.get("status"))
        status_kind = clean_cell_value(item.get("statusKind")) or "bool"

        if status == "Не требуется":
            return False

        if status_kind == "bool":
            return status == "Да"

        return bool(status)

    active_items = [item for item in items if is_active(item)]
    completed_items = [item for item in active_items if is_completed(item)]

    active_count = len(active_items)
    completed_count = len(completed_items)
    progress_percent = round((completed_count / active_count) * 100) if active_count else 0

    return {
        "activeCount": active_count,
        "completedCount": completed_count,
        "progressPercent": progress_percent,
    }


def calculate_opr_progress(items: list) -> dict:
    items = items or []

    active_items = [item for item in items if normalize_status(item.get("status")) != "Не требуется"]
    completed_items = [item for item in active_items if normalize_status(item.get("status")) == "Есть"]

    active_count = len(active_items)
    completed_count = len(completed_items)
    progress_percent = round((completed_count / active_count) * 100) if active_count else 0

    return {
        "activeCount": active_count,
        "completedCount": completed_count,
        "progressPercent": progress_percent,
    }


def normalize_checklist_data(data: dict, checklist_key: str = "id") -> dict:
    data = dict(data or {})
    checklist_key = str(checklist_key or data.get("checklistKey") or "id").strip() or "id"

    def prepare_item_common(item: dict, default_name: str = "") -> tuple[dict, list[dict], dict, str, str, str]:
        item = migrate_legacy_document_fields(dict(item or {}))

        name = clean_cell_value(item.get("name")) or clean_cell_value(default_name)
        documents = normalize_documents_list(item.get("documents"))
        first_doc = documents[0] if documents else {}

        folder_key = clean_cell_value(item.get("folderKey")) or build_folder_key(
            checklist_key,
            name,
            item.get("id") or ""
        )
        folder_path = clean_cell_value(item.get("folderPath"))
        folder_url = clean_cell_value(item.get("folderUrl"))

        legacy_document_url = clean_cell_value(item.get("documentUrl")) or clean_cell_value(first_doc.get("fileUrl"))
        legacy_document_name = clean_cell_value(item.get("documentName")) or clean_cell_value(first_doc.get("name"))

        return item, documents, first_doc, folder_key, folder_path, folder_url, legacy_document_url, legacy_document_name

    if checklist_key == "concept":
        raw_items = data.get("items", []) or []
        normalized_items = []

        for raw_item in raw_items:
            item, documents, first_doc, folder_key, folder_path, folder_url, legacy_document_url, legacy_document_name = prepare_item_common(raw_item)

            name = clean_cell_value(item.get("name"))
            if not name:
                continue

            normalized_items.append({
                "id": str(item.get("id") or ""),
                "group": int(item.get("group") or 0),
                "order": int(item.get("order") or 0),
                "name": name,
                "source": clean_cell_value(item.get("source")),
                "statusKind": clean_cell_value(item.get("statusKind")) or "bool",
                "statusOptions": item.get("statusOptions") or [],
                "statusPlaceholder": clean_cell_value(item.get("statusPlaceholder")),
                "status": clean_cell_value(item.get("status")),
                "extraInfo": clean_cell_value(item.get("extraInfo")),
                "extraInfoPlaceholder": clean_cell_value(item.get("extraInfoPlaceholder")),
                "folderKey": folder_key,
                "folderPath": folder_path,
                "folderUrl": folder_url,
                "documents": documents,
                "documentUrl": legacy_document_url,
                "documentName": legacy_document_name,
                "isCustom": bool(item.get("isCustom", False)),
                "priority": clean_cell_value(item.get("priority")) or "white",
            })

        normalized_items.sort(key=lambda x: (x["group"], x["order"], x["name"]))

        for group in CONCEPT_GROUPS:
            group_items = [x for x in normalized_items if x["group"] == group["id"]]
            for order, item in enumerate(group_items, start=1):
                item["order"] = order
                if not item.get("id"):
                    item["id"] = f"concept_g{group['id']}_{order}"
                if not item.get("folderKey"):
                    item["folderKey"] = build_folder_key("concept", item.get("name"), item.get("id"))

        progress = calculate_concept_progress(normalized_items)

        return {
            "title": data.get("title") or "Чек-лист Концепция",
            "checklistKey": "concept",
            "collabTitle": clean_cell_value(data.get("collabTitle")),
            "contractDeadline": "",
            "startDate": "",
            "groups": [{"id": group["id"], "title": group["title"]} for group in CONCEPT_GROUPS],
            "projectChecklists": build_project_checklists(),
            "items": normalized_items,
            "notice": data.get("notice", ""),
            "activeCount": progress["activeCount"],
            "completedCount": progress["completedCount"],
            "progressPercent": progress["progressPercent"],
        }

    if checklist_key == "opr":
        raw_items = data.get("items", []) or []
        normalized_items = []

        for raw_item in raw_items:
            item, documents, first_doc, folder_key, folder_path, folder_url, legacy_document_url, legacy_document_name = prepare_item_common(raw_item)

            name = clean_cell_value(item.get("name"))
            if not name:
                continue

            status = normalize_status(item.get("status"))
            group_id = int(item.get("group") or 0)

            if status == "Не требуется":
                group_id = 2
            elif group_id == 2 or not group_id:
                group_id = resolve_opr_group_id_by_item_id_or_name(item)

            normalized_items.append({
                "id": str(item.get("id") or ""),
                "group": group_id,
                "order": int(item.get("order") or 0),
                "name": name,
                "priority": derive_indicator_from_status(status),
                "status": status,
                "plan": normalize_date_string(item.get("plan") or item.get("plannedDate")),
                "fact": normalize_date_string(item.get("fact")),
                "folderKey": folder_key,
                "folderPath": folder_path,
                "folderUrl": folder_url,
                "documents": documents,
                "documentUrl": legacy_document_url,
                "documentName": legacy_document_name,
                "isCustom": bool(item.get("isCustom", False)),
            })

        existing_names = {
            clean_cell_value(existing_item.get("name")).lower()
            for existing_item in normalized_items
            if clean_cell_value(existing_item.get("name"))
        }

        for group in OPR_GROUPS:
            if group["id"] == 2:
                continue

            for default_order, default_name in enumerate(group["items"], start=1):
                normalized_name = clean_cell_value(default_name).lower()
                if not normalized_name or normalized_name in existing_names:
                    continue

                migrated_item_id = f"opr_g{group['id']}_{default_order}_migrated_{slugify_folder_part(default_name)}"

                normalized_items.append({
                    "id": migrated_item_id,
                    "group": group["id"],
                    "order": default_order,
                    "name": default_name,
                    "priority": "white",
                    "status": "",
                    "plan": "",
                    "fact": "",
                    "folderKey": build_folder_key("opr", default_name, migrated_item_id),
                    "folderPath": "",
                    "folderUrl": "",
                    "documents": [],
                    "documentUrl": "",
                    "documentName": "",
                    "isCustom": False,
                })

                existing_names.add(normalized_name)

        normalized_items.sort(key=lambda x: (x["group"], x["order"], x["name"]))

        for group in OPR_GROUPS:
            group_items = [x for x in normalized_items if x["group"] == group["id"]]
            for order, item in enumerate(group_items, start=1):
                item["order"] = order
                if not item.get("id"):
                    item["id"] = f"opr_g{group['id']}_{order}"
                if not item.get("folderKey"):
                    item["folderKey"] = build_folder_key("opr", item.get("name"), item.get("id"))

        progress = calculate_opr_progress(normalized_items)

        return {
            "title": data.get("title") or "Чек-лист ОПР",
            "checklistKey": "opr",
            "collabTitle": clean_cell_value(data.get("collabTitle")),
            "contractDeadline": "",
            "startDate": "",
            "groups": [{"id": group["id"], "title": group["title"]} for group in OPR_GROUPS],
            "projectChecklists": build_project_checklists(),
            "items": normalized_items,
            "notice": data.get("notice", ""),
            "activeCount": progress["activeCount"],
            "completedCount": progress["completedCount"],
            "progressPercent": progress["progressPercent"],
        }

    raw_items = data.get("items", []) or []
    normalized_items = []

    group_order_counters = {1: 0, 2: 0, 3: 0, 4: 0}
    current_id_default_names = build_current_id_default_name_set()

    for raw_item in raw_items:
        item, documents, first_doc, folder_key, folder_path, folder_url, legacy_document_url, legacy_document_name = prepare_item_common(raw_item)

        item_name = normalize_id_builtin_name(item.get("name"))
        if not item_name:
            continue

        is_custom = bool(item.get("isCustom", False))
        status = normalize_status(item.get("status"))
        target_group_id = move_item_to_required_group({
            "group": item.get("group"),
            "name": item_name,
            "status": status,
        })

        if not is_custom and target_group_id != 4 and item_name.lower() not in current_id_default_names:
            continue

        group_order_counters[target_group_id] += 1
        default_order = group_order_counters[target_group_id]

        normalized_items.append({
            "id": str(item.get("id") or build_item_id(target_group_id, default_order)),
            "group": target_group_id,
            "order": int(item.get("order") or default_order),
            "name": item_name,
            "priority": derive_indicator_from_status(status),
            "status": status,
            "plan": normalize_date_string(item.get("plan")),
            "fact": normalize_date_string(item.get("fact")),
            "folderKey": folder_key,
            "folderPath": folder_path,
            "folderUrl": folder_url,
            "documents": documents,
            "documentUrl": legacy_document_url,
            "documentName": legacy_document_name,
            "isCustom": is_custom,
        })

    deduped_items = []
    seen_builtin_names = set()

    for existing_item in normalized_items:
        name_key = clean_cell_value(existing_item.get("name")).lower()

        if not existing_item.get("isCustom") and int(existing_item.get("group") or 0) != 4:
            if name_key in seen_builtin_names:
                continue
            seen_builtin_names.add(name_key)

        deduped_items.append(existing_item)

    normalized_items = deduped_items

    existing_names = {
        clean_cell_value(existing_item.get("name")).lower()
        for existing_item in normalized_items
        if clean_cell_value(existing_item.get("name"))
    }

    for default_group_id, group_data in CHECKLIST_GROUPS.items():
        if default_group_id == 4:
            continue

        for default_order, default_name in enumerate(group_data["items"], start=1):
            normalized_name = clean_cell_value(default_name).lower()
            if not normalized_name or normalized_name in existing_names:
                continue

            migrated_item_id = f"{build_item_id(default_group_id, default_order)}_migrated_{slugify_folder_part(default_name)}"

            normalized_items.append({
                "id": migrated_item_id,
                "group": default_group_id,
                "order": default_order,
                "name": default_name,
                "priority": "white",
                "status": "",
                "plan": "",
                "fact": "",
                "folderKey": build_folder_key("id", default_name, migrated_item_id),
                "folderPath": "",
                "folderUrl": "",
                "documents": [],
                "documentUrl": "",
                "documentName": "",
                "isCustom": False,
            })

            existing_names.add(normalized_name)

    default_id_order = {}
    for default_group_id, group_data in CHECKLIST_GROUPS.items():
        if default_group_id == 4:
            continue

        for default_order, default_name in enumerate(group_data["items"], start=1):
            default_id_order[(default_group_id, clean_cell_value(default_name).lower())] = default_order

    normalized_items.sort(
        key=lambda x: (
            x["group"],
            default_id_order.get(
                (int(x["group"]), clean_cell_value(x.get("name")).lower()),
                10000
            ),
            x["order"],
            x["name"],
        )
    )

    for group_id in [1, 2, 3, 4]:
        group_items = [x for x in normalized_items if x["group"] == group_id]
        for order, item in enumerate(group_items, start=1):
            item["order"] = order
            if not item.get("id"):
                item["id"] = build_item_id(group_id, order)
            if not item.get("folderKey"):
                item["folderKey"] = build_folder_key("id", item.get("name"), item.get("id"))

    progress = calculate_progress(normalized_items)

    return {
        "title": data.get("title") or "Чек-лист ИД",
        "checklistKey": "id",
        "collabTitle": clean_cell_value(data.get("collabTitle")),
        "contractDeadline": "",
        "startDate": "",
        "groups": [
            {"id": 1, "title": "ИД"},
            {"id": 2, "title": "ТУ"},
            {"id": 3, "title": "Прочее"},
            {"id": 4, "title": "Не требуется"},
        ],
        "projectChecklists": build_project_checklists(),
        "items": normalized_items,
        "notice": data.get("notice", ""),
        "activeCount": progress["activeCount"],
        "completedCount": progress["completedCount"],
        "progressPercent": progress["progressPercent"],
    }

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS checklists (
            dialog_id TEXT PRIMARY KEY,
            title TEXT,
            data_json TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS project_storage_contexts (
            dialog_id TEXT PRIMARY KEY,
            project_id TEXT,
            project_name TEXT,
            provider TEXT,
            storage_mode_json TEXT,
            yandex_json TEXT,
            item_mappings_json TEXT,
            updated_at TEXT
        )
    """)
    conn.commit()
    conn.close()


# Страховочный вызов при импорте модуля
init_db()


@app.on_event("startup")
def startup_event():
    init_db()


def normalize_domain(value: str) -> str:
    value = (value or "").strip()
    value = value.replace("https://", "").replace("http://", "").strip("/")
    return value

def normalize_base_path(value: str) -> str:
    value = str(value or "").strip()
    if not value or value == "/":
        return ""
    if not value.startswith("/"):
        value = "/" + value
    return value.rstrip("/")


def get_public_app_base_path(request: Request) -> str:
    if APP_BASE_PATH:
        return normalize_base_path(APP_BASE_PATH)

    forwarded_prefix = request.headers.get("x-forwarded-prefix", "").split(",")[0].strip()
    if forwarded_prefix:
        return normalize_base_path(forwarded_prefix)

    root_path = str(request.scope.get("root_path") or "").strip()
    return normalize_base_path(root_path)


def get_public_origin(request: Request) -> str:
    if PUBLIC_APP_BASE_URL:
        parsed = urlparse(PUBLIC_APP_BASE_URL)
        if parsed.scheme and parsed.netloc:
            return f"{parsed.scheme}://{parsed.netloc}"

    proto = (request.headers.get("x-forwarded-proto") or request.url.scheme or "https").split(",")[0].strip()
    host = (request.headers.get("x-forwarded-host") or request.headers.get("host") or request.url.netloc).split(",")[0].strip()
    return f"{proto}://{host}"


def get_public_app_base_url(request: Request) -> str:
    if PUBLIC_APP_BASE_URL:
        return PUBLIC_APP_BASE_URL.rstrip("/")
    return f"{get_public_origin(request)}{get_public_app_base_path(request)}"

def bitrix_rest_call(domain: str, method: str, access_token: str, payload: dict):
    url = f"https://{domain}/rest/{method}.json"
    response = requests.post(
        url,
        data={**payload, "auth": access_token},
        timeout=30
    )
    try:
        return response.json()
    except Exception:
        return {
            "http_status": response.status_code,
            "text": response.text
        }


def bitrix_webhook_call(method: str, payload: dict):
    if not BITRIX_TECH_WEBHOOK_URL:
        return {
            "error": "TECH_WEBHOOK_NOT_CONFIGURED",
            "error_description": "BITRIX_TECH_WEBHOOK_URL is empty"
        }

    base = BITRIX_TECH_WEBHOOK_URL.rstrip("/")
    url = f"{base}/{method}.json"

    response = requests.post(url, data=payload, timeout=30)

    try:
        return response.json()
    except Exception:
        return {
            "http_status": response.status_code,
            "text": response.text
        }


def write_debug_log(event: str, payload: dict):
    record = {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "event": event,
        "payload": payload,
    }

    with open(DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def make_checklist_lock_key(dialog_id: str, checklist_key: str) -> str:
    return f"{normalize_dialog_id(dialog_id)}::{normalize_checklist_key(checklist_key)}"


def _cleanup_expired_checklist_locks(now_ts: float | None = None):
    now_ts = now_ts or datetime.now().timestamp()
    expired_keys = []

    for lock_key, lock_data in ACTIVE_CHECKLIST_LOCKS.items():
        updated_at = float(lock_data.get("updatedAtTs") or 0)
        if now_ts - updated_at > EDIT_LOCK_TTL_SECONDS:
            expired_keys.append(lock_key)

    for lock_key in expired_keys:
        ACTIVE_CHECKLIST_LOCKS.pop(lock_key, None)


def acquire_checklist_lock(
    dialog_id: str,
    checklist_key: str,
    user_id: str,
    user_name: str,
    lock_id: str,
) -> dict:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    user_id = str(user_id or "").strip()
    user_name = str(user_name or "").strip()
    lock_id = str(lock_id or "").strip()
    now = datetime.now()
    now_ts = now.timestamp()
    lock_key = make_checklist_lock_key(dialog_id, checklist_key)

    with ACTIVE_CHECKLIST_LOCKS_GUARD:
        _cleanup_expired_checklist_locks(now_ts)
        existing = ACTIVE_CHECKLIST_LOCKS.get(lock_key)

        if existing and existing.get("lockId") != lock_id:
            return {
                "ok": True,
                "owned": False,
                "lockedByOther": True,
                "lockId": str(existing.get("lockId") or ""),
                "dialogId": dialog_id,
                "checklistKey": checklist_key,
                "userId": str(existing.get("userId") or ""),
                "userName": str(existing.get("userName") or "Другой сотрудник"),
                "updatedAt": str(existing.get("updatedAt") or ""),
            }

        if not lock_id:
            lock_id = uuid.uuid4().hex

        ACTIVE_CHECKLIST_LOCKS[lock_key] = {
            "lockId": lock_id,
            "dialogId": dialog_id,
            "checklistKey": checklist_key,
            "userId": user_id,
            "userName": user_name or "Неизвестный пользователь",
            "updatedAt": now.isoformat(),
            "updatedAtTs": now_ts,
        }

        return {
            "ok": True,
            "owned": True,
            "lockedByOther": False,
            "lockId": lock_id,
            "dialogId": dialog_id,
            "checklistKey": checklist_key,
            "userId": user_id,
            "userName": user_name or "Неизвестный пользователь",
            "updatedAt": now.isoformat(),
        }


def heartbeat_checklist_lock(
    dialog_id: str,
    checklist_key: str,
    user_id: str,
    user_name: str,
    lock_id: str,
) -> dict:
    return acquire_checklist_lock(dialog_id, checklist_key, user_id, user_name, lock_id)


def release_checklist_lock(dialog_id: str, checklist_key: str, lock_id: str = "", user_id: str = "") -> dict:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    lock_id = str(lock_id or "").strip()
    user_id = str(user_id or "").strip()
    lock_key = make_checklist_lock_key(dialog_id, checklist_key)

    with ACTIVE_CHECKLIST_LOCKS_GUARD:
        _cleanup_expired_checklist_locks()
        existing = ACTIVE_CHECKLIST_LOCKS.get(lock_key)
        if not existing:
            return {
                "ok": True,
                "released": False,
                "dialogId": dialog_id,
                "checklistKey": checklist_key,
            }

        if lock_id and existing.get("lockId") != lock_id:
            return {
                "ok": True,
                "released": False,
                "dialogId": dialog_id,
                "checklistKey": checklist_key,
                "lockedByOther": True,
                "userName": str(existing.get("userName") or ""),
            }

        if not lock_id and user_id and str(existing.get("userId") or "") != user_id:
            return {
                "ok": True,
                "released": False,
                "dialogId": dialog_id,
                "checklistKey": checklist_key,
                "lockedByOther": True,
                "userName": str(existing.get("userName") or ""),
            }

        ACTIVE_CHECKLIST_LOCKS.pop(lock_key, None)
        return {
            "ok": True,
            "released": True,
            "dialogId": dialog_id,
            "checklistKey": checklist_key,
        }

def status_emoji(status: str) -> str:
    status = display_status_text(status)

    if status in {"Есть", "Да"}:
        return "🟢"
    if status == "Нет":
        return "🔴"
    if status == "Не требуется":
        return "🚫"
    return "🟢" if status else "⚪️"


MESSAGE_ALIGNMENT_SPACE = " "
MESSAGE_ALIGNMENT_SPACE_FACTOR = 2
MESSAGE_CHECKLIST_ORDER = ["id", "opr", "concept"]
MESSAGE_CHECKLIST_TITLES = {
    "id": "Чек-лист ИД",
    "concept": "Чек-лист Концепция",
    "opr": "Чек-лист ОПР",
}
MESSAGE_SECTION_LABELS = {
    "status": "Статусы",
    "date": "Даты",
    "extraInfo": "Доп информация",
    "source": "Нормативы",
    "name": "Пункты",
    "document": "Документы",
    "add-item": "Пункты",
}


def display_status_text(status: str) -> str:
    status = clean_cell_value(status)
    normalized = normalize_status(status)
    return normalized or status


def build_progress_text(data: dict) -> str:
    checklist_key = normalize_checklist_key(data.get("checklistKey") or "id")
    items = data.get("items") or []

    if checklist_key == "concept":
        active_items = [item for item in items if clean_cell_value(item.get("status")) != "Не требуется"]
        completed_items = []
        for item in active_items:
            status = clean_cell_value(item.get("status"))
            status_kind = clean_cell_value(item.get("statusKind")) or "bool"
            if status_kind == "bool":
                if status == "Да":
                    completed_items.append(item)
            elif status:
                completed_items.append(item)
    else:
        active_items = [item for item in items if display_status_text(item.get("status")) != "Не требуется"]
        completed_items = [item for item in active_items if display_status_text(item.get("status")) == "Есть"]

    percent = round((len(completed_items) / len(active_items)) * 100) if active_items else 0
    return f"📊Прогресс: {percent}%"


def split_changes(changes: list) -> dict:
    groups = {
        "status": [],
        "date": [],
        "extraInfo": [],
        "source": [],
        "name": [],
        "document": [],
        "add-item": [],
    }

    for change in changes or []:
        field = str(change.get("field") or "").strip()
        if field == "status":
            groups["status"].append(change)
        elif field in {"plan", "plannedDate", "fact"}:
            groups["date"].append(change)
        elif field == "extraInfo":
            groups["extraInfo"].append(change)
        elif field == "source":
            groups["source"].append(change)
        elif field == "name":
            groups["name"].append(change)
        elif field == "document":
            groups["document"].append(change)
        elif field == "add-item":
            groups["add-item"].append(change)

    return groups


def strip_message_markup(value: str) -> str:
    return re.sub(r"\[[^\]]+\]", "", str(value or ""))


def message_visible_length(value: str) -> int:
    return len(strip_message_markup(value))


def pad_message_left(value: str, target_width: int) -> str:
    missing = max(target_width - message_visible_length(value), 0)
    return value + (MESSAGE_ALIGNMENT_SPACE * (missing * MESSAGE_ALIGNMENT_SPACE_FACTOR))


def get_checklist_message_title(checklist_key: str, checklist_title: str = "") -> str:
    normalized_key = normalize_checklist_key(checklist_key)
    cleaned_title = clean_cell_value(checklist_title)
    return cleaned_title or MESSAGE_CHECKLIST_TITLES.get(normalized_key, "Чек-лист ИД")


def get_checklist_link_caption(checklist_key: str) -> str:
    mapping = {
        "id": "ИД",
        "concept": "КОНЦЕПЦИЯ",
        "opr": "ОПР",
    }
    return mapping.get(normalize_checklist_key(checklist_key), "ЧЕК-ЛИСТ")


def build_checklist_message_link(dialog_id: str, checklist_key: str) -> str:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)

    if not dialog_id or not BITRIX_TECH_WEBHOOK_URL:
        return ""

    parsed = urlparse(BITRIX_TECH_WEBHOOK_URL)
    if not parsed.scheme or not parsed.netloc:
        return ""

    base_url = f"{parsed.scheme}://{parsed.netloc}"
    app_path = APP_PORTAL_PATH if APP_PORTAL_PATH.endswith("/") else APP_PORTAL_PATH + "/"
    query = f"dialogId={quote(dialog_id)}&checklistKey={quote(checklist_key)}"
    return f"{base_url}{app_path}?{query}#{query}"


def build_editor_text(editor: dict) -> str:
    editor_name = str(editor.get("name") or "").strip()
    editor_id = str(editor.get("id") or "").strip()

    if editor_name and editor_id:
        return f"👤Кем изменено: {editor_name} (ID {editor_id})"
    if editor_name:
        return f"👤Кем изменено: {editor_name}"
    if editor_id:
        return f"👤Кем изменено: ID {editor_id}"
    return "👤Кем изменено: неизвестно"


def normalize_message_value(field: str, value, checklist_key: str) -> str:
    raw = clean_cell_value(value)
    if field == "status":
        return display_status_text(raw)
    return raw


def build_change_left_label(field: str, checklist_key: str) -> str:
    if field == "status":
        return "Статус"
    if field in {"date", "plan", "plannedDate"}:
        return "План"
    if field == "fact":
        return "Факт"
    if field == "extraInfo":
        return "Доп информация"
    if field == "source":
        return "Нормативы"
    if field == "name":
        return "Пункт"
    if field == "document":
        return "Документ"
    if field == "add-item":
        return "Пункт"
    return field


def build_change_emoji(field: str, new_value: str, checklist_key: str) -> str:
    if field == "status":
        return status_emoji(new_value)
    if field in {"date", "plan", "plannedDate", "fact"}:
        return "📆"
    if field == "extraInfo":
        return "📝"
    if field == "source":
        return "📚"
    if field == "name":
        return "✏️"
    if field == "document":
        return "📎"
    if field == "add-item":
        return "➕"
    return "✏️"


def should_show_empty_old_value(field: str) -> bool:
    return field in {"extraInfo", "source", "name"}


def build_change_entry(change: dict, field: str, checklist_key: str) -> dict:
    item_name = clean_cell_value(change.get("itemName")) or "Без названия"
    old_value = normalize_message_value(field, change.get("oldValue"), checklist_key)
    new_value = normalize_message_value(field, change.get("newValue"), checklist_key) or "—"
    label = build_change_left_label(field, checklist_key)
    prefix = build_change_emoji(field, new_value, checklist_key)

    if field == "add-item":
        left = f"{prefix}{item_name} / {label}: →"
        right = "добавлен"
        return {
            "field": field,
            "itemName": item_name,
            "oldValue": old_value,
            "newValue": new_value,
            "leftText": left,
            "rightText": right,
        }

    if old_value:
        left = f"{prefix}{item_name} / {label}: {old_value} →"
    elif should_show_empty_old_value(field):
        left = f"{prefix}{item_name} / {label}: — →"
    else:
        left = f"{prefix}{item_name} / {label}: →"

    return {
        "field": field,
        "itemName": item_name,
        "oldValue": old_value,
        "newValue": new_value,
        "leftText": left,
        "rightText": new_value,
    }


def collect_global_alignment_width(section_blocks: list[dict]) -> int:
    widths = []
    for block in section_blocks:
        for section in block.get("sections") or []:
            for row in section.get("rows") or []:
                widths.append(message_visible_length(row.get("leftText") or ""))

    return max(widths) if widths else 0


def format_aligned_change_line(row: dict, target_width: int) -> str:
    left = str(row.get("leftText") or "")
    right = str(row.get("rightText") or "—")
    return f"{pad_message_left(left, target_width)}|[b]{right}[/b]"


def build_aligned_section_lines(section_title: str, rows: list[dict], target_width: int) -> list[str]:
    if not rows:
        return []

    lines = [f"[b]{section_title}:[/b]", ""]
    for row in rows:
        lines.append(format_aligned_change_line(row, target_width))
    lines.append("")
    return lines


def build_recent_changes_sections(changes: list, checklist_key: str) -> list[dict]:
    grouped = split_changes(changes)
    key = normalize_checklist_key(checklist_key)
    section_order = ["status", "date"]
    if key == "concept":
        section_order.extend(["source", "extraInfo", "name", "add-item"])
    elif key == "id":
        section_order.extend(["document", "add-item"])
    elif key == "opr":
        section_order.extend(["add-item"])

    sections = []
    for field in section_order:
        field_changes = grouped.get(field) or []
        if not field_changes:
            continue
        rows = [build_change_entry(change, field, key) for change in field_changes]
        sections.append({
            "field": field,
            "title": MESSAGE_SECTION_LABELS.get(field, field),
            "rows": rows,
        })
    return sections


def build_recent_changes_text(
    changes: list,
    checklist_title: str = "Чек-лист ИД",
    checklist_key: str = "id",
    sections: list[dict] | None = None,
    alignment_width: int = 0,
) -> str:
    title = get_checklist_message_title(checklist_key, checklist_title).upper()
    sections = sections if sections is not None else build_recent_changes_sections(changes, checklist_key)
    lines = [f"✏️[b]ИЗМЕНЕНИЯ В {title}[/b]", ""]

    if sections:
        target_width = alignment_width or collect_global_alignment_width([{"sections": sections}])
        for section in sections:
            lines.extend(build_aligned_section_lines(section["title"], section["rows"], target_width))
    else:
        lines.extend(["Изменений нет", ""])

    return "\n".join(lines).strip()


def build_checklist_message_block(
    data: dict,
    changes: list,
    sections: list[dict] | None = None,
    alignment_width: int = 0,
) -> str:
    checklist_key = normalize_checklist_key(data.get("checklistKey") or "id")
    checklist_title = get_checklist_message_title(checklist_key, data.get("title") or "")
    dialog_id = normalize_dialog_id(data.get("resolvedDialogId") or data.get("dialogId") or "")
    link_url = build_checklist_message_link(dialog_id, checklist_key)
    link_caption = get_checklist_link_caption(checklist_key)

    parts = [
        build_recent_changes_text(
            changes,
            checklist_title,
            checklist_key,
            sections=sections,
            alignment_width=alignment_width,
        ),
        "",
        build_progress_text(data),
        "",
    ]

    if link_url:
        parts.append(f"[URL={link_url}]Чтобы посмотреть весь чек-лист {link_caption} нажмите на этот текст[/URL]")
    else:
        parts.append(f"Чтобы посмотреть весь чек-лист {link_caption} нажмите на этот текст")

    return "\n".join(parts).strip()


def build_multi_checklist_chat_message(sessions: list, editor: dict) -> str:
    if not sessions:
        return ""

    indexed = []
    for session in sessions:
        data = session.get("data") or {}
        changes = session.get("changes") or []
        checklist_key = normalize_checklist_key(session.get("checklistKey") or data.get("checklistKey") or "id")
        if not changes:
            continue
        indexed.append({
            "checklistKey": checklist_key,
            "data": data,
            "changes": changes,
            "sections": build_recent_changes_sections(changes, checklist_key),
        })

    if not indexed:
        return ""

    order_map = {key: index for index, key in enumerate(MESSAGE_CHECKLIST_ORDER)}
    indexed.sort(key=lambda item: order_map.get(item["checklistKey"], 999))
    alignment_width = collect_global_alignment_width(indexed)

    parts = []
    for entry in indexed:
        block = build_checklist_message_block(
            entry["data"],
            entry["changes"],
            sections=entry["sections"],
            alignment_width=alignment_width,
        )
        if block:
            parts.append(block)

    if not parts:
        return ""

    parts.extend([
        "",
        build_editor_text(editor),
    ])
    return "\n\n".join(parts).strip()


def build_checklist_chat_message(data: dict, changes: list, editor: dict) -> str:
    checklist_key = normalize_checklist_key(data.get("checklistKey") or "id")
    sections = build_recent_changes_sections(changes, checklist_key)
    alignment_width = collect_global_alignment_width([{"sections": sections}])
    return build_multi_checklist_chat_message([
        {
            "checklistKey": checklist_key,
            "data": data,
            "changes": changes,
            "sections": sections,
            "alignmentWidth": alignment_width,
        }
    ], editor)


def install_finish_block():
    return """
    <script src="https://api.bitrix24.com/api/v1/"></script>
    <script>
        if (typeof BX24 !== 'undefined') {
            BX24.init(function () {
                try {
                    BX24.installFinish();
                } catch (e) {
                    console.log(e);
                }
            });
        }
    </script>
    """


def format_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def parse_xlsx_to_checklist(file_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    items = []

    for row in range(3, ws.max_row + 1):
        name = clean_cell_value(ws[f"A{row}"].value)
        status = normalize_status(ws[f"B{row}"].value)
        plan = normalize_date_string(format_cell(ws[f"C{row}"].value))
        fact = normalize_date_string(format_cell(ws[f"D{row}"].value))

        if not name:
            continue

        group_id = resolve_group_id(name)

        items.append({
            "id": build_item_id(group_id, len([x for x in items if x["group"] == group_id]) + 1),
            "group": group_id,
            "order": len([x for x in items if x["group"] == group_id]) + 1,
            "name": name,
            "priority": derive_indicator_from_status(status),
            "status": status,
            "plan": plan,
            "fact": fact,
            "documentUrl": "",
            "documentName": "",
            "isCustom": False,
        })

    data = {
        "title": "Чек-лист ИД",
        "collabTitle": ws.title,
        "contractDeadline": "",
        "startDate": "",
        "groups": build_default_groups(),
        "items": items,
    }

    return normalize_checklist_data(data)


def normalize_dialog_id(value: str) -> str:
    s = str(value or "").strip()
    s = s.strip('"').strip("'")

    if not s:
        return ""

    if s.startswith("chat") and s[4:].isdigit():
        return s

    if s.isdigit():
        return f"chat{s}"

    return s


def normalize_checklist_key(value: str) -> str:
    value = str(value or "").strip().lower()
    if value in {"id", "opr", "concept"}:
        return value
    return "id"


def make_storage_dialog_id(dialog_id: str, checklist_key: str = "id") -> str:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    return dialog_id if checklist_key == "id" else f"{dialog_id}::{checklist_key}"


def save_checklist(dialog_id: str, data: dict, checklist_key: str = "id"):
    dialog_id = make_storage_dialog_id(dialog_id, checklist_key)
    checklist_key = normalize_checklist_key(checklist_key)
    data = normalize_checklist_data(data, checklist_key)

    conn = get_conn()
    conn.execute("""
        INSERT INTO checklists(dialog_id, title, data_json)
        VALUES (?, ?, ?)
        ON CONFLICT(dialog_id) DO UPDATE SET
            title=excluded.title,
            data_json=excluded.data_json
    """, (dialog_id, data.get("title", "Чек-лист"), json.dumps(data, ensure_ascii=False)))
    conn.commit()
    conn.close()

def save_project_storage_context(dialog_id: str, payload: dict):
    dialog_id = normalize_dialog_id(dialog_id)
    if not dialog_id:
        raise ValueError("dialogId is required")

    project_id = str(payload.get("projectId") or "").strip()
    project_name = clean_cell_value(payload.get("projectName"))
    storage_mode = payload.get("storageMode") or {}
    yandex_disk = payload.get("yandexDisk") or {}
    item_mappings = payload.get("itemMappings") or []

    conn = get_conn()
    conn.execute("""
        INSERT INTO project_storage_contexts(
            dialog_id,
            project_id,
            project_name,
            provider,
            storage_mode_json,
            yandex_json,
            item_mappings_json,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(dialog_id) DO UPDATE SET
            project_id=excluded.project_id,
            project_name=excluded.project_name,
            provider=excluded.provider,
            storage_mode_json=excluded.storage_mode_json,
            yandex_json=excluded.yandex_json,
            item_mappings_json=excluded.item_mappings_json,
            updated_at=excluded.updated_at
    """, (
        dialog_id,
        project_id,
        project_name,
        str(yandex_disk.get("provider") or "yandex_disk").strip(),
        json.dumps(storage_mode, ensure_ascii=False),
        json.dumps(yandex_disk, ensure_ascii=False),
        json.dumps(item_mappings, ensure_ascii=False),
        datetime.now().isoformat()
    ))
    conn.commit()
    conn.close()


def get_project_storage_context(dialog_id: str):
    dialog_id = normalize_dialog_id(dialog_id)
    if not dialog_id:
        return None

    conn = get_conn()
    row = conn.execute(
        "SELECT * FROM project_storage_contexts WHERE dialog_id = ?",
        (dialog_id,)
    ).fetchone()
    conn.close()

    if not row:
        return None

    return {
        "dialogId": row["dialog_id"],
        "projectId": row["project_id"],
        "projectName": row["project_name"],
        "provider": row["provider"],
        "storageMode": json.loads(row["storage_mode_json"] or "{}"),
        "yandexDisk": json.loads(row["yandex_json"] or "{}"),
        "itemMappings": json.loads(row["item_mappings_json"] or "[]"),
        "updatedAt": row["updated_at"],
    }


def get_item_yandex_mapping(dialog_id: str, checklist_key: str, item_name: str):
    context = get_project_storage_context(dialog_id)
    if not context:
        return None

    checklist_key = normalize_checklist_key(checklist_key)
    item_name = clean_cell_value(item_name).lower()

    for mapping in context.get("itemMappings", []):
        mapping_key = normalize_checklist_key(mapping.get("checklistKey"))
        mapping_name = clean_cell_value(mapping.get("itemName")).lower()

        if mapping_key == checklist_key and mapping_name == item_name:
            return mapping

    return None


def get_item_yandex_folder(dialog_id: str, checklist_key: str, item_name: str):
    context = get_project_storage_context(dialog_id)
    if not context:
        return None

    mapping = get_item_yandex_mapping(dialog_id, checklist_key, item_name)
    if not mapping:
        return None

    folder_alias = str(mapping.get("folderAlias") or "").strip()
    if not folder_alias:
        return None

    yandex_disk = context.get("yandexDisk") or {}
    folders = yandex_disk.get("folders") or {}
    folder = folders.get(folder_alias)

    if not folder:
        return None

    return {
        "folderAlias": folder_alias,
        "folder": folder,
        "mapping": mapping,
        "context": context,
    }

def is_yandex_disk_enabled() -> bool:
    return bool(YANDEX_DISK_OAUTH_TOKEN)


def get_yandex_disk_headers() -> dict:
    return {
        "Authorization": f"OAuth {YANDEX_DISK_OAUTH_TOKEN}"
    }


def normalize_yandex_disk_path(path: str) -> str:
    value = clean_cell_value(path)
    if not value:
        return ""

    if value.startswith("disk:/"):
        return value

    if value.startswith("/"):
        return "disk:" + value

    return "disk:/" + value


def yandex_disk_get_upload_href(target_path: str, overwrite: bool = True) -> str:
    normalized_path = normalize_yandex_disk_path(target_path)
    response = requests.get(
        f"{YANDEX_DISK_API_BASE}/resources/upload",
        headers=get_yandex_disk_headers(),
        params={
            "path": normalized_path,
            "overwrite": "true" if overwrite else "false",
        },
        timeout=30
    )
    response.raise_for_status()
    data = response.json() or {}
    href = clean_cell_value(data.get("href"))
    if not href:
        raise RuntimeError("Yandex Disk upload href not found")
    return href


def yandex_disk_upload_bytes(target_path: str, file_bytes: bytes) -> dict:
    upload_href = yandex_disk_get_upload_href(target_path, overwrite=True)
    upload_response = requests.put(upload_href, data=file_bytes, timeout=120)
    upload_response.raise_for_status()

    return {
        "ok": True,
        "path": normalize_yandex_disk_path(target_path),
    }


def yandex_disk_delete_path(target_path: str, permanently: bool = True):
    normalized_path = normalize_yandex_disk_path(target_path)
    response = requests.delete(
        f"{YANDEX_DISK_API_BASE}/resources",
        headers=get_yandex_disk_headers(),
        params={
            "path": normalized_path,
            "permanently": "true" if permanently else "false"
        },
        timeout=30
    )

    if response.status_code not in (200, 202, 204):
        try:
            payload = response.json()
        except Exception:
            payload = {"text": response.text}
        raise RuntimeError(f"Yandex Disk delete failed: {payload}")

def sanitize_yandex_folder_name(value: str) -> str:
    value = clean_cell_value(value)
    if not value:
        return "Новый пункт"

    value = re.sub(r'[\\\\/:*?"<>|]+', " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value or "Новый пункт"


def yandex_disk_ensure_folder(target_path: str) -> dict:
    normalized_path = normalize_yandex_disk_path(target_path)

    response = requests.put(
        f"{YANDEX_DISK_API_BASE}/resources",
        headers=get_yandex_disk_headers(),
        params={"path": normalized_path},
        timeout=30
    )

    if response.status_code not in (201, 409):
        try:
            payload = response.json()
        except Exception:
            payload = {"text": response.text}
        raise RuntimeError(f"Yandex Disk create folder failed: {payload}")

    return {
        "ok": True,
        "path": normalized_path,
        "alreadyExists": response.status_code == 409,
    }


def yandex_disk_publish_path(target_path: str) -> dict:
    normalized_path = normalize_yandex_disk_path(target_path)

    response = requests.put(
        f"{YANDEX_DISK_API_BASE}/resources/publish",
        headers=get_yandex_disk_headers(),
        params={"path": normalized_path},
        timeout=30
    )

    if response.status_code not in (200, 201, 202):
        try:
            payload = response.json()
        except Exception:
            payload = {"text": response.text}
        raise RuntimeError(f"Yandex Disk publish failed: {payload}")

    return {
        "ok": True,
        "path": normalized_path,
    }


def yandex_disk_get_resource_meta(target_path: str) -> dict:
    normalized_path = normalize_yandex_disk_path(target_path)

    response = requests.get(
        f"{YANDEX_DISK_API_BASE}/resources",
        headers=get_yandex_disk_headers(),
        params={
            "path": normalized_path,
            "fields": "name,path,public_url"
        },
        timeout=30
    )
    response.raise_for_status()

    data = response.json() or {}
    return {
        "name": clean_cell_value(data.get("name")),
        "path": clean_cell_value(data.get("path")) or normalized_path,
        "public_url": clean_cell_value(data.get("public_url")),
    }


def extract_yandex_folder_number(name: str) -> int:
    name = clean_cell_value(name)
    match = re.match(r"^(\d+)_", name)
    if not match:
        return 0
    try:
        return int(match.group(1))
    except Exception:
        return 0


def get_next_id_stage_folder_number(yandex_disk: dict) -> int:
    folders = dict((yandex_disk or {}).get("folders") or {})
    max_number = 0

    for folder in folders.values():
        folder_name = clean_cell_value((folder or {}).get("name"))
        current_number = extract_yandex_folder_number(folder_name)
        if current_number > max_number:
            max_number = current_number

    return max_number + 1 if max_number > 0 else 9


def build_custom_id_yandex_folder_paths(id_stage_root_path: str, yandex_disk: dict, item_name: str) -> dict:
    base_path = normalize_yandex_disk_path(id_stage_root_path).rstrip("/")
    next_number = get_next_id_stage_folder_number(yandex_disk)
    folder_name = f"{next_number:02d}_{sanitize_yandex_folder_name(item_name)}"
    item_folder_path = f"{base_path}/{folder_name}"

    return {
        "itemPath": item_folder_path,
        "folderName": folder_name,
        "folderNumber": next_number,
    }


def upsert_item_yandex_mapping(
    dialog_id: str,
    checklist_key: str,
    item_name: str,
    folder_alias: str,
    folder_name: str,
    folder_path: str,
    folder_url: str,
):
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    item_name = clean_cell_value(item_name)
    folder_alias = clean_cell_value(folder_alias)
    folder_name = clean_cell_value(folder_name)
    folder_path = normalize_yandex_disk_path(folder_path)
    folder_url = clean_cell_value(folder_url)

    if not dialog_id:
        raise ValueError("dialogId is required")
    if not item_name:
        raise ValueError("itemName is required")
    if not folder_alias:
        raise ValueError("folderAlias is required")
    if not folder_path:
        raise ValueError("folderPath is required")

    context = get_project_storage_context(dialog_id)
    if not context:
        raise RuntimeError("project storage context not found")

    yandex_disk = dict(context.get("yandexDisk") or {})
    folders = dict(yandex_disk.get("folders") or {})
    item_mappings = list(context.get("itemMappings") or [])

    folders[folder_alias] = {
        "name": folder_name,
        "path": folder_path,
        "url": folder_url,
    }
    yandex_disk["folders"] = folders

    normalized_item_name = item_name.lower()
    filtered_mappings = []

    for mapping in item_mappings:
        mapping_key = normalize_checklist_key(mapping.get("checklistKey"))
        mapping_name = clean_cell_value(mapping.get("itemName")).lower()

        if mapping_key == checklist_key and mapping_name == normalized_item_name:
            continue

        filtered_mappings.append(mapping)

    filtered_mappings.append({
        "checklistKey": checklist_key,
        "itemName": item_name,
        "folderAlias": folder_alias,
    })

    payload = {
        "dialogId": dialog_id,
        "projectId": context.get("projectId") or "",
        "projectName": context.get("projectName") or "",
        "storageMode": context.get("storageMode") or {},
        "yandexDisk": yandex_disk,
        "itemMappings": filtered_mappings,
    }

    save_project_storage_context(dialog_id, payload)
    return payload


def ensure_yandex_folder_for_custom_item(
    dialog_id: str,
    checklist_key: str,
    group_id: int,
    item_name: str,
    item_id: str,
) -> dict:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    item_name = clean_cell_value(item_name)
    item_id = str(item_id or "").strip()

    if checklist_key != "id":
        raise RuntimeError("custom yandex folder is supported only for checklist id")

    if not is_yandex_disk_enabled():
        raise RuntimeError("yandex token is empty")

    context = get_project_storage_context(dialog_id)
    if not context:
        raise RuntimeError("project storage context not found")

    yandex_disk = context.get("yandexDisk") or {}
    id_stage_root_path = clean_cell_value(yandex_disk.get("idStageRootPath"))
    if not id_stage_root_path:
        raise RuntimeError("idStageRootPath is empty in project storage context")

    paths = build_custom_id_yandex_folder_paths(id_stage_root_path, yandex_disk, item_name)

    yandex_disk_ensure_folder(paths["itemPath"])
    yandex_disk_publish_path(paths["itemPath"])

    meta = yandex_disk_get_resource_meta(paths["itemPath"])

    suffix = re.sub(r"[^a-zA-Z0-9]+", "", item_id)[-8:].lower() or uuid.uuid4().hex[:8]
    folder_alias = f"id_custom_g{int(group_id)}_{slugify_folder_part(item_name)}_{suffix}"

    folder_name = clean_cell_value(meta.get("name")) or paths["folderName"]
    folder_path = clean_cell_value(meta.get("path")) or paths["itemPath"]
    folder_url = clean_cell_value(meta.get("public_url"))

    upsert_item_yandex_mapping(
        dialog_id=dialog_id,
        checklist_key=checklist_key,
        item_name=item_name,
        folder_alias=folder_alias,
        folder_name=folder_name,
        folder_path=folder_path,
        folder_url=folder_url,
    )

    return {
        "folderAlias": folder_alias,
        "folderName": folder_name,
        "folderPath": folder_path,
        "folderUrl": folder_url,
    }

def get_standard_id_yandex_spec(item_name: str):
    item_name = clean_cell_value(item_name)
    return STANDARD_ID_YANDEX_FOLDER_SPECS.get(item_name)


def get_id_stage_root_path_from_context(dialog_id: str) -> str:
    context = get_project_storage_context(dialog_id)
    if not context:
        return ""

    yandex_disk = context.get("yandexDisk") or {}
    return clean_cell_value(yandex_disk.get("idStageRootPath"))


def build_standard_id_yandex_folder_path(id_stage_root_path: str, relative_path: str) -> str:
    base_path = normalize_yandex_disk_path(id_stage_root_path).rstrip("/")
    relative_path = str(relative_path or "").strip().strip("/")
    if not base_path or not relative_path:
        return ""
    return f"{base_path}/{relative_path}"


def ensure_standard_yandex_folder_for_item(
    dialog_id: str,
    checklist_key: str,
    item_name: str,
) -> dict:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    item_name = clean_cell_value(item_name)

    if checklist_key != "id":
        raise RuntimeError("standard yandex folder restore is supported only for checklist id")

    if not is_yandex_disk_enabled():
        raise RuntimeError("yandex token is empty")

    spec = get_standard_id_yandex_spec(item_name)
    if not spec:
        raise RuntimeError("standard folder spec not found")

    id_stage_root_path = get_id_stage_root_path_from_context(dialog_id)
    if not id_stage_root_path:
        raise RuntimeError("idStageRootPath is empty in project storage context")

    full_folder_path = build_standard_id_yandex_folder_path(
        id_stage_root_path,
        spec.get("relativePath") or ""
    )
    if not full_folder_path:
        raise RuntimeError("full standard folder path is empty")

    path_parts = [part for part in str(spec.get("relativePath") or "").split("/") if part]
    current_path = normalize_yandex_disk_path(id_stage_root_path).rstrip("/")

    for part in path_parts:
        current_path = f"{current_path}/{part}"
        yandex_disk_ensure_folder(current_path)

    yandex_disk_publish_path(full_folder_path)
    meta = yandex_disk_get_resource_meta(full_folder_path)

    folder_alias = clean_cell_value(spec.get("alias"))
    folder_name = clean_cell_value(meta.get("name")) or clean_cell_value(spec.get("folderName"))
    folder_path = clean_cell_value(meta.get("path")) or full_folder_path
    folder_url = clean_cell_value(meta.get("public_url"))

    upsert_item_yandex_mapping(
        dialog_id=dialog_id,
        checklist_key=checklist_key,
        item_name=item_name,
        folder_alias=folder_alias,
        folder_name=folder_name,
        folder_path=folder_path,
        folder_url=folder_url,
    )

    return {
        "folderAlias": folder_alias,
        "folderName": folder_name,
        "folderPath": folder_path,
        "folderUrl": folder_url,
    }

def get_standard_opr_yandex_spec(item_name: str):
    item_name = clean_cell_value(item_name)
    return STANDARD_OPR_YANDEX_FOLDER_SPECS.get(item_name)


def get_opr_root_path_from_context(dialog_id: str) -> str:
    context = get_project_storage_context(dialog_id)
    if not context:
        return ""

    yandex_disk = context.get("yandexDisk") or {}
    project_root_path = clean_cell_value(yandex_disk.get("projectRootPath"))
    if not project_root_path:
        return ""

    return normalize_yandex_disk_path(project_root_path).rstrip("/") + "/02_Выдача документации/02_ОПР"


def ensure_opr_root_path(dialog_id: str) -> str:
    context = get_project_storage_context(dialog_id)
    if not context:
        raise RuntimeError("project storage context not found")

    yandex_disk = context.get("yandexDisk") or {}
    project_root_path = clean_cell_value(yandex_disk.get("projectRootPath"))
    if not project_root_path:
        raise RuntimeError("projectRootPath is empty in project storage context")

    current_path = normalize_yandex_disk_path(project_root_path).rstrip("/")
    for part in ["02_Выдача документации", "02_ОПР"]:
        current_path = f"{current_path}/{part}"
        yandex_disk_ensure_folder(current_path)

    return current_path


def build_standard_opr_yandex_folder_path(opr_root_path: str, relative_path: str) -> str:
    base_path = normalize_yandex_disk_path(opr_root_path).rstrip("/")
    relative_path = str(relative_path or "").strip().strip("/")
    if not base_path or not relative_path:
        return ""
    return f"{base_path}/{relative_path}"


def ensure_standard_yandex_folder_for_opr_item(
    dialog_id: str,
    checklist_key: str,
    item_name: str,
) -> dict:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    item_name = clean_cell_value(item_name)

    if checklist_key != "opr":
        raise RuntimeError("standard yandex folder restore is supported only for checklist opr")

    if not is_yandex_disk_enabled():
        raise RuntimeError("yandex token is empty")

    spec = get_standard_opr_yandex_spec(item_name)
    if not spec:
        raise RuntimeError("standard OPR folder spec not found")

    opr_root_path = ensure_opr_root_path(dialog_id)
    full_folder_path = build_standard_opr_yandex_folder_path(
        opr_root_path,
        spec.get("relativePath") or ""
    )
    if not full_folder_path:
        raise RuntimeError("full OPR folder path is empty")

    path_parts = [part for part in str(spec.get("relativePath") or "").split("/") if part]
    current_path = normalize_yandex_disk_path(opr_root_path).rstrip("/")

    for part in path_parts:
        current_path = f"{current_path}/{part}"
        yandex_disk_ensure_folder(current_path)

    yandex_disk_publish_path(full_folder_path)
    meta = yandex_disk_get_resource_meta(full_folder_path)

    folder_alias = clean_cell_value(spec.get("alias"))
    folder_name = clean_cell_value(meta.get("name")) or clean_cell_value(spec.get("folderName"))
    folder_path = clean_cell_value(meta.get("path")) or full_folder_path
    folder_url = clean_cell_value(meta.get("public_url"))

    upsert_item_yandex_mapping(
        dialog_id=dialog_id,
        checklist_key=checklist_key,
        item_name=item_name,
        folder_alias=folder_alias,
        folder_name=folder_name,
        folder_path=folder_path,
        folder_url=folder_url,
    )

    return {
        "folderAlias": folder_alias,
        "folderName": folder_name,
        "folderPath": folder_path,
        "folderUrl": folder_url,
    }


def get_next_opr_root_folder_number(yandex_disk: dict, opr_root_path: str) -> int:
    opr_root_path = normalize_yandex_disk_path(opr_root_path).rstrip("/")
    folders = dict((yandex_disk or {}).get("folders") or {})
    max_number = 0

    for folder in folders.values():
        folder_path = normalize_yandex_disk_path((folder or {}).get("path"))
        if not folder_path.startswith(opr_root_path + "/"):
            continue

        relative = folder_path[len(opr_root_path) + 1:]
        first_part = relative.split("/", 1)[0]
        current_number = extract_yandex_folder_number(first_part)

        if current_number > max_number:
            max_number = current_number

    return max_number + 1 if max_number > 0 else 4


def build_custom_opr_yandex_folder_paths(opr_root_path: str, yandex_disk: dict, item_name: str) -> dict:
    base_path = normalize_yandex_disk_path(opr_root_path).rstrip("/")
    next_number = get_next_opr_root_folder_number(yandex_disk, base_path)
    folder_name = f"{next_number:02d}_{sanitize_yandex_folder_name(item_name)}"
    item_folder_path = f"{base_path}/{folder_name}"

    return {
        "itemPath": item_folder_path,
        "folderName": folder_name,
        "folderNumber": next_number,
    }


def ensure_yandex_folder_for_custom_opr_item(
    dialog_id: str,
    checklist_key: str,
    group_id: int,
    item_name: str,
    item_id: str,
) -> dict:
    dialog_id = normalize_dialog_id(dialog_id)
    checklist_key = normalize_checklist_key(checklist_key)
    item_name = clean_cell_value(item_name)
    item_id = str(item_id or "").strip()

    if checklist_key != "opr":
        raise RuntimeError("custom yandex folder is supported only for checklist opr")

    if not is_yandex_disk_enabled():
        raise RuntimeError("yandex token is empty")

    context = get_project_storage_context(dialog_id)
    if not context:
        raise RuntimeError("project storage context not found")

    yandex_disk = context.get("yandexDisk") or {}
    opr_root_path = ensure_opr_root_path(dialog_id)
    paths = build_custom_opr_yandex_folder_paths(opr_root_path, yandex_disk, item_name)

    yandex_disk_ensure_folder(paths["itemPath"])
    yandex_disk_publish_path(paths["itemPath"])

    meta = yandex_disk_get_resource_meta(paths["itemPath"])

    suffix = re.sub(r"[^a-zA-Z0-9]+", "", item_id)[-8:].lower() or uuid.uuid4().hex[:8]
    folder_alias = f"opr_custom_g{int(group_id)}_{slugify_folder_part(item_name)}_{suffix}"

    folder_name = clean_cell_value(meta.get("name")) or paths["folderName"]
    folder_path = clean_cell_value(meta.get("path")) or paths["itemPath"]
    folder_url = clean_cell_value(meta.get("public_url"))

    upsert_item_yandex_mapping(
        dialog_id=dialog_id,
        checklist_key=checklist_key,
        item_name=item_name,
        folder_alias=folder_alias,
        folder_name=folder_name,
        folder_path=folder_path,
        folder_url=folder_url,
    )

    return {
        "folderAlias": folder_alias,
        "folderName": folder_name,
        "folderPath": folder_path,
        "folderUrl": folder_url,
    }

def ensure_item_yandex_folder_for_upload(
    dialog_id: str,
    checklist_key: str,
    item_name: str,
    item_id: str = "",
    item_group: int = 0,
    is_custom: bool = False,
) -> dict:
    existing = get_item_yandex_folder(dialog_id, checklist_key, item_name)
    if existing:
        folder = existing.get("folder") or {}
        folder_path = clean_cell_value(folder.get("path"))
        if folder_path:
            try:
                yandex_disk_ensure_folder(folder_path)
                yandex_disk_publish_path(folder_path)
                meta = yandex_disk_get_resource_meta(folder_path)

                folder_alias = clean_cell_value(existing.get("folderAlias"))
                upsert_item_yandex_mapping(
                    dialog_id=dialog_id,
                    checklist_key=checklist_key,
                    item_name=item_name,
                    folder_alias=folder_alias,
                    folder_name=clean_cell_value(meta.get("name")) or clean_cell_value(folder.get("name")),
                    folder_path=clean_cell_value(meta.get("path")) or folder_path,
                    folder_url=clean_cell_value(meta.get("public_url")),
                )

                return get_item_yandex_folder(dialog_id, checklist_key, item_name) or existing
            except Exception:
                pass

    if checklist_key == "id" and not is_custom:
        restored = ensure_standard_yandex_folder_for_item(dialog_id, checklist_key, item_name)
        return get_item_yandex_folder(dialog_id, checklist_key, item_name) or {
            "folderAlias": restored.get("folderAlias"),
            "folder": {
                "name": restored.get("folderName"),
                "path": restored.get("folderPath"),
                "url": restored.get("folderUrl"),
            },
            "mapping": {
                "checklistKey": checklist_key,
                "itemName": item_name,
                "folderAlias": restored.get("folderAlias"),
            },
            "context": get_project_storage_context(dialog_id),
        }

    if checklist_key == "id" and is_custom:
        restored = ensure_yandex_folder_for_custom_item(
            dialog_id=dialog_id,
            checklist_key=checklist_key,
            group_id=int(item_group or 0),
            item_name=item_name,
            item_id=item_id,
        )
        return get_item_yandex_folder(dialog_id, checklist_key, item_name) or {
            "folderAlias": restored.get("folderAlias"),
            "folder": {
                "name": restored.get("folderName"),
                "path": restored.get("folderPath"),
                "url": restored.get("folderUrl"),
            },
            "mapping": {
                "checklistKey": checklist_key,
                "itemName": item_name,
                "folderAlias": restored.get("folderAlias"),
            },
            "context": get_project_storage_context(dialog_id),
        }

    if checklist_key == "opr" and not is_custom:
        restored = ensure_standard_yandex_folder_for_opr_item(dialog_id, checklist_key, item_name)
        return get_item_yandex_folder(dialog_id, checklist_key, item_name) or {
            "folderAlias": restored.get("folderAlias"),
            "folder": {
                "name": restored.get("folderName"),
                "path": restored.get("folderPath"),
                "url": restored.get("folderUrl"),
            },
            "mapping": {
                "checklistKey": checklist_key,
                "itemName": item_name,
                "folderAlias": restored.get("folderAlias"),
            },
            "context": get_project_storage_context(dialog_id),
        }

    if checklist_key == "opr" and is_custom:
        restored = ensure_yandex_folder_for_custom_opr_item(
            dialog_id=dialog_id,
            checklist_key=checklist_key,
            group_id=int(item_group or 0),
            item_name=item_name,
            item_id=item_id,
        )
        return get_item_yandex_folder(dialog_id, checklist_key, item_name) or {
            "folderAlias": restored.get("folderAlias"),
            "folder": {
                "name": restored.get("folderName"),
                "path": restored.get("folderPath"),
                "url": restored.get("folderUrl"),
            },
            "mapping": {
                "checklistKey": checklist_key,
                "itemName": item_name,
                "folderAlias": restored.get("folderAlias"),
            },
            "context": get_project_storage_context(dialog_id),
        }

    return None

def build_yandex_file_target_path(folder_path: str, filename: str) -> str:
    folder_path = normalize_yandex_disk_path(folder_path).rstrip("/")
    safe_name = Path(filename or "file.bin").name
    return f"{folder_path}/{safe_name}"


def mirror_document_to_yandex(
    dialog_id: str,
    checklist_key: str,
    item_name: str,
    filename: str,
    file_bytes: bytes,
    item_id: str = "",
    item_group: int = 0,
    is_custom: bool = False,
) -> dict:
    if not is_yandex_disk_enabled():
        return {
            "ok": False,
            "reason": "yandex token is empty"
        }

    folder_data = ensure_item_yandex_folder_for_upload(
        dialog_id=dialog_id,
        checklist_key=checklist_key,
        item_name=item_name,
        item_id=item_id,
        item_group=item_group,
        is_custom=is_custom,
    )
    if not folder_data:
        return {
            "ok": False,
            "reason": "folder mapping not found"
        }

    folder = folder_data.get("folder") or {}
    folder_path = clean_cell_value(folder.get("path"))
    folder_url = clean_cell_value(folder.get("url"))
    folder_alias = clean_cell_value(folder_data.get("folderAlias"))

    if not folder_path:
        return {
            "ok": False,
            "reason": "folder path is empty"
        }

    target_path = build_yandex_file_target_path(folder_path, filename)
    upload_result = yandex_disk_upload_bytes(target_path, file_bytes)

    return {
        "ok": True,
        "folderAlias": folder_alias,
        "folderPath": normalize_yandex_disk_path(folder_path),
        "folderUrl": folder_url,
        "filePath": upload_result.get("path") or normalize_yandex_disk_path(target_path),
    }

def extract_dialog_id_from_form(form: dict) -> str:
    def pick(value) -> str:
        value = normalize_dialog_id(value)
        return value or ""

    direct_candidates = [
        form.get("dialogId"),
        form.get("DIALOG_ID"),
        form.get("DIALOGID"),
        form.get("dialog_id"),
        form.get("chatId"),
        form.get("CHAT_ID"),
        form.get("chat_id"),
    ]

    for value in direct_candidates:
        found = pick(value)
        if found:
            return found

    def walk(obj) -> str:
        if isinstance(obj, dict):
            preferred_keys = [
                "dialogId", "DIALOG_ID", "dialog_id",
                "chatId", "CHAT_ID", "chat_id"
            ]
            for key in preferred_keys:
                found = pick(obj.get(key))
                if found:
                    return found

            for value in obj.values():
                found = walk(value)
                if found:
                    return found

        elif isinstance(obj, list):
            for value in obj:
                found = walk(value)
                if found:
                    return found

        return ""

    json_candidates = [
        form.get("PLACEMENT_OPTIONS"),
        form.get("placementOptions"),
        form.get("options"),
    ]

    for raw in json_candidates:
        if not raw:
            continue
        try:
            data = json.loads(raw) if isinstance(raw, str) else raw
            found = walk(data)
            if found:
                return found
        except Exception:
            pass

    for key, value in form.items():
        key_l = str(key).lower()
        if "dialog" in key_l or "chat" in key_l:
            found = pick(value)
            if found:
                return found

    return ""

def get_checklist(dialog_id: str, checklist_key: str = "id"):
    checklist_key = normalize_checklist_key(checklist_key)
    normalized_id = normalize_dialog_id(dialog_id)

    aliases = []
    for candidate in [dialog_id, normalized_id]:
        candidate = str(candidate or "").strip()
        if candidate and candidate not in aliases:
            aliases.append(candidate)

    if normalized_id.startswith("chat") and normalized_id[4:].isdigit():
        numeric_id = normalized_id[4:]
        if numeric_id not in aliases:
            aliases.append(numeric_id)

    storage_aliases = [make_storage_dialog_id(alias, checklist_key) for alias in aliases]
    storage_dialog_id = make_storage_dialog_id(normalized_id, checklist_key)

    conn = get_conn()

    row = None
    found_alias = None

    for candidate in storage_aliases:
        try:
            row = conn.execute(
                "SELECT data_json FROM checklists WHERE dialog_id = ?",
                (candidate,)
            ).fetchone()
        except sqlite3.OperationalError as e:
            if "no such table" in str(e).lower():
                conn.close()
                init_db()
                conn = get_conn()
                row = conn.execute(
                    "SELECT data_json FROM checklists WHERE dialog_id = ?",
                    (candidate,)
                ).fetchone()
            else:
                conn.close()
                raise

        if row:
            found_alias = candidate
            break

    conn.close()

    if row:
        data = json.loads(row["data_json"])
        data = normalize_checklist_data(data, checklist_key)
        data["resolvedDialogId"] = normalized_id
        data["lookupAliases"] = aliases
        data["foundAlias"] = found_alias
        return data

    data = build_default_checklist_template(normalized_id, checklist_key)
    save_checklist(normalized_id, data, checklist_key)

    data["resolvedDialogId"] = normalized_id
    data["lookupAliases"] = aliases
    data["foundAlias"] = storage_dialog_id
    return data


def app_home_html(
    initial_dialog_id: str = "",
    initial_checklist_key: str = "id",
    initial_context_text: str = ""
):
    initial_dialog_id_json = json.dumps(
        normalize_dialog_id(initial_dialog_id or ""),
        ensure_ascii=False
    )
    initial_checklist_key_json = json.dumps(
        normalize_checklist_key(initial_checklist_key or "id"),
        ensure_ascii=False
    )
    initial_context_text_json = json.dumps(initial_context_text or "", ensure_ascii=False)

    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>Чек-листы проекта</title>
        <script src="https://api.bitrix24.com/api/v1/"></script>
        <script>
            (function () {{
                const initialDialogId = {initial_dialog_id_json};
                const initialChecklistKey = {initial_checklist_key_json};
                const initialContextText = {initial_context_text_json};

                function detectAppBasePath() {{
                    const path = String(window.location.pathname || '/').replace(/\/+$/, '');
                    const suffixes = ['/launch', '/popup', '/textarea', '/install', '/health', '/debug/logs', '/admin', '/admin/upload'];

                    for (const suffix of suffixes) {{
                        if (path === suffix) return '';
                        if (path.endsWith(suffix)) {{
                            return path.slice(0, -suffix.length) || '';
                        }}
                    }}

                    return path === '/' ? '' : path;
                }}

                const APP_BASE_PATH = detectAppBasePath();

                function appPath(path) {{
                    return (APP_BASE_PATH || '') + '/' + String(path || '').replace(/^\/+/, '');
                }}

                function pickValue(searchParams, hashParams, key, fallback) {{
                    return (searchParams.get(key) || hashParams.get(key) || fallback || '').trim();
                }}

                function normalizeChecklistKey(value) {{
                    const v = String(value || '').trim().toLowerCase();
                    if (v === 'concept' || v === 'opr' || v === 'id') return v;
                    return 'id';
                }}

                function extractFromBx24() {{
                    let dialogId = '';
                    let checklistKey = '';

                    try {{
                        if (!(window.BX24 && typeof window.BX24.placement === 'object' && typeof window.BX24.placement.info === 'function')) {{
                            return {{ dialogId: '', checklistKey: '' }};
                        }}

                        const info = window.BX24.placement.info() || {{}};
                        const options = info.options || {{}};

                        const dialogCandidates = [
                            options.dialogId,
                            options.DIALOG_ID,
                            options.dialog_id,
                            info.dialogId,
                            info.DIALOG_ID,
                            info.dialog_id,
                            options.chatId,
                            options.CHAT_ID,
                            options.chat_id,
                            info.chatId,
                            info.CHAT_ID,
                            info.chat_id
                        ];

                        for (let i = 0; i < dialogCandidates.length; i++) {{
                            const candidate = String(dialogCandidates[i] || '').trim();
                            if (candidate) {{
                                dialogId = candidate;
                                break;
                            }}
                        }}

                        const checklistCandidates = [
                            options.checklistKey,
                            options.CHECKLIST_KEY,
                            options.checklist_key,
                            info.checklistKey,
                            info.CHECKLIST_KEY,
                            info.checklist_key
                        ];

                        for (let i = 0; i < checklistCandidates.length; i++) {{
                            const candidate = String(checklistCandidates[i] || '').trim();
                            if (candidate) {{
                                checklistKey = normalizeChecklistKey(candidate);
                                break;
                            }}
                        }}

                        try {{
                            console.log('app_home placement.info =', info);
                        }} catch (e) {{}}
                    }} catch (e) {{
                        console.log('app_home extractFromBx24 error:', e);
                    }}

                    return {{
                        dialogId: dialogId,
                        checklistKey: checklistKey || 'id'
                    }};
                }}

                function resizeCurrentPopupFrame() {{
                    try {{
                        if (window.BX24 && typeof window.BX24.resizeWindow === 'function') {{
                            window.BX24.resizeWindow(1180, 720);
                        }}
                        if (window.BX24 && typeof window.BX24.fitWindow === 'function') {{
                            window.BX24.fitWindow();
                        }}
                    }} catch (e) {{
                        console.log('BX24 resize skipped:', e);
                    }}
                }}

                function rememberAndRedirect(dialogId, checklistKey) {{
                    if (!dialogId) return;

                    try {{
                        localStorage.setItem('checklist_pending_dialog', JSON.stringify({{
                            dialogId: dialogId,
                            checklistKey: checklistKey || 'id',
                            ts: Date.now()
                        }}));
                    }} catch (e) {{
                        console.log('pending dialog save skipped:', e);
                    }}

                    const popupUrl =
                        appPath('popup') +
                        '?dialogId=' + encodeURIComponent(dialogId) +
                        '&checklistKey=' + encodeURIComponent(checklistKey || 'id');

                    resizeCurrentPopupFrame();
                    setTimeout(resizeCurrentPopupFrame, 80);

                    setTimeout(function () {{
                        window.location.replace(popupUrl);
                    }}, 120);
                }}

                try {{
                    if (initialContextText) {{
                        console.log('HOME initial context:', initialContextText);
                    }}

                    const searchParams = new URLSearchParams(window.location.search || '');
                    const hashRaw = String(window.location.hash || '').replace(/^#/, '');
                    const hashParams = new URLSearchParams(hashRaw);

                    const raw = localStorage.getItem('checklist_pending_dialog');
                    let localPayload = null;

                    if (raw) {{
                        try {{
                            localPayload = JSON.parse(raw);
                        }} catch (e) {{
                            console.log('pending dialog parse skipped:', e);
                        }}
                    }}

                    const dialogId = pickValue(
                        searchParams,
                        hashParams,
                        'dialogId',
                        initialDialogId || (localPayload && localPayload.dialogId)
                    );

                    const checklistKey = normalizeChecklistKey(
                        pickValue(
                            searchParams,
                            hashParams,
                            'checklistKey',
                            initialChecklistKey || (localPayload && localPayload.checklistKey)
                        ) || 'id'
                    );

                    const ts = Number((localPayload && localPayload.ts) || 0);
                    const age = ts ? (Date.now() - ts) : 0;

                    if (dialogId) {{
                        rememberAndRedirect(dialogId, checklistKey);
                        return;
                    }}

                    if (window.BX24 && typeof window.BX24.init === 'function') {{
                        window.BX24.init(function () {{
                            const bxData = extractFromBx24();
                            if (bxData.dialogId) {{
                                rememberAndRedirect(bxData.dialogId, bxData.checklistKey || checklistKey || 'id');
                                return;
                            }}

                            if (localPayload && localPayload.dialogId && age < 60000) {{
                                rememberAndRedirect(
                                    localPayload.dialogId,
                                    normalizeChecklistKey(localPayload.checklistKey || 'id')
                                );
                            }}
                        }});
                        return;
                    }}

                    if (localPayload && localPayload.dialogId && age < 60000) {{
                        rememberAndRedirect(
                            localPayload.dialogId,
                            normalizeChecklistKey(localPayload.checklistKey || 'id')
                        );
                        return;
                    }}
                }} catch (e) {{
                    console.log('launcher redirect skipped:', e);
                }}
            }})();
        </script>
    </head>
    <body style="margin:0;font-family:Arial,sans-serif;background:#f8fafc;color:#344054;display:flex;align-items:center;justify-content:center;min-height:100vh;">
        <div style="padding:24px 28px;border:1px solid #e5e7eb;border-radius:14px;background:#fff;box-shadow:0 12px 30px rgba(15,23,42,0.08);font-size:14px;">
            Открываем нужный чек-лист...
        </div>
    </body>
    </html>
    """

def textarea_html(initial_dialog_id: str = "", initial_context_text: str = ""):
    initial_dialog_id_json = json.dumps(initial_dialog_id or "", ensure_ascii=False)
    initial_context_text_json = json.dumps(initial_context_text or "", ensure_ascii=False)


    return f"""
    <!doctype html>
    <html lang="ru">
    <head>
        <meta charset="utf-8">
        <title>Чек-лист ИД — textarea</title>
        <script src="https://api.bitrix24.com/api/v1/"></script>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 10px 12px;
                background: #fff;
            }}
            .wrap {{
                display: flex;
                align-items: center;
                gap: 10px;
            }}
            .dot {{
                width: 28px;
                height: 28px;
                border-radius: 999px;
                background: #ef5b8d;
                color: #fff;
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: 700;
                flex: 0 0 28px;
            }}
            .main {{
                min-width: 0;
                flex: 1;
            }}
            .title {{
                font-size: 13px;
                font-weight: 700;
                margin-bottom: 3px;
            }}
            .meta {{
                font-size: 11px;
                color: #666;
                margin-bottom: 6px;
                word-break: break-word;
            }}
            .btn {{
                display: inline-block;
                padding: 6px 10px;
                border: 1px solid #d0d7de;
                border-radius: 8px;
                background: #f6f8fa;
                cursor: pointer;
                font-size: 12px;
            }}
            .btn:hover {{
                background: #eef2f7;
            }}
            .error {{
                color: #b42318;
                font-size: 11px;
                margin-top: 6px;
                word-break: break-word;
            }}
        </style>
    </head>
    <body>
        <div class="wrap">
            <div class="dot">≡</div>
            <div class="main">
                <div class="title">Чек-лист ИД</div>
                <div class="meta" id="meta">Инициализация...</div>
                <button class="btn" id="openBtn" type="button">Открыть чек-лист</button>
                <div class="error" id="error"></div>
            </div>
        </div>

        <script>
            var initialDialogId = {initial_dialog_id_json};
            var initialContextText = {initial_context_text_json};
            var autoOpened = false;

            function detectAppBasePath() {{
                const path = String(window.location.pathname || '/').replace(/\/+$/, '');
                const suffixes = ['/launch', '/popup', '/textarea', '/install', '/health', '/debug/logs', '/admin', '/admin/upload'];

                for (const suffix of suffixes) {{
                    if (path === suffix) return '';
                    if (path.endsWith(suffix)) {{
                        return path.slice(0, -suffix.length) || '';
                    }}
                }}

                return path === '/' ? '' : path;
            }}

            const APP_BASE_PATH = detectAppBasePath();

            function appPath(path) {{
                return (APP_BASE_PATH || '') + '/' + String(path || '').replace(/^\/+/, '');
            }}
            function setMeta(text) {{
                document.getElementById('meta').textContent = text;
            }}

            function setError(text) {{
                document.getElementById('error').textContent = text || '';
            }}

            function openChecklist(dialogId, checklistKey = 'id') {{
                if (!dialogId) {{
                    setError('dialogId не найден');
                    return;
                }}

                try {{
                    localStorage.setItem('checklist_pending_dialog', JSON.stringify({{
                        dialogId: dialogId,
                        checklistKey: checklistKey,
                        ts: Date.now()
                    }}));
                }} catch (e) {{
                    console.log('localStorage save error:', e);
                }}

                try {{
                    if (window.BX24 && typeof window.BX24.openApplication === 'function') {{
                        BX24.openApplication({{
                            dialogId: dialogId,
                            checklistKey: checklistKey,
                            source: 'textarea'
                        }});
                        autoOpened = true;
                        setMeta('Открываем popup для ' + dialogId);
                        return;
                    }}
                }} catch (e) {{
                    setError('BX24.openApplication error: ' + String(e));
                }}

                window.open(
                    appPath('popup') +
                    '?dialogId=' + encodeURIComponent(dialogId) +
                    '&checklistKey=' + encodeURIComponent(checklistKey),
                    '_blank'
                );
            }}

            document.getElementById('openBtn').addEventListener('click', function () {{
                try {{
                    if (window.__dialogId) {{
                        openChecklist(window.__dialogId);
                    }} else {{
                        setError('dialogId ещё не определён');
                    }}
                }} catch (e) {{
                    setError(String(e));
                }}
            }});

            function finish(dialogId, sourceText) {{
                window.__dialogId = dialogId || '';
                setMeta('dialogId: ' + (window.__dialogId || 'не передан') + ' | source: ' + sourceText);

                try {{
                    if (window.BX24 && typeof window.BX24.fitWindow === 'function') {{
                        window.BX24.fitWindow();
                    }}
                }} catch (e) {{}}

                if (window.__dialogId && !autoOpened) {{
                    setTimeout(function() {{
                        openChecklist(window.__dialogId);
                    }}, 250);
                }}
            }}

            function canUseBx24() {{
                return !!(window.BX24 && typeof window.BX24.init === 'function');
            }}

            if (initialDialogId) {{
                finish(initialDialogId, 'server-post');
            }} else if (canUseBx24()) {{
                try {{
                    window.BX24.init(function () {{
                        var dialogId = '';
                        try {{
                            var info = window.BX24.placement.info() || {{}};
                            var options = info.options || {{}};

                            var candidates = [
                                options.dialogId,
                                options.DIALOG_ID,
                                options.dialog_id,
                                info.dialogId,
                                info.DIALOG_ID,
                                info.dialog_id,
                                options.chatId,
                                options.CHAT_ID,
                                options.chat_id,
                                info.chatId,
                                info.CHAT_ID,
                                info.chat_id
                            ];

                            for (var i = 0; i < candidates.length; i++) {{
                                var candidate = String(candidates[i] || '').trim();
                                if (candidate) {{
                                    dialogId = candidate;
                                    break;
                                }}
                            }}

                            try {{
                                console.log('placement.info =', info);
                            }} catch (e) {{}}
                        }} catch (e) {{
                            setError('placement.info error: ' + String(e));
                        }}

                        finish(dialogId, 'BX24-js');
                    }});
                }} catch (e) {{
                    setError('BX24.init error: ' + String(e));
                    finish('', 'BX24-init-failed');
                }}
            }} else {{
                setError(initialContextText || 'BX24 не найден');
                finish('', 'local');
            }}
        </script>
    </body>
    </html>
    """


@app.get("/health")
def health(request: Request):
    return {
        "ok": True,
        "appBasePathEnv": APP_BASE_PATH,
        "publicAppBaseUrlEnv": PUBLIC_APP_BASE_URL,
        "portalPath": APP_PORTAL_PATH,
        "requestBaseUrl": str(request.base_url).rstrip("/"),
        "publicBasePathDetected": get_public_app_base_path(request),
        "publicBaseUrlDetected": get_public_app_base_url(request),
        "xForwardedPrefix": request.headers.get("x-forwarded-prefix", ""),
        "xForwardedProto": request.headers.get("x-forwarded-proto", ""),
        "xForwardedHost": request.headers.get("x-forwarded-host", ""),
        "rootPath": request.scope.get("root_path", ""),
        "launchRoute": "/launch",
        "popupRoute": "/popup",
        "textareaRoute": "/textarea",
    }


@app.get("/", response_class=HTMLResponse)
def home_get(dialogId: str = "", checklistKey: str = "id", mode: str = ""):
    return app_home_html()


@app.post("/", response_class=HTMLResponse)
async def home_post(request: Request):
    form = dict(await request.form())

    def extract_checklist_key_from_form(form_data: dict) -> str:
        def pick(value) -> str:
            raw = str(value or "").strip()
            return normalize_checklist_key(raw) if raw else ""

        direct_candidates = [
            form_data.get("checklistKey"),
            form_data.get("CHECKLIST_KEY"),
            form_data.get("checklist_key"),
        ]

        for value in direct_candidates:
            found = pick(value)
            if found:
                return found

        def walk(obj) -> str:
            if isinstance(obj, dict):
                preferred_keys = [
                    "checklistKey", "CHECKLIST_KEY", "checklist_key"
                ]
                for key in preferred_keys:
                    found = pick(obj.get(key))
                    if found:
                        return found

                for value in obj.values():
                    found = walk(value)
                    if found:
                        return found

            elif isinstance(obj, list):
                for value in obj:
                    found = walk(value)
                    if found:
                        return found

            return ""

        json_candidates = [
            form_data.get("PLACEMENT_OPTIONS"),
            form_data.get("placementOptions"),
            form_data.get("options"),
        ]

        for raw in json_candidates:
            if not raw:
                continue
            try:
                data = json.loads(raw) if isinstance(raw, str) else raw
                found = walk(data)
                if found:
                    return found
            except Exception:
                pass

        return "id"

    dialog_id = extract_dialog_id_from_form(form)
    checklist_key = extract_checklist_key_from_form(form)
    raw_context = json.dumps(form, ensure_ascii=False, indent=2)

    print("HOME POST FORM:", raw_context)
    print("HOME EXTRACTED DIALOG ID:", dialog_id)
    print("HOME EXTRACTED CHECKLIST KEY:", checklist_key)

    return app_home_html(dialog_id, checklist_key, raw_context)

@app.get("/launch", response_class=HTMLResponse)
def launch_get(dialogId: str = "", checklistKey: str = "id"):
    return app_home_html()


@app.post("/launch", response_class=HTMLResponse)
async def launch_post(request: Request):
    return app_home_html()

@app.get("/install", response_class=HTMLResponse)
def install_get():
    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>Bitrix24 Install</title>
    </head>
    <body style="font-family:Arial,sans-serif;padding:40px">
        <h1>Установка приложения</h1>
        <p>Если эта страница открыта внутри Bitrix24, она завершит установку приложения.</p>
        {install_finish_block()}
    </body>
    </html>
    """


@app.post("/install", response_class=HTMLResponse)
@app.post("/install/", response_class=HTMLResponse)
async def install_post(request: Request):
    form = dict(await request.form())
    query = dict(request.query_params)
    params = {**query, **form}

    access_token = params.get("AUTH_ID") or params.get("access_token") or ""
    domain = normalize_domain(params.get("DOMAIN") or params.get("domain") or "")
    base_url = get_public_app_base_url(request)

    app_sid = params.get("APP_SID") or ""

    if app_sid and domain and not access_token:
        try:
            auth_response = requests.get(
                f"https://{domain}/rest/app.auth.json",
                params={"app_sid": app_sid},
                timeout=10
            )
            auth_data = auth_response.json()
            access_token = auth_data.get("result", {}).get("access_token", "")
        except Exception:
            pass

    bind_result = {
        "im_textarea": {"skipped": True}
    }
    placement_get_result = {
        "all": {"skipped": True}
    }

    if domain and access_token:
        bind_result["im_textarea"] = bitrix_rest_call(
            domain,
            "placement.bind",
            access_token,
            {
                "PLACEMENT": "IM_TEXTAREA",
                "HANDLER": f"{base_url}/textarea",
                "TITLE": "ТЕСТ",
                "OPTIONS[iconName]": "fa-bars",
                "OPTIONS[context]": "CHAT",
                "OPTIONS[role]": "ADMIN"
            }
        )

        placement_get_result["all"] = bitrix_rest_call(
            domain,
            "placement.get",
            access_token,
            {}
        )

    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>Bitrix24 Install Callback</title>
    </head>
    <body style="font-family:Arial,sans-serif;padding:40px">
        <h1>Install callback получен</h1>
        <p>Если bind прошёл успешно, launcher будет зарегистрирован в IM_TEXTAREA.</p>

        <h2>Что прислал Bitrix24 (query)</h2>
        <pre>{html.escape(json.dumps(query, ensure_ascii=False, indent=2))}</pre>

        <h2>Что прислал Bitrix24 (form)</h2>
        <pre>{html.escape(json.dumps(form, ensure_ascii=False, indent=2))}</pre>

        <h2>Ответ placement.bind</h2>
        <pre>{html.escape(json.dumps(bind_result, ensure_ascii=False, indent=2))}</pre>
        <h2>Ответ placement.get</h2>
        <pre>{html.escape(json.dumps(placement_get_result, ensure_ascii=False, indent=2))}</pre>

        {install_finish_block()}
    </body>
    </html>
    """


@app.get("/textarea", response_class=HTMLResponse)
def textarea_get(dialogId: str = ""):
    dialog_id = normalize_dialog_id(dialogId)
    return textarea_html(dialog_id, "GET /textarea")


@app.post("/textarea", response_class=HTMLResponse)
async def textarea_post(request: Request):
    form = dict(await request.form())
    dialog_id = extract_dialog_id_from_form(form)
    raw_context = json.dumps(form, ensure_ascii=False, indent=2)

    print("TEXTAREA POST FORM:", raw_context)
    print("TEXTAREA EXTRACTED DIALOG ID:", dialog_id)

    return textarea_html(dialog_id, raw_context)


@app.get("/popup", response_class=HTMLResponse)
def popup_get(dialogId: str = "", checklistKey: str = "id"):
    dialog_id = normalize_dialog_id(dialogId)
    checklist_key = normalize_checklist_key(checklistKey)
    data = get_checklist(dialog_id, checklist_key)

    title_raw = data.get("title", "Чек-лист ИД")
    title = html.escape(title_raw)
    collab_title_raw = (data.get("collabTitle", "") or "").strip()
    collab_title = html.escape(collab_title_raw)
    full_title = f"{title} — {collab_title}" if collab_title_raw else title
    progress_percent = int(data.get("progressPercent", 0) or 0)

    items_json = json.dumps(data.get("items", []), ensure_ascii=False)
    groups_json = json.dumps(data.get("groups", []), ensure_ascii=False)
    project_checklists_json = json.dumps(data.get("projectChecklists", []), ensure_ascii=False)
    dialog_id_json = json.dumps(dialog_id, ensure_ascii=False)
    collab_title_json = json.dumps(collab_title_raw, ensure_ascii=False)
    checklist_key_json = json.dumps(checklist_key, ensure_ascii=False)
    checklist_title_json = json.dumps(title_raw, ensure_ascii=False)

    popup_session_enhancements_js = """
            const clientSessionId = 'popup_' + Date.now() + '_' + Math.random().toString(36).slice(2, 8);
            let checklistSessionState = {};
            let currentChecklistLock = {
                owned: false,
                lockedByOther: false,
                lockId: '',
                userId: '',
                userName: '',
                checklistKey: ''
            };
            let lockHeartbeatTimer = null;
            let suppressAutoCloseSave = false;
            let activeInlineEditor = { role: '', itemId: '' };

            const headerRightEl = document.querySelector('.header-right');
            if (headerRightEl && !document.getElementById('lockNotice')) {
                const lockNode = document.createElement('div');
                lockNode.id = 'lockNotice';
                lockNode.style.fontSize = '12px';
                lockNode.style.fontWeight = '700';
                lockNode.style.padding = '7px 10px';
                lockNode.style.borderRadius = '999px';
                lockNode.style.background = '#fff4e5';
                lockNode.style.color = '#b26a00';
                lockNode.style.whiteSpace = 'nowrap';
                lockNode.style.display = 'none';
                headerRightEl.insertBefore(lockNode, saveStateEl);
            }
            const lockNoticeEl = document.getElementById('lockNotice');

            const contentEl = document.querySelector('.content');
            if (contentEl && !document.getElementById('footerActions')) {
                const footerNode = document.createElement('div');
                footerNode.id = 'footerActions';
                footerNode.style.position = 'sticky';
                footerNode.style.bottom = '0';
                footerNode.style.zIndex = '30';
                footerNode.style.marginTop = '14px';
                footerNode.style.padding = '10px 12px';
                footerNode.style.display = 'flex';
                footerNode.style.gap = '10px';
                footerNode.style.justifyContent = 'flex-end';
                footerNode.style.alignItems = 'center';
                footerNode.style.border = '1px solid #e5e7eb';
                footerNode.style.borderRadius = '12px';
                footerNode.style.background = '#fafbfc';
                footerNode.style.boxShadow = '0 -4px 18px rgba(15,23,42,.06)';
                footerNode.innerHTML = `
                    <button id="saveCloseBtn" type="button" style="min-width:150px;height:34px;border:none;border-radius:8px;background:#16a34a;color:#fff;font-size:12px;font-weight:700;cursor:pointer;">
                        Сохранить и закрыть
                    </button>
                    <button id="cancelBtn" type="button" style="min-width:100px;height:34px;border:none;border-radius:8px;background:#dc2626;color:#fff;font-size:12px;font-weight:700;cursor:pointer;">
                        Отмена
                    </button>
                `;
                contentEl.appendChild(footerNode);
            }
            const saveCloseBtn = document.getElementById('saveCloseBtn');
            const cancelBtn = document.getElementById('cancelBtn');

            function getChecklistDefaultTitle(key) {
                if (key === 'concept') return 'Чек-лист Концепция';
                if (key === 'opr') return 'Чек-лист ОПР';
                return 'Чек-лист ИД';
            }

            function getChecklistState(key) {
                const stateKey = String(key || '').trim() || 'id';
                if (!checklistSessionState[stateKey]) {
                    checklistSessionState[stateKey] = { changes: [], dirty: false };
                }
                return checklistSessionState[stateKey];
            }

            function getCurrentEditorIdentity() {
                return {
                    userId: String(currentEditor.id || clientSessionId),
                    userName: String(currentEditor.name || 'Пользователь')
                };
            }

            function isChecklistLockedByOther() {
                return !!(currentChecklistLock.lockedByOther && currentChecklistLock.checklistKey === currentChecklistKey);
            }

            function isEditingAllowed() {
                return !isChecklistLockedByOther();
            }

            function disabledAttr() {
                return isEditingAllowed() ? '' : 'disabled';
            }

            function setActionButtonState() {
                if (saveCloseBtn) {
                    saveCloseBtn.disabled = !isEditingAllowed();
                    saveCloseBtn.style.opacity = saveCloseBtn.disabled ? '0.55' : '1';
                    saveCloseBtn.style.cursor = saveCloseBtn.disabled ? 'not-allowed' : 'pointer';
                }
                if (cancelBtn) {
                    cancelBtn.disabled = false;
                }
            }

            function updateSaveStateBySession() {
                if (isChecklistLockedByOther()) {
                    setSaveState('error', 'Только просмотр');
                    return;
                }
                if (sessionDirty) {
                    setSaveState('saving', 'Есть несохраненные изменения');
                    return;
                }
                setSaveState('', 'Сохранено');
            }

            function updateLockNotice() {
                if (!lockNoticeEl) return;
                if (isChecklistLockedByOther()) {
                    const ownerName = currentChecklistLock.userName || 'другой сотрудник';
                    lockNoticeEl.textContent = 'Сейчас с этим чек-листом работает ' + ownerName + ', дождитесь завершения сессии';
                    lockNoticeEl.style.display = '';
                } else {
                    lockNoticeEl.textContent = '';
                    lockNoticeEl.style.display = 'none';
                }
                setActionButtonState();
                updateSaveStateBySession();
            }

            const previousSyncChecklistCache = syncChecklistCache;
            syncChecklistCache = function () {
                checklistCache[currentChecklistKey] = buildChecklistSnapshot();
                checklistSessionState[currentChecklistKey] = {
                    changes: deepClone(sessionChanges),
                    dirty: !!sessionDirty
                };
                if (typeof previousSyncChecklistCache === 'function') {
                    previousSyncChecklistCache();
                }
            };

            const previousApplyChecklistData = applyChecklistData;
            applyChecklistData = function (data) {
                previousApplyChecklistData(data);
                checklistTitle = String((data && data.title) || checklistTitle || getChecklistDefaultTitle(currentChecklistKey));
                const cachedState = getChecklistState(currentChecklistKey);
                sessionChanges = deepClone(cachedState.changes || []);
                sessionDirty = !!cachedState.dirty;
                closeSummarySent = false;
                activeInlineEditor = { role: '', itemId: '' };
                currentChecklistLock.checklistKey = currentChecklistKey;
                updateLockNotice();
            };

            function clearChecklistSessionState(checklistKey) {
                const targetKey = String(checklistKey || '').trim() || 'id';
                checklistSessionState[targetKey] = { changes: [], dirty: false };
                if (targetKey === currentChecklistKey) {
                    sessionChanges = [];
                    sessionDirty = false;
                    closeSummarySent = false;
                    updateLockNotice();
                }
            }

            function getDirtyChecklistKeys() {
                syncChecklistCache();
                return Object.keys(checklistSessionState).filter(key => {
                    const state = checklistSessionState[key];
                    return !!(state && state.dirty && Array.isArray(state.changes) && state.changes.length && checklistCache[key]);
                });
            }

            function buildClosePayloadForKey(checklistKey, closeEvent) {
                const key = String(checklistKey || '').trim() || 'id';
                const snapshot = key === currentChecklistKey
                    ? buildChecklistSnapshot()
                    : deepClone(checklistCache[key] || {});
                const state = key === currentChecklistKey
                    ? { changes: deepClone(sessionChanges), dirty: !!sessionDirty }
                    : deepClone(checklistSessionState[key] || { changes: [], dirty: false });

                return {
                    dialogId,
                    checklistKey: key,
                    editor: currentEditor,
                    data: snapshot,
                    changes: state.changes || [],
                    closeEvent,
                    ts: new Date().toISOString()
                };
            }

            function buildCloseBatchPayload(closeEvent) {
                const dirtyKeys = getDirtyChecklistKeys();
                return {
                    dialogId,
                    editor: currentEditor,
                    closeEvent,
                    ts: new Date().toISOString(),
                    sessions: dirtyKeys.map(key => buildClosePayloadForKey(key, closeEvent))
                };
            }

            async function persistDirtyChecklists(closeEvent, useBeacon = false) {
                const dirtyKeys = getDirtyChecklistKeys();
                if (!dirtyKeys.length) {
                    return {
                        savedCount: 0,
                        messageOk: true,
                        messageSkipped: true,
                        result: {}
                    };
                }

                const payload = buildCloseBatchPayload(closeEvent);

                if (useBeacon && navigator.sendBeacon) {
                    const blob = new Blob([JSON.stringify(payload)], { type: 'application/json' });
                    navigator.sendBeacon(APP_BASE_URL + '/api/checklist/close-session', blob);
                    dirtyKeys.forEach(clearChecklistSessionState);
                    return {
                        savedCount: dirtyKeys.length,
                        messageOk: true,
                        messageSkipped: false,
                        result: {}
                    };
                }

                const response = await fetch(appUrl('api/checklist/close-session'), {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify(payload),
                    keepalive: true
                });
                const result = await response.json().catch(() => ({}));
                if (!response.ok) {
                    throw new Error(result.error || 'close-session failed');
                }

                if (result && result.messageOk === false) {
                    debugLog('close_session_message_warning', {
                        closeEvent,
                        dialogId,
                        checklistKeys: dirtyKeys,
                        result
                    });
                } else {
                    debugLog('close_session_completed', {
                        closeEvent,
                        dialogId,
                        checklistKeys: dirtyKeys,
                        result
                    });
                }

                dirtyKeys.forEach(clearChecklistSessionState);
                return {
                    savedCount: dirtyKeys.length,
                    messageOk: result.messageOk !== false,
                    messageSkipped: !!result.messageSkipped,
                    result
                };
            }

            function stopLockHeartbeat() {
                if (lockHeartbeatTimer) {
                    clearInterval(lockHeartbeatTimer);
                    lockHeartbeatTimer = null;
                }
            }

            async function acquireChecklistLock(checklistKey = currentChecklistKey, silent = false) {
                const targetKey = String(checklistKey || '').trim() || 'id';
                const editorIdentity = getCurrentEditorIdentity();
                const payload = {
                    dialogId,
                    checklistKey: targetKey,
                    userId: editorIdentity.userId,
                    userName: editorIdentity.userName,
                    lockId: currentChecklistLock.checklistKey === targetKey ? currentChecklistLock.lockId : ''
                };

                try {
                    const response = await fetch(appUrl('api/checklist/lock/acquire'), {
                        method: 'POST',
                        headers: {
                            'Content-Type': 'application/json'
                        },
                        body: JSON.stringify(payload),
                        keepalive: true
                    });
                    const result = await response.json();
                    if (!response.ok) {
                        throw new Error(result.error || 'lock acquire failed');
                    }

                    if (targetKey !== currentChecklistKey) {
                        return result;
                    }

                    const prevOwned = currentChecklistLock.owned;
                    const prevBlocked = currentChecklistLock.lockedByOther;
                    const prevLockId = currentChecklistLock.lockId;
                    currentChecklistLock = {
                        owned: !!result.owned,
                        lockedByOther: !!result.lockedByOther,
                        lockId: String(result.lockId || ''),
                        userId: String(result.userId || ''),
                        userName: String(result.userName || ''),
                        checklistKey: targetKey
                    };
                    updateLockNotice();

                    if (prevOwned !== currentChecklistLock.owned || prevBlocked !== currentChecklistLock.lockedByOther || prevLockId !== currentChecklistLock.lockId) {
                        renderAll();
                    }
                    return result;
                } catch (e) {
                    if (!silent) {
                        console.log('acquireChecklistLock error:', e);
                    }
                    return null;
                }
            }

            async function heartbeatChecklistLock(silent = true) {
                if (!currentChecklistLock.lockId || currentChecklistLock.checklistKey !== currentChecklistKey) {
                    return null;
                }

                const editorIdentity = getCurrentEditorIdentity();
                try {
                    const response = await fetch(appUrl('api/checklist/lock/heartbeat'), {
                        method: 'POST',
                        headers: {
                            'Content-Type': 'application/json'
                        },
                        body: JSON.stringify({
                            dialogId,
                            checklistKey: currentChecklistKey,
                            userId: editorIdentity.userId,
                            userName: editorIdentity.userName,
                            lockId: currentChecklistLock.lockId
                        }),
                        keepalive: true
                    });
                    const result = await response.json();
                    if (!response.ok) {
                        throw new Error(result.error || 'lock heartbeat failed');
                    }
                    currentChecklistLock = {
                        owned: !!result.owned,
                        lockedByOther: !!result.lockedByOther,
                        lockId: String(result.lockId || ''),
                        userId: String(result.userId || ''),
                        userName: String(result.userName || ''),
                        checklistKey: currentChecklistKey
                    };
                    updateLockNotice();
                    return result;
                } catch (e) {
                    if (!silent) {
                        console.log('heartbeatChecklistLock error:', e);
                    }
                    return null;
                }
            }

            function startLockHeartbeat() {
                stopLockHeartbeat();
                lockHeartbeatTimer = setInterval(async function () {
                    if (currentChecklistLock.owned) {
                        await heartbeatChecklistLock(true);
                    } else {
                        await acquireChecklistLock(currentChecklistKey, true);
                    }
                }, 15000);
            }

            async function releaseChecklistLock(checklistKey = currentChecklistKey, useBeacon = false) {
                const targetKey = String(checklistKey || '').trim() || 'id';
                const lockId = currentChecklistLock.checklistKey === targetKey ? currentChecklistLock.lockId : '';
                const editorIdentity = getCurrentEditorIdentity();
                stopLockHeartbeat();

                if (!lockId && currentChecklistLock.checklistKey === targetKey) {
                    currentChecklistLock = {
                        owned: false,
                        lockedByOther: false,
                        lockId: '',
                        userId: '',
                        userName: '',
                        checklistKey: targetKey
                    };
                    updateLockNotice();
                    return;
                }

                const payload = {
                    dialogId,
                    checklistKey: targetKey,
                    lockId,
                    userId: editorIdentity.userId
                };

                if (useBeacon && navigator.sendBeacon) {
                    const blob = new Blob([JSON.stringify(payload)], { type: 'application/json' });
                    navigator.sendBeacon(APP_BASE_URL + '/api/checklist/lock/release', blob);
                } else {
                    try {
                        await fetch(appUrl('api/checklist/lock/release'), {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json'
                            },
                            body: JSON.stringify(payload),
                            keepalive: true
                        });
                    } catch (e) {
                        console.log('releaseChecklistLock error:', e);
                    }
                }

                if (currentChecklistLock.checklistKey === targetKey) {
                    currentChecklistLock = {
                        owned: false,
                        lockedByOther: false,
                        lockId: '',
                        userId: '',
                        userName: '',
                        checklistKey: targetKey
                    };
                    updateLockNotice();
                }
            }

            function closePopupWindow() {
                try {
                    if (window.BX24 && typeof window.BX24.closeApplication === 'function') {
                        window.BX24.closeApplication();
                        return;
                    }
                } catch (e) {
                    console.log('BX24.closeApplication error:', e);
                }

                try {
                    window.close();
                } catch (e) {
                    console.log('window.close error:', e);
                }
            }

            async function finalizePopupSession(saveChanges) {
                suppressAutoCloseSave = true;
                closeSummarySent = true;
                syncChecklistCache();

                let persistResult = {
                    savedCount: 0,
                    messageOk: true,
                    messageSkipped: true,
                    result: {}
                };

                try {
                    if (saveChanges) {
                        persistResult = await persistDirtyChecklists('save_and_close', false);
                    }
                } catch (e) {
                    console.log('finalizePopupSession error:', e);
                    setSaveState('error', saveChanges ? 'Ошибка сохранения' : 'Ошибка отмены');
                    return;
                }

                if (saveChanges) {
                    if (persistResult.messageOk === false) {
                        setSaveState('saving', 'Сохранено, сообщение не отправлено');
                    } else if (persistResult.savedCount > 0) {
                        setSaveState('', 'Сохранено');
                    }
                }

                await releaseChecklistLock(currentChecklistKey, false);
                closePopupWindow();
            }

            if (saveCloseBtn) {
                saveCloseBtn.addEventListener('click', async function () {
                    await finalizePopupSession(true);
                });
            }

            if (cancelBtn) {
                cancelBtn.addEventListener('click', async function () {
                    await finalizePopupSession(false);
                });
            }

            sendCloseSummaryOnce = function (eventName) {
                if (eventName === 'popup_hidden' || suppressAutoCloseSave || closeSummarySent) {
                    return;
                }

                closeSummarySent = true;
                persistDirtyChecklists(eventName, true);
                releaseChecklistLock(currentChecklistKey, true);
            };

            loadChecklistByKey = async function (checklistKey) {
                const targetKey = String(checklistKey || '').trim() || 'id';
                if (targetKey === currentChecklistKey) {
                    return;
                }

                syncChecklistCache();
                await releaseChecklistLock(currentChecklistKey, false);
                setSaveState('saving', 'Загружаем...');

                try {
                    const cachedData = checklistCache[targetKey];
                    if (cachedData) {
                        applyChecklistData(deepClone(cachedData));
                        renderAll();
                        await acquireChecklistLock(targetKey, true);
                        startLockHeartbeat();
                        return;
                    }

                    const response = await fetch(
                        appUrl('api/checklist') +
                        '?dialogId=' + encodeURIComponent(dialogId) +
                        '&checklistKey=' + encodeURIComponent(targetKey)
                    );
                    const result = await response.json();
                    if (!response.ok) {
                        throw new Error(result.error || 'load checklist failed');
                    }

                    applyChecklistData(result);
                    renderAll();
                    await acquireChecklistLock(targetKey, true);
                    startLockHeartbeat();
                } catch (e) {
                    console.log('loadChecklistByKey enhanced error:', e);
                    setSaveState('error', 'Ошибка загрузки чек-листа');
                }
            };

            buildDocumentCell = function (item) {
                if (normalizeStatus(item && item.status) === 'Не требуется') {
                    return '';
                }
                const documents = getItemDocuments(item);
                const itemId = String(item && item.id || '');
                const folderViewUrl = String(item.folderUrl || '').trim() || (documents.length ? (
                    appUrl('api/checklist/folder') +
                    '?dialogId=' + encodeURIComponent(dialogId) +
                    '&checklistKey=' + encodeURIComponent(currentChecklistKey) +
                    '&itemId=' + encodeURIComponent(itemId)
                ) : '');
                const showViewFolder = documents.length > 0 && !!folderViewUrl;

                const filesHtml = documents.map(doc => {
                    const docId = String(doc.id || '');
                    const docName = String(doc.name || 'Файл');
                    const openUrl = appUrl('api/checklist/file') +
                        '?dialogId=' + encodeURIComponent(dialogId) +
                        '&checklistKey=' + encodeURIComponent(currentChecklistKey) +
                        '&itemId=' + encodeURIComponent(itemId) +
                        '&documentId=' + encodeURIComponent(docId);
                    const sizeText = formatFileSize(doc.size || 0);

                    return `
                        <div class="doc-file-row">
                            <a
                                href="javascript:void(0)"
                                class="doc-file-link"
                                data-role="view-file"
                                data-item-id="${esc(itemId)}"
                                data-document-id="${esc(docId)}"
                                data-open-url="${esc(openUrl)}"
                                title="${esc(docName)}"
                            >
                                ${esc(docName)}
                            </a>
                            <button
                                type="button"
                                class="doc-file-remove"
                                data-role="remove-file"
                                data-item-id="${esc(itemId)}"
                                data-document-id="${esc(docId)}"
                                data-document-name="${esc(docName)}"
                                title="Удалить файл"
                                ${typeof disabledAttr === 'function' ? disabledAttr() : ''}
                            >
                                ×
                            </button>
                            ${sizeText ? `<span class="doc-file-meta">${esc(sizeText)}</span>` : ''}
                        </div>
                    `;
                }).join('');

                return `
                    <div class="doc-cell">
                        <div class="doc-actions">
                            <button
                                class="upload-btn"
                                type="button"
                                data-role="upload"
                                data-item-id="${esc(itemId)}"
                                ${typeof disabledAttr === 'function' ? disabledAttr() : ''}
                            >
                                Загрузить
                            </button>

                            ${showViewFolder ? `
                                <button
                                    class="doc-btn"
                                    type="button"
                                    data-role="view-folder"
                                    data-item-id="${esc(itemId)}"
                                    data-folder-url="${esc(folderViewUrl)}"
                                >
                                    Посмотреть
                                </button>
                            ` : ''}
                        </div>

                        ${documents.length ? `
                            <div class="doc-files">
                                ${filesHtml}
                            </div>
                        ` : ''}

                        <input
                            type="file"
                            data-role="file-input"
                            data-item-id="${esc(itemId)}"
                            style="display:none;"
                            multiple
                            ${typeof disabledAttr === 'function' ? disabledAttr() : ''}
                        >
                    </div>
                `;
            };

            renderGroup = function (group) {
                const groupItems = getItemsByGroup(group.id);
                const allowAdd = Number(group.id) !== 4;
                const rows = groupItems.map(item => {
                    const rowClass = normalizeStatus(item.status) === 'Не требуется' ? 'row not-required' : 'row';
                    return `
                        <div class="${rowClass}" data-item-id="${esc(item.id)}">
                            <div class="td"><div class="cell-name"><div class="${indicatorClass(item.status)}"></div><div class="item-name">${esc(item.name)}</div></div></div>
                            <div class="td">${buildDocumentCell(item)}</div>
                            <div class="td">
                                <select class="status-select" data-role="status" data-item-id="${esc(item.id)}" ${disabledAttr()}>
                                    <option value="" ${normalizeStatus(item.status) === '' ? 'selected' : ''}></option>
                                    <option value="Есть" ${normalizeStatus(item.status) === 'Есть' ? 'selected' : ''}>Есть</option>
                                    <option value="Нет" ${normalizeStatus(item.status) === 'Нет' ? 'selected' : ''}>Нет</option>
                                    <option value="Не требуется" ${normalizeStatus(item.status) === 'Не требуется' ? 'selected' : ''}>Не требуется</option>
                                </select>
                            </div>
                            <div class="td"><input class="date-input" type="date" data-role="plan" data-item-id="${esc(item.id)}" value="${esc(toInputDate(item.plan))}" ${disabledAttr()}></div>
                            <div class="td"><input class="date-input" type="date" data-role="fact" data-item-id="${esc(item.id)}" value="${esc(toInputDate(item.fact))}" ${disabledAttr()}></div>
                        </div>
                    `;
                }).join('');
                const addBlock = allowAdd ? `
                    <div class="add-item-row">
                        <input class="add-item-input" id="addItemInput_${group.id}" type="text" placeholder="Новый пункт" ${disabledAttr()}>
                        <button class="add-item-btn" type="button" data-role="add-item" data-group-id="${group.id}" ${disabledAttr()}>Добавить пункт</button>
                    </div>` : '';
                return `<div class="group-block"><div class="group-title">${esc(group.title)}</div>${rows}${addBlock}</div>`;
            };

            function getInlineEditorDomId(role, itemId) {
                return 'inline_' + String(role || '').replace(/[^a-z0-9_-]/ig, '_') + '_' + String(itemId || '').replace(/[^a-z0-9_-]/ig, '_');
            }

            function isInlineEditorActive(role, itemId) {
                return activeInlineEditor.role === role && String(activeInlineEditor.itemId || '') === String(itemId || '');
            }

            function startInlineEditor(role, itemId) {
                if (!isEditingAllowed()) {
                    return;
                }
                activeInlineEditor = {
                    role,
                    itemId: String(itemId || '')
                };
                renderAll();
            }

            function stopInlineEditor() {
                activeInlineEditor = { role: '', itemId: '' };
            }

            function focusActiveInlineEditor() {
                if (!activeInlineEditor.role || !activeInlineEditor.itemId) {
                    return;
                }
                const el = document.getElementById(getInlineEditorDomId(activeInlineEditor.role, activeInlineEditor.itemId));
                if (!el) {
                    return;
                }
                if (typeof el.focus === 'function') {
                    el.focus();
                }
                if (typeof el.select === 'function') {
                    el.select();
                }
                if (el.tagName === 'TEXTAREA') {
                    el.dataset.initialValue = String(el.value || '');
                    el.dataset.baseHeight = '32';
                    autoGrowTextarea(el);
                }
            }

            function buildConceptInlineCell(role, item, value, placeholder, extraStyle = '') {
                const itemId = String(item.id || '');
                const isActive = isInlineEditorActive(role, itemId) && isEditingAllowed();
                const displayValue = String(value || '').trim();
                const content = displayValue || String(placeholder || '').trim() || '';
                const displayClass = 'concept-inline-display' + (displayValue ? '' : ' empty') + (isEditingAllowed() ? '' : ' disabled');
                const styleAttr = extraStyle ? ` style="${extraStyle}"` : '';

                if (isActive) {
                    return `
                        <textarea
                            id="${esc(getInlineEditorDomId(role, itemId))}"
                            class="concept-inline-input"
                            data-role="${esc(role + '-edit')}"
                            data-item-id="${esc(itemId)}"
                            data-initial-value="${esc(value || '')}"
                            placeholder="${esc(placeholder || '')}"
                            ${disabledAttr()}
                        >${esc(value || '')}</textarea>
                    `;
                }

                return `
                    <div
                        class="${displayClass}"
                        data-role="${esc(role + '-display')}"
                        data-item-id="${esc(itemId)}"
                        tabindex="${isEditingAllowed() ? '0' : '-1'}"
                        ${styleAttr}
                    >${esc(content) || '&nbsp;'}</div>
                `;
            }

            function buildConceptNameCell(item) {
                const textStyle = item.status === 'Не требуется' ? 'text-decoration:line-through;color:#98a2b3;' : '';
                return buildConceptInlineCell('concept-name', item, item.name || '', 'Название пункта', textStyle);
            }

            function buildConceptSourceCell(item) {
                return buildConceptInlineCell('concept-source', item, item.source || '', 'Нормативы');
            }

            buildConceptStatusCell = function (item) {
                if (item.statusKind === 'bool') {
                    return `
                        <select class="status-select" data-role="concept-status" data-item-id="${esc(item.id)}" ${disabledAttr()}>
                            <option value="" ${item.status === '' ? 'selected' : ''}></option>
                            <option value="Да" ${item.status === 'Да' ? 'selected' : ''}>Да</option>
                            <option value="Нет" ${item.status === 'Нет' ? 'selected' : ''}>Нет</option>
                            <option value="Не требуется" ${item.status === 'Не требуется' ? 'selected' : ''}>Не требуется</option>
                        </select>
                    `;
                }
                if (item.statusKind === 'select') {
                    const options = [''].concat(item.statusOptions || [], ['Не требуется']);
                    return `
                        <select class="status-select" data-role="concept-status" data-item-id="${esc(item.id)}" ${disabledAttr()}>
                            ${options.map(option => `<option value="${esc(option)}" ${item.status === option ? 'selected' : ''}>${esc(option)}</option>`).join('')}
                        </select>
                    `;
                }
                return `<input class="status-select" type="text" data-role="concept-status" data-item-id="${esc(item.id)}" placeholder="${esc(item.statusPlaceholder || '')}" value="${esc(item.status || '')}" ${disabledAttr()}>`;
            };

            buildConceptExtraCell = function (item) {
                return `<textarea class="concept-extra-textarea" data-role="concept-extra" data-item-id="${esc(item.id)}" placeholder="${esc(item.extraInfoPlaceholder || '')}" ${disabledAttr()}>${esc(item.extraInfo || '')}</textarea>`;
            };

            renderConceptGroup = function (group) {
                const groupItems = getItemsByGroup(group.id);
                const allowAdd = Number(group.id) !== 10;
                const rows = groupItems.map(item => `
                    <div class="row" style="grid-template-columns: 1.05fr 0.66fr 190px 150px 1.15fr;" data-item-id="${esc(item.id)}">
                        <div class="td"><div class="cell-name"><div class="${conceptIndicatorClass(item)}"></div>${buildConceptNameCell(item)}</div></div>
                        <div class="td">${buildConceptSourceCell(item)}</div>
                        <div class="td">${buildDocumentCell(item)}</div>
                        <div class="td">${buildConceptStatusCell(item)}</div>
                        <div class="td">${buildConceptExtraCell(item)}</div>
                    </div>
                `).join('');

                const addBlock = allowAdd ? `
                    <div class="add-item-row">
                        <input class="add-item-input" id="conceptAddItemInput_${group.id}" type="text" placeholder="Новый пункт" ${disabledAttr()}>
                        <button class="add-item-btn" type="button" data-role="concept-add-item" data-group-id="${group.id}" ${disabledAttr()}>Добавить пункт</button>
                    </div>
                ` : '';

                return `<div class="group-block"><div class="group-title">${esc(group.title)}</div>${rows}${addBlock}</div>`;
            };

            function buildConceptTableHtmlEnhanced(conceptGroups) {
                return `
                    <div class="thead">
                        <div class="thead-top" style="grid-template-columns: 1.05fr 0.66fr 190px 150px 1.15fr;">
                            <div class="th">Пункт</div>
                            <div class="th">Нормативы</div>
                            <div class="th">Документ</div>
                            <div class="th">Статус</div>
                            <div class="th">Доп информация</div>
                        </div>
                    </div>
                    <div>${conceptGroups.map(renderConceptGroup).join('')}</div>
                `;
            }

            renderConceptTable = function () {
                if (!leftTableEl || !rightTableEl || !tablesGridEl) {
                    throw new Error('concept table containers not found');
                }

                tablesGridEl.style.gridTemplateColumns = '1fr 1fr';
                if (tablePanels[1]) {
                    tablePanels[1].style.display = '';
                }

                const visibleGroups = groups.filter(group => {
                    if (Number(group.id) !== 10) return true;
                    return items.some(x => Number(x.group) === 10);
                });

                const leftGroups = visibleGroups.filter(group => [1, 3, 5, 7, 9].includes(Number(group.id)));
                const rightGroups = visibleGroups.filter(group => [2, 4, 6, 8, 10].includes(Number(group.id)));

                leftTableEl.innerHTML = buildConceptTableHtmlEnhanced(leftGroups);
                rightTableEl.innerHTML = buildConceptTableHtmlEnhanced(rightGroups);
            };

            renderProjectChecklistList = function () {
                if (!projectChecklistListEl) {
                    return;
                }

                const normalizedList = (Array.isArray(projectChecklists) ? projectChecklists : []).map(item => {
                    const key = String(item && item.key || '').trim();
                    return {
                        key,
                        title: getChecklistDefaultTitle(key)
                    };
                });

                projectChecklistListEl.innerHTML = normalizedList.map(item => {
                    const active = item.key === currentChecklistKey ? 'side-link active' : 'side-link';
                    return `<button type="button" class="${active}" data-checklist-key="${esc(item.key)}">${esc(item.title)}</button>`;
                }).join('');

                projectChecklistListEl.querySelectorAll('[data-checklist-key]').forEach(btn => {
                    btn.addEventListener('click', async function () {
                        const key = this.dataset.checklistKey;
                        await loadChecklistByKey(key);
                    });
                });
            };

            oprIndicatorClass = function (item) {
                const status = normalizeStatus(item && item.status);
                if (status === 'Есть') return 'status-indicator green';
                if (status === 'Нет' || status === 'Не требуется') return 'status-indicator gray';
                return 'status-indicator';
            };

            function resolveOprGroupIdByItemIdOrName(item) {
                const itemId = String(item && item.id || '');
                if (itemId.startsWith('opr_g')) {
                    const match = itemId.match(/^opr_g(\d+)_/);
                    if (match) {
                        const groupId = Number(match[1]);
                        if (groupId && groupId !== 2) {
                            return groupId;
                        }
                    }
                }

                const name = String(item && item.name || '').trim();
                const matchedGroup = (Array.isArray(groups) ? groups : []).find(group => {
                    const gid = Number(group && group.id);
                    if (gid === 2) return false;

                    return Array.isArray(items) && items.some(existing =>
                        existing !== item &&
                        Number(existing.group) === gid &&
                        String(existing.name || '').trim() === name
                    );
                });

                return matchedGroup ? Number(matchedGroup.id) : 1;
            }

            function buildOprDatesToggle() {
                const active = !!oprDateVisibility[1];
                return `
                    <button
                        type="button"
                        class="id-dates-toggle"
                        data-role="opr-toggle-dates"
                        title="${active ? 'Скрыть даты' : 'Показать даты'}"
                        aria-label="${active ? 'Скрыть даты' : 'Показать даты'}"
                        ${disabledAttr()}
                    >
                        📅
                    </button>
                `;
            }

            function buildOprStatusCellUi(item) {
                return `
                    <select class="status-select" data-role="opr-status" data-item-id="${esc(item.id)}" ${disabledAttr()}>
                        <option value="" ${normalizeStatus(item.status) === '' ? 'selected' : ''}></option>
                        <option value="Есть" ${normalizeStatus(item.status) === 'Есть' ? 'selected' : ''}>Есть</option>
                        <option value="Нет" ${normalizeStatus(item.status) === 'Нет' ? 'selected' : ''}>Нет</option>
                        <option value="Не требуется" ${normalizeStatus(item.status) === 'Не требуется' ? 'selected' : ''}>Не требуется</option>
                    </select>
                `;
            }

            function renderOprGroupUi(group) {
                const groupItems = getItemsByGroup(group.id);
                const allowAdd = Number(group.id) !== 2;
                const showDates = !!oprDateVisibility[1];
                const gridClass = showDates ? 'id-grid id-grid-expanded' : 'id-grid id-grid-compact';

                const rows = groupItems.map(item => {
                    const rowClass = normalizeStatus(item.status) === 'Не требуется'
                        ? `row not-required ${gridClass}`
                        : `row ${gridClass}`;

                    return `
                        <div class="${rowClass}" data-item-id="${esc(item.id)}">
                            <div class="td">
                                <div class="cell-name">
                                    <div class="${oprIndicatorClass(item)}"></div>
                                    <div class="item-name">${esc(item.name)}</div>
                                </div>
                            </div>
                            <div class="td">${buildDocumentCell(item)}</div>
                            <div class="td">${buildOprStatusCellUi(item)}</div>
                            ${showDates ? `
                                <div class="td">
                                    <input class="date-input" type="date" data-role="opr-plan" data-item-id="${esc(item.id)}" value="${esc(toInputDate(item.plan || ''))}" ${disabledAttr()}>
                                </div>
                                <div class="td">
                                    <input class="date-input" type="date" data-role="opr-fact" data-item-id="${esc(item.id)}" value="${esc(toInputDate(item.fact || ''))}" ${disabledAttr()}>
                                </div>
                            ` : ''}
                        </div>
                    `;
                }).join('');

                const addBlock = allowAdd ? `
                    <div class="add-item-row">
                        <input class="add-item-input" id="oprAddItemInput_${group.id}" type="text" placeholder="Новый пункт" ${disabledAttr()}>
                        <button class="add-item-btn" type="button" data-role="opr-add-item" data-group-id="${group.id}" ${disabledAttr()}>Добавить пункт</button>
                    </div>
                ` : '';

                return `<div class="group-block"><div class="group-title">${esc(group.title)}</div>${rows}${addBlock}</div>`;
            }

            renderOprTables = function () {
                if (!leftTableEl || !rightTableEl || !tablesGridEl) {
                    throw new Error('opr table containers not found');
                }

                const showDates = !!oprDateVisibility[1];
                const gridClass = showDates ? 'id-grid id-grid-expanded' : 'id-grid id-grid-compact';

                tablesGridEl.style.gridTemplateColumns = 'minmax(0, 37%)';
                tablesGridEl.style.justifyContent = 'start';

                if (tablePanels[0]) {
                    tablePanels[0].style.display = '';
                    tablePanels[0].style.flex = '0 0 37%';
                    tablePanels[0].style.maxWidth = '37%';
                }
                if (tablePanels[1]) {
                    tablePanels[1].style.display = 'none';
                    tablePanels[1].style.flex = '';
                    tablePanels[1].style.maxWidth = '';
                }
                if (tablePanels[2]) {
                    tablePanels[2].style.display = 'none';
                    tablePanels[2].style.flex = '';
                    tablePanels[2].style.maxWidth = '';
                }

                leftTableEl.style.width = '100%';
                leftTableEl.style.maxWidth = '100%';

                const visibleGroups = groups.filter(group => {
                    if (Number(group.id) !== 2) return true;
                    return items.some(x => Number(x.group) === 2);
                });

                leftTableEl.classList.add('id-table');
                leftTableEl.innerHTML = `
                    <div class="thead">
                        <div class="thead-top ${gridClass}">
                            <div class="th">ОПР</div>
                            <div class="th">Документ</div>
                            <div class="th">
                                <div class="th-status-with-toggle">
                                    <span>Статус</span>
                                    ${buildOprDatesToggle()}
                                </div>
                            </div>
                            ${showDates ? `<div class="th center" style="grid-column: 4 / span 2;">Даты</div>` : ''}
                        </div>
                        ${showDates ? `
                            <div class="thead-bottom ${gridClass}">
                                <div class="th"></div>
                                <div class="th"></div>
                                <div class="th"></div>
                                <div class="th">План</div>
                                <div class="th">Факт</div>
                            </div>
                        ` : ''}
                    </div>
                    <div>
                        ${visibleGroups.map(renderOprGroupUi).join('')}
                    </div>
                `;

                if (middleTableEl) middleTableEl.innerHTML = '';
                if (rightTableEl) rightTableEl.innerHTML = '';
            };

            const previousRenderAll = renderAll;
            renderAll = function () {
                previousRenderAll();
                updateLockNotice();
                focusActiveInlineEditor();
            };

            const previousBindEventsEnhanced = bindEvents;
            bindEvents = function () {
                previousBindEventsEnhanced();

                document.querySelectorAll('[data-role="concept-extra"], [data-role="concept-name-edit"], [data-role="concept-source-edit"]').forEach(el => {
                    el.dataset.initialValue = String(el.value || '');
                    el.dataset.baseHeight = '32';
                    el.style.height = '32px';
                    el.addEventListener('input', function () {
                        autoGrowTextarea(this);
                    });
                });

                document.querySelectorAll('[data-role="concept-name-display"]').forEach(el => {
                    const openEditor = function () {
                        startInlineEditor('concept-name', this.dataset.itemId);
                    };
                    el.addEventListener('click', openEditor);
                    el.addEventListener('keydown', function (e) {
                        if (e.key === 'Enter' || e.key === ' ') {
                            e.preventDefault();
                            openEditor.call(this);
                        }
                    });
                });

                document.querySelectorAll('[data-role="concept-source-display"]').forEach(el => {
                    const openEditor = function () {
                        startInlineEditor('concept-source', this.dataset.itemId);
                    };
                    el.addEventListener('click', openEditor);
                    el.addEventListener('keydown', function (e) {
                        if (e.key === 'Enter' || e.key === ' ') {
                            e.preventDefault();
                            openEditor.call(this);
                        }
                    });
                });

                document.querySelectorAll('[data-role="concept-name-edit"]').forEach(el => {
                    const commit = function () {
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;
                        const oldValue = item.name || '';
                        const newValue = String(this.value || '').trim();
                        item.name = newValue;
                        pushSessionChange(item.id, newValue || oldValue || item.id, 'name', oldValue, newValue);
                        stopInlineEditor();
                        renderAll();
                    };
                    const cancel = function () {
                        stopInlineEditor();
                        renderAll();
                    };
                    el.addEventListener('blur', commit);
                    el.addEventListener('keydown', function (e) {
                        if (e.key === 'Escape') {
                            e.preventDefault();
                            cancel();
                            return;
                        }
                        if (e.key === 'Enter' && !e.shiftKey) {
                            e.preventDefault();
                            commit.call(this);
                        }
                    });
                });

                document.querySelectorAll('[data-role="concept-source-edit"]').forEach(el => {
                    const commit = function () {
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;
                        const oldValue = item.source || '';
                        const newValue = String(this.value || '').trim();
                        item.source = newValue;
                        pushSessionChange(item.id, item.name || item.id, 'source', oldValue, newValue);
                        stopInlineEditor();
                        renderAll();
                    };
                    const cancel = function () {
                        stopInlineEditor();
                        renderAll();
                    };
                    el.addEventListener('blur', commit);
                    el.addEventListener('keydown', function (e) {
                        if (e.key === 'Escape') {
                            e.preventDefault();
                            cancel();
                            return;
                        }
                        if (e.key === 'Enter' && !e.shiftKey) {
                            e.preventDefault();
                            commit.call(this);
                        }
                    });
                });

                document.querySelectorAll('[data-role="opr-toggle-dates"]').forEach(btn => {
                    btn.addEventListener('click', function () {
                        oprDateVisibility[1] = !oprDateVisibility[1];
                        renderAll();
                    });
                });

                document.querySelectorAll('[data-role="opr-status"]').forEach(el => {
                    el.addEventListener('change', function() {
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;

                        const oldValue = item.status || '';
                        const newValue = this.value;

                        item.status = newValue;
                        if (newValue === 'Не требуется') {
                            item.group = 2;
                        } else if (Number(item.group) === 2) {
                            item.group = resolveOprGroupIdByItemIdOrName ? resolveOprGroupIdByItemIdOrName(item) : 1;
                            if (Number(item.group) === 2) {
                                item.group = 1;
                            }
                        }

                        pushSessionChange(item.id, item.name, 'status', oldValue, newValue);
                        renderAll();
                    });
                });

                document.querySelectorAll('[data-role="opr-plan"]').forEach(el => {
                    el.addEventListener('change', function() {
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;
                        const oldValue = item.plan || '';
                        const newValue = fromInputDate(this.value);
                        item.plan = newValue;
                        pushSessionChange(item.id, item.name, 'plan', oldValue, newValue);
                        renderAll();
                    });
                });

                document.querySelectorAll('[data-role="opr-fact"]').forEach(el => {
                    el.addEventListener('change', function() {
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;
                        const oldValue = item.fact || '';
                        const newValue = fromInputDate(this.value);
                        item.fact = newValue;
                        pushSessionChange(item.id, item.name, 'fact', oldValue, newValue);
                        renderAll();
                    });
                });

                document.querySelectorAll('[data-role="opr-add-item"]').forEach(btn => {
                    btn.addEventListener('click', async function() {
                        const groupId = Number(this.dataset.groupId);
                        const input = document.getElementById('oprAddItemInput_' + groupId);
                        if (!input) return;

                        const name = (input.value || '').trim();
                        if (!name) return;

                        this.disabled = true;

                        try {
                            const result = await addItem(groupId, name, 'opr');
                            if (!result || !result.item) {
                                throw new Error('add opr item failed');
                            }

                            replaceItem(result.item);
                            pushSessionChange(result.item.id, result.item.name, 'add-item', '', result.item.name);

                            debugLog('opr_item_added', {
                                itemId: result.item.id,
                                itemName: result.item.name,
                                groupId: groupId
                            });

                            input.value = '';
                            renderAll();
                        } catch (e) {
                            console.log('opr add item error:', e);
                            setSaveState('error', 'Ошибка добавления пункта');
                        } finally {
                            this.disabled = false;
                        }
                    });
                });
            };

            setTimeout(function () {
                acquireChecklistLock(currentChecklistKey, true);
                startLockHeartbeat();
                updateLockNotice();
            }, 0);
    """

    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>{full_title}</title>
        <script src="https://api.bitrix24.com/api/v1/"></script>
        <style>
            * {{ box-sizing: border-box; }}
            body {{ margin:0; font-family:Arial,sans-serif; background:#f3f6fb; color:#1f2328; }}
            .shell {{ padding:14px; }}
            .modal {{ background:#fff; border:1px solid #e5e7eb; border-radius:14px; overflow:hidden; box-shadow:0 16px 40px rgba(0,0,0,.12); }}
            .header {{ padding:14px 16px 12px; border-bottom:1px solid #edf0f2; display:flex; justify-content:space-between; align-items:flex-start; gap:14px; }}
            .title {{ font-size:22px; font-weight:700; line-height:1.2; }}
            .title small {{ font-size:20px; font-weight:600; color:#344054; }}
            .header-main {{ display:flex; align-items:flex-start; gap:18px; flex:1 1 auto; min-width:0; }}
            .header-right {{ display:flex; align-items:center; gap:14px; flex:0 0 auto; }}
            .progress-box {{ min-width:150px; }}
            .progress-label {{ font-size:12px; color:#667085; margin-bottom:4px; }}
            .progress-value {{ font-size:21px; font-weight:700; margin-bottom:5px; }}
            .progress-track {{ width:100%; height:8px; background:#edf2f7; border-radius:999px; overflow:hidden; }}
            .progress-bar {{ height:100%; width:0%; background:#22c55e; transition:width .2s ease; }}
            .save-state {{ font-size:12px; font-weight:700; padding:7px 10px; border-radius:999px; background:#eef2ff; color:#3730a3; white-space:nowrap; }}
            .progress-box.id-accent {{ min-width:172px; }}
            .progress-box.id-accent .progress-label {{ font-size:13px; }}
            .progress-box.id-accent .progress-value {{ font-size:24px; }}
            .progress-box.id-accent .progress-track {{ height:9px; }}
            .save-state.saving {{ background:#fff4e5; color:#b26a00; }}
            .save-state.error {{ background:#fdecec; color:#b42318; }}
            .content {{ padding:14px 16px 16px; max-height:82vh; overflow:auto; }}
            .layout {{ display:flex; flex-direction:column; gap:12px; align-items:stretch; }}
            .tables-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; align-items:start; }}
            .table-panel {{ min-width:0; display:flex; }}
            .table-panel .table {{ flex:1 1 auto; }}
            .side-panel {{ order:-1; border:1px solid #e5e7eb; border-radius:12px; background:#fff; overflow:hidden; position:static; }}
            .side-panel-title {{ padding:12px 14px; background:#fafbfc; border-bottom:1px solid #e5e7eb; font-size:13px; font-weight:700; color:#344054; }}
            .side-panel-list {{ padding:10px; display:flex; flex-wrap:wrap; gap:8px; }}
            .side-link {{ display:inline-flex; width:auto; text-align:left; border:1px solid #d0d7de; border-radius:8px; background:#fff; padding:9px 12px; font-size:13px; cursor:pointer; align-items:center; }}
            .side-link.active {{ background:#eef2ff; border-color:#c7d2fe; font-weight:700; }}
            .table {{ width:100%; border:1px solid #e5e7eb; border-radius:12px; overflow:hidden; background:#fff; }}
            .thead {{ position:sticky; top:0; z-index:10; background:#f8fafc; border-bottom:1px solid #e5e7eb; }}
            .thead-top,.thead-bottom {{ min-height:38px; }}
            .thead-top,.thead-bottom,.row {{ display:grid; grid-template-columns:190px 190px 100px 136px 136px; gap:0; align-items:stretch; justify-content:start; }}
            .th,.td {{ padding:8px 9px; border-right:1px solid #edf0f2; }}
            .th:last-child,.td:last-child {{ border-right:none; }}
            .th {{ font-size:12px; font-weight:700; color:#475467; min-height:38px; display:flex; align-items:center; }}
            .thead-top .th,.thead-bottom .th {{ min-height:38px; }}
            .th.center {{ text-align:center; justify-content:center; }}
            .group-block {{ border-top:8px solid #f8fafc; }}
            .group-title {{ padding:9px 12px; min-height:40px; background:#fafbfc; border-top:1px solid #e5e7eb; border-bottom:1px solid #e5e7eb; font-size:13px; font-weight:700; color:#344054; display:flex; align-items:center; }}
            .row {{ border-top:1px solid #edf0f2; background:#fff; }}
            .row.not-required {{ background:#fafafa; }}
            .row.not-required .item-name {{ text-decoration:line-through; color:#98a2b3; }}
            .cell-name {{ display:flex; align-items:center; gap:8px; min-width:0; }}
            .status-indicator {{ width:15px; height:15px; border-radius:999px; border:1px solid #d0d7de; flex:0 0 15px; background:#fff; }}
            .status-indicator.green {{ background:#22c55e; border-color:#22c55e; }}
            .status-indicator.gray {{ background:#9ca3af; border-color:#9ca3af; }}
            .item-name {{ font-size:13px; font-weight:700; color:#1f2328; line-height:1.15; min-width:0; word-break:break-word; max-width:165px; }}
            .status-select,.date-input {{ width:100%; border:1px solid #d0d7de; border-radius:8px; padding:6px 8px; font-size:12px; background:#fff; }}
            .concept-extra-textarea {{ width:100%; height:32px; min-height:32px; border:1px solid #d0d7de; border-radius:8px; padding:6px 8px; font-size:12px; background:#fff; resize:vertical; overflow:hidden; line-height:1.35; }}
            .concept-inline-display {{ min-height:32px; padding:6px 8px; border:1px solid transparent; border-radius:8px; font-size:12px; line-height:1.35; white-space:pre-wrap; word-break:break-word; cursor:text; }}
            .concept-inline-display:hover {{ background:#f8fafc; border-color:#e5e7eb; }}
            .concept-inline-display.empty {{ color:#98a2b3; }}
            .concept-inline-display.disabled {{ cursor:default; background:#f8fafc; border-color:transparent; }}
            .concept-inline-input {{ width:100%; min-height:32px; height:32px; border:1px solid #d0d7de; border-radius:8px; padding:6px 8px; font-size:12px; background:#fff; resize:vertical; overflow:hidden; line-height:1.35; }}
            .doc-btn,.upload-btn,.add-item-btn {{ display:inline-block; width:100%; text-align:center; padding:6px 8px; border:1px solid #d0d7de; border-radius:8px; background:#f8fafc; color:#1f2328; font-size:12px; text-decoration:none; cursor:pointer; }}
            .doc-btn:hover,.upload-btn:hover,.side-link:hover,.add-item-btn:hover {{ background:#f1f5f9; }}
            .doc-cell {{
                display: flex;
                flex-direction: column;
                gap: 6px;
            }}

            .doc-actions {{
                display: flex;
                gap: 6px;
            }}

            .doc-actions .upload-btn,
            .doc-actions .doc-btn {{
                flex: 1 1 0;
                width: auto;
            }}

            .doc-files {{
                display: flex;
                flex-direction: column;
                gap: 4px;
            }}

            .doc-file-row {{
                display: flex;
                align-items: center;
                gap: 6px;
                min-width: 0;
            }}

            .doc-file-link {{
                flex: 1 1 auto;
                min-width: 0;
                font-size: 12px;
                color: #175cd3;
                text-decoration: none;
                cursor: pointer;
                overflow: hidden;
                text-overflow: ellipsis;
                white-space: nowrap;
            }}

            .doc-file-link:hover {{
                text-decoration: underline;
            }}

            .doc-file-remove {{
                flex: 0 0 auto;
                border: none;
                background: transparent;
                color: #b42318;
                cursor: pointer;
                font-size: 14px;
                line-height: 1;
                padding: 0 2px;
                position: relative;
                z-index: 2;
                pointer-events: auto;
            }}

            .doc-file-meta {{
                flex: 0 0 auto;
                font-size: 11px;
                color: #667085;
                white-space: nowrap;
            }}
            
            .doc-file-remove:hover {{
                opacity: .8;
            }}
            .add-item-row {{ padding:9px 12px 10px; min-height:52px; border-top:1px solid #edf0f2; background:#fcfcfd; display:flex; gap:8px; align-items:center; justify-content:flex-start; }}
            .add-item-row::after {{ content:''; flex:1 1 auto; }}
            .add-item-input {{ flex:0 0 190px; width:190px; min-width:190px; height:32px; border:1px solid #d0d7de; border-radius:8px; padding:7px 9px; font-size:12px; }}
            .add-item-btn {{ flex:0 0 100px; width:100px; min-width:100px; height:32px; border:1px solid #d0d7de; border-radius:8px; padding:6px 8px; background:#f8fafc; cursor:pointer; font-size:12px; white-space:nowrap; }}
            @media (max-width:1320px) {{ .layout {{ grid-template-columns:1fr; }} .side-panel {{ position:static; }} }}
            @media (max-width:1120px) {{ .tables-grid {{ grid-template-columns:1fr; }} }}
            .tables-grid.id-three-cols {{
                grid-template-columns: minmax(0, 1fr) minmax(0, 1fr) minmax(0, 1fr);
            }}

            .id-table .thead {{
                position: sticky;
                top: 0;
                z-index: 10;
                background: #f8fafc;
                border-bottom: 1px solid #e5e7eb;
            }}

            .id-grid {{
                display: grid;
                gap: 0;
                align-items: stretch;
                justify-content: start;
            }}

            .id-grid.id-grid-compact {{
                grid-template-columns: minmax(0, 1.18fr) minmax(0, 1.02fr) 108px;
            }}

            .id-grid.id-grid-expanded {{
                grid-template-columns: minmax(0, 1.08fr) minmax(0, 0.96fr) 108px 112px 112px;
            }}

            .id-table .thead-top,
            .id-table .thead-bottom,
            .id-table .row {{
                min-height: 38px;
            }}

            .id-table .thead-bottom {{
                border-top: 1px solid #edf0f2;
            }}

            .th-status-with-toggle {{
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 6px;
            }}

            .id-dates-toggle {{
                width: 22px;
                min-width: 22px;
                height: 22px;
                border: 1px solid #d0d7de;
                background: #fff;
                color: #344054;
                border-radius: 6px;
                padding: 0;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                font-size: 12px;
                line-height: 1;
                cursor: pointer;
            }}

            .id-dates-toggle:hover {{
                background: #f8fafc;
            }}

            .id-table .item-name {{
                max-width: none;
            }}

            @media (max-width:980px) {{
                .thead-top,.thead-bottom,.row {{ grid-template-columns:1fr; }}
                .th,.td {{ border-right:none; border-bottom:1px solid #edf0f2; }}
                .th:last-child,.td:last-child {{ border-bottom:none; }}
            }}
        </style>
    </head>
    <body>
        <div class="shell">
            <div class="modal">
                <div class="header">
                    <div class="header-main">
                        <div class="title" id="popupTitle">Чек-лист ИД</div>
                        <div class="progress-box">
                            <div class="progress-label">Прогресс</div>
                            <div class="progress-value" id="progressValue">{progress_percent}%</div>
                            <div class="progress-track"><div class="progress-bar" id="progressBar"></div></div>
                        </div>
                    </div>
                    <div class="header-right">
                        <div id="saveState" class="save-state">Сохранено</div>
                    </div>
                </div>
                <div class="content">
                    <div id="debugPanel" style="
                        display:none;
                        margin-bottom:12px;
                        padding:10px 12px;
                        border:1px solid #e5e7eb;
                        border-radius:10px;
                        background:#fafbfc;
                        font-size:12px;
                        color:#344054;
                    ">
                        <div><b>Debug:</b> <span id="debugLastEvent">popup init</span></div>
                        <div style="margin-top:4px;">
                            <a id="debugLogsLink" href="debug/logs" target="_blank">Открыть /debug/logs</a>
                        </div>
                    </div>
                    <div class="layout">
                        <div class="tables-grid">
                            <div class="table-panel">
                                <div class="table">
                                    <div class="thead">
                                        <div class="thead-top">
                                            <div class="th">ИД</div>
                                            <div class="th">Документ</div>
                                            <div class="th">Статус</div>
                                            <div class="th center" style="grid-column: 4 / span 2;">Дата получения</div>
                                        </div>
                                        <div class="thead-bottom">
                                            <div class="th"></div>
                                            <div class="th"></div>
                                            <div class="th"></div>
                                            <div class="th">План</div>
                                            <div class="th">Факт</div>
                                        </div>
                                    </div>
                                    <div id="leftTableBody"></div>
                                </div>
                            </div>

                            <div class="table-panel">
                                <div class="table">
                                    <div class="thead">
                                        <div class="thead-top">
                                            <div class="th">ТУ</div>
                                            <div class="th">Документ</div>
                                            <div class="th">Статус</div>
                                            <div class="th center" style="grid-column: 4 / span 2;">Дата получения</div>
                                        </div>
                                        <div class="thead-bottom">
                                            <div class="th"></div>
                                            <div class="th"></div>
                                            <div class="th"></div>
                                            <div class="th">План</div>
                                            <div class="th">Факт</div>
                                        </div>
                                    </div>
                                    <div id="middleTableBody"></div>
                                </div>
                            </div>

                            <div class="table-panel">
                                <div class="table">
                                    <div class="thead">
                                        <div class="thead-top">
                                            <div class="th">Прочее</div>
                                            <div class="th">Документ</div>
                                            <div class="th">Статус</div>
                                            <div class="th center" style="grid-column: 4 / span 2;">Дата получения</div>
                                        </div>
                                        <div class="thead-bottom">
                                            <div class="th"></div>
                                            <div class="th"></div>
                                            <div class="th"></div>
                                            <div class="th">План</div>
                                            <div class="th">Факт</div>
                                        </div>
                                    </div>
                                    <div id="rightTableBody"></div>
                                </div>
                            </div>
                        </div>
                        <div class="side-panel">
                            <div class="side-panel-title">Список чек-листов по проекту</div>
                            <div class="side-panel-list" id="projectChecklistList"></div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <script>
            const dialogId = {dialog_id_json};
            let rawGroups = {groups_json};
            let rawProjectChecklists = {project_checklists_json};
            let rawItems = {items_json};
            let collabTitle = {collab_title_json};

            let groups = Array.isArray(rawGroups) ? rawGroups : [];
            let projectChecklists = Array.isArray(rawProjectChecklists) ? rawProjectChecklists : [];
            let items = Array.isArray(rawItems) ? rawItems : [];

            let currentChecklistKey = {checklist_key_json};
            let checklistTitle = {checklist_title_json};
            let checklistCache = {{}};
            let sessionChanges = [];
            let currentEditor = {{
                id: "",
                name: ""
            }};
                        const saveStateEl = document.getElementById('saveState');
            const leftTableBodyEl = document.getElementById('leftTableBody');
            const middleTableBodyEl = document.getElementById('middleTableBody');
            const rightTableBodyEl = document.getElementById('rightTableBody');
            const progressValueEl = document.getElementById('progressValue');
            const progressBarEl = document.getElementById('progressBar');
            const progressBoxEl = document.querySelector('.progress-box');
            const popupTitleEl = document.getElementById('popupTitle');
            const projectChecklistListEl = document.getElementById('projectChecklistList');
            const tablePanels = document.querySelectorAll('.table-panel');
            const tablesGridEl = document.querySelector('.tables-grid');
            const leftTableEl = tablePanels[0] ? tablePanels[0].querySelector('.table') : null;
            const middleTableEl = tablePanels[1] ? tablePanels[1].querySelector('.table') : null;
            const rightTableEl = tablePanels[2] ? tablePanels[2].querySelector('.table') : null;
            const idTableShellHtml = leftTableEl ? leftTableEl.innerHTML : '';
            const idDateVisibility = {{ 1: false, 2: false, 3: false }};
            const oprDateVisibility = {{ 1: false }};
            const debugLastEventEl = document.getElementById('debugLastEvent');
            const debugPanelEl = document.getElementById('debugPanel');
            const debugLogsLinkEl = document.getElementById('debugLogsLink');
            const allowedDebugUserIds = new Set(['138', '18']);
            function updateDebugPanelAccess() {{
                const currentUserId = String(currentEditor.id || '');
                if (debugPanelEl) {{
                    debugPanelEl.style.display = allowedDebugUserIds.has(currentUserId) ? '' : 'none';
                }}
                if (debugLogsLinkEl) {{
                    debugLogsLinkEl.href = 'debug/logs?userId=' + encodeURIComponent(currentUserId);
                }}
            }}
            function detectAppBasePath() {{
                const path = String(window.location.pathname || '/').replace(/\/+$/, '');
                const suffixes = ['/popup', '/launch', '/textarea', '/install', '/health', '/debug/logs', '/admin', '/admin/upload'];

                for (const suffix of suffixes) {{
                    if (path === suffix) return '';
                    if (path.endsWith(suffix)) {{
                        return path.slice(0, -suffix.length) || '';
                    }}
                }}

                return '';
            }}

            const APP_BASE_PATH = detectAppBasePath();
            const APP_BASE_URL = window.location.origin + (APP_BASE_PATH || '');

            function appUrl(path) {{
                return APP_BASE_URL + '/' + String(path || '').replace(/^\/+/, '');
            }}
            let closeSummarySent = false;
            let sessionDirty = false;

            function setSaveState(mode, text) {{
                saveStateEl.classList.remove('saving', 'error');
                if (mode === 'saving') saveStateEl.classList.add('saving');
                if (mode === 'error') saveStateEl.classList.add('error');
                saveStateEl.textContent = text;
            }}
            function esc(v) {{
                if (v === null || v === undefined) return '';
                return String(v).replaceAll('&', '&amp;').replaceAll('<', '&lt;').replaceAll('>', '&gt;').replaceAll('"', '&quot;');
            }}
            function toInputDate(value) {{
                if (!value) return '';
                const parts = value.split('.');
                if (parts.length !== 3) return '';
                return `${{parts[2]}}-${{parts[1]}}-${{parts[0]}}`;
            }}
            function fromInputDate(value) {{
                if (!value) return '';
                const parts = value.split('-');
                if (parts.length !== 3) return '';
                return `${{parts[2]}}.${{parts[1]}}.${{parts[0]}}`;
            }}
            function normalizeStatus(status) {{
                const s = String(status || '').trim();
                if (s === 'Есть') return 'Есть';
                if (s === 'Нет') return 'Нет';
                if (s === 'Не требуется') return 'Не требуется';
                return '';
            }}
            function indicatorClass(status) {{
                const s = normalizeStatus(status);
                if (s === 'Есть') return 'status-indicator green';
                if (s === 'Нет' || s === 'Не требуется') return 'status-indicator gray';
                return 'status-indicator';
            }}

            function getItemDocuments(item) {{
                const docs = Array.isArray(item && item.documents) ? item.documents : [];
                if (docs.length) {{
                    return docs;
                }}

                const legacyUrl = String(item && item.documentUrl || '').trim();
                const legacyName = String(item && item.documentName || '').trim();

                if (legacyUrl || legacyName) {{
                    return [{{
                        id: 'legacy_' + String(item && item.id || ''),
                        name: legacyName || 'Файл',
                        path: legacyUrl,
                        fileUrl: legacyUrl,
                        previewUrl: legacyUrl,
                        size: 0,
                        modifiedAt: '',
                        source: 'local'
                    }}];
                }}

                return [];
            }}

            function formatFileSize(size) {{
                const value = Number(size || 0);
                if (!value || value <= 0) return '';

                const units = ['Б', 'КБ', 'МБ', 'ГБ'];
                let current = value;
                let unitIndex = 0;

                while (current >= 1024 && unitIndex < units.length - 1) {{
                    current /= 1024;
                    unitIndex += 1;
                }}

                if (unitIndex === 0) {{
                    return Math.round(current) + ' ' + units[unitIndex];
                }}

                if (current >= 100) return current.toFixed(0) + ' ' + units[unitIndex];
                if (current >= 10) return current.toFixed(1) + ' ' + units[unitIndex];
                return current.toFixed(2) + ' ' + units[unitIndex];
            }}

            function confirmFileRemoval(fileName) {{
                const safeName = String(fileName || 'файл').trim() || 'файл';
                return window.confirm('Удалить файл "' + safeName + '"?');
            }}

            function renderTitle() {{
                if (collabTitle) {{
                    popupTitleEl.innerHTML = esc(checklistTitle) + ' <small>— ' + esc(collabTitle) + '</small>';
                }} else {{
                    popupTitleEl.textContent = checklistTitle;
                }}
            }}
            function fetchCurrentUserIfPossible() {{
                try {{
                    if (!(window.BX24 && typeof window.BX24.init === 'function')) {{
                        return;
                    }}

                    window.BX24.init(function () {{
                        try {{
                            window.BX24.callMethod('user.current', {{}}, function(result) {{
                                try {{
                                    if (result.error()) {{
                                        return;
                                    }}

                                    const data = result.data() || {{}};
                                    const fullName = [data.NAME, data.LAST_NAME].filter(Boolean).join(' ').trim();

                                    currentEditor = {{
                                        id: String(data.ID || ''),
                                        name: fullName || String(data.NAME || '') || ''
                                    }};
                                    updateDebugPanelAccess();
                                }} catch (e) {{
                                    console.log('user.current parse error:', e);
                                }}
                            }});
                        }} catch (e) {{
                            console.log('user.current call error:', e);
                        }}
                    }});
                }} catch (e) {{
                    console.log('fetchCurrentUserIfPossible skipped:', e);
                }}
            }}
            function setDebugText(text) {{
                if (debugLastEventEl) {{
                    debugLastEventEl.textContent = text;
                }}
            }}
            function debugLog(event, payload = {{}}, useBeacon = false) {{
                const body = JSON.stringify({{
                    event,
                    dialogId,
                    checklistKey: currentChecklistKey,
                    payload,
                    href: window.location.href,
                    ts: new Date().toISOString()
                }});

                setDebugText(event);

                try {{
                    const url = APP_BASE_URL + '/api/debug/event';

                    if (useBeacon && navigator.sendBeacon) {{
                        const blob = new Blob([body], {{ type: 'application/json' }});
                        const ok = navigator.sendBeacon(url, blob);
                        setDebugText(event + ' | beacon=' + ok);
                        return;
                    }}

                    fetch(url, {{
                        method: 'POST',
                        headers: {{
                            'Content-Type': 'application/json'
                        }},
                        body
                    }})
                    .then(r => {{
                        setDebugText(event + ' | http=' + r.status);
                    }})
                    .catch(err => {{
                        console.log('debugLog fetch error:', err);
                        setDebugText(event + ' | fetch error');
                    }});
                }} catch (e) {{
                    console.log('debugLog error:', e);
                    setDebugText(event + ' | js error');
                }}
            }}
            function logRenderState(stage) {{
                debugLog('render_state', {{
                    stage,
                    groupsType: typeof rawGroups,
                    itemsType: typeof rawItems,
                    projectChecklistsType: typeof rawProjectChecklists,
                    groupsIsArray: Array.isArray(groups),
                    itemsIsArray: Array.isArray(items),
                    projectChecklistsIsArray: Array.isArray(projectChecklists),
                    groupsLength: groups.length,
                    itemsLength: items.length,
                    projectChecklistsLength: projectChecklists.length,
                    leftTableExists: !!leftTableBodyEl,
                    rightTableExists: !!rightTableBodyEl,
                    titleExists: !!popupTitleEl,
                    sidePanelExists: !!projectChecklistListEl
                }});
            }}

            function logRenderError(stage, error) {{
                const message = (error && error.message) ? error.message : String(error || 'unknown error');
                const stack = (error && error.stack) ? error.stack : '';

                setDebugText(stage + ' | ERROR: ' + message);

                debugLog('render_error', {{
                    stage,
                    message,
                    stack
                }});
            }}
            function deepClone(value) {{
                return JSON.parse(JSON.stringify(value));
            }}

            function buildChecklistSnapshot() {{
                return {{
                    checklistKey: currentChecklistKey,
                    title: checklistTitle,
                    collabTitle,
                    groups: deepClone(groups),
                    projectChecklists: deepClone(projectChecklists),
                    items: deepClone(items)
                }};
            }}

            function syncChecklistCache() {{
                checklistCache[currentChecklistKey] = buildChecklistSnapshot();
            }}

            function applyChecklistData(data) {{
                const nextData = data || {{}};
                const nextKey = String(nextData.checklistKey || currentChecklistKey || 'id').trim() || 'id';

                currentChecklistKey = nextKey;
                checklistTitle = String(nextData.title || getChecklistDefaultTitle(nextKey));
                collabTitle = String(nextData.collabTitle || collabTitle || '');

                rawGroups = Array.isArray(nextData.groups) ? nextData.groups : [];
                rawProjectChecklists = Array.isArray(nextData.projectChecklists)
                    ? nextData.projectChecklists
                    : rawProjectChecklists;
                rawItems = Array.isArray(nextData.items) ? nextData.items : [];

                groups = Array.isArray(rawGroups) ? rawGroups : [];
                projectChecklists = Array.isArray(rawProjectChecklists) ? rawProjectChecklists : [];
                items = Array.isArray(rawItems) ? rawItems : [];

                document.title = collabTitle
                    ? checklistTitle + ' — ' + collabTitle
                    : checklistTitle;
            }}

            async function flushCurrentChecklistSummary(reason = 'checklist_switch') {{
                syncChecklistCache();
                debugLog('close_summary_switch_skipped', {{
                    checklistKey: currentChecklistKey,
                    reason,
                    changesCount: sessionChanges.length,
                    dirty: !!sessionDirty
                }});
            }}

            async function loadChecklistByKey(checklistKey) {{
                const targetKey = String(checklistKey || '').trim() || 'id';
                if (targetKey === currentChecklistKey) {{
                    return;
                }}

                syncChecklistCache();
                await releaseChecklistLock(currentChecklistKey, false);
                setSaveState('saving', 'Загружаем...');

                try {{
                    const cachedData = checklistCache[targetKey];
                    if (cachedData) {{
                        applyChecklistData(deepClone(cachedData));
                        renderAll();
                        debugLog('checklist_switched_cached', {{
                            checklistKey: targetKey
                        }});
                        await acquireChecklistLock(targetKey, true);
                        startLockHeartbeat();
                        setSaveState('', 'Сохранено');
                        return;
                    }}

                    const response = await fetch(
                        appUrl('api/checklist') +
                        '?dialogId=' + encodeURIComponent(dialogId) +
                        '&checklistKey=' + encodeURIComponent(targetKey)
                    );
                    const result = await response.json();

                    if (!response.ok) {{
                        throw new Error(result.error || 'load checklist failed');
                    }}

                    applyChecklistData(result);
                    renderAll();
                    debugLog('checklist_switched', {{
                        checklistKey: targetKey
                    }});
                    await acquireChecklistLock(targetKey, true);
                    startLockHeartbeat();
                    setSaveState('', 'Сохранено');
                }} catch (e) {{
                    console.log('loadChecklistByKey error:', e);
                    setSaveState('error', 'Ошибка загрузки чек-листа');
                }}
            }}

            async function reloadCurrentChecklistFromServer() {{
                const response = await fetch(
                    appUrl('api/checklist') +
                    '?dialogId=' + encodeURIComponent(dialogId) +
                    '&checklistKey=' + encodeURIComponent(currentChecklistKey)
                );
                const result = await response.json();

                if (!response.ok) {{
                    throw new Error(result.error || 'reload checklist failed');
                }}

                applyChecklistData(result);
                return result;
            }}

            function sendCloseSummaryOnce(eventName) {{
                if (eventName === 'popup_hidden' || suppressAutoCloseSave || closeSummarySent) {{
                    return;
                }}

                closeSummarySent = true;
                persistDirtyChecklists(eventName, true);
                releaseChecklistLock(currentChecklistKey, true);
            }}

            window.addEventListener('message', async function (event) {{
                const data = event && event.data ? event.data : {{}};
                if (!data || data.type !== 'checklist-document-removed') return;
                if (String(data.dialogId || '') !== String(dialogId || '')) return;
                if (String(data.checklistKey || '') !== String(currentChecklistKey || '')) return;

                try {{
                    const response = await fetch(
                        appUrl('api/checklist') +
                        '?dialogId=' + encodeURIComponent(dialogId) +
                        '&checklistKey=' + encodeURIComponent(currentChecklistKey)
                    );
                    const result = await response.json();
                    if (!response.ok) throw new Error(result.error || 'reload after folder remove failed');

                    applyChecklistData(result);
                    renderAll();
                }} catch (e) {{
                    console.log('folder remove sync error:', e);
                }}
            }});

            document.addEventListener('visibilitychange', function () {{
                if (document.visibilityState === 'hidden') {{
                    sendCloseSummaryOnce('popup_hidden');
                }}
            }});

            window.addEventListener('pagehide', function () {{
                sendCloseSummaryOnce('popup_pagehide');
            }});

            window.addEventListener('beforeunload', function () {{
                sendCloseSummaryOnce('popup_beforeunload');
            }});
            async function fetchChatTitleIfMissing() {{
                if (collabTitle) {{ renderTitle(); return; }}
                try {{
                    if (!(window.BX24 && typeof window.BX24.init === 'function')) {{ renderTitle(); return; }}
                    window.BX24.init(function () {{
                        try {{
                            window.BX24.callMethod('im.dialog.get', {{ dialog_id: dialogId }}, async function(result) {{
                                try {{
                                    if (result.error()) {{ renderTitle(); return; }}
                                    const data = result.data() || {{}};
                                    let title = data.title || data.name || (data.dialog && (data.dialog.title || data.dialog.name)) || (data.chat && (data.chat.title || data.chat.name)) || '';
                                    title = String(title || '').trim();
                                    if (!title) {{ renderTitle(); return; }}
                                    collabTitle = title;
                                    renderTitle();
                                    debugLog('chat_title_loaded', {{
                                        title: title
                                    }});
                                    try {{
                                        await fetch(appUrl('api/checklist/update-meta'), {{
                                            method: 'POST',
                                            headers: {{ 'Content-Type': 'application/json' }},
                                            body: JSON.stringify({{ dialogId, checklistKey: currentChecklistKey, field: 'collabTitle', value: title }})
                                        }});
                                    }} catch (e) {{
                                        console.log('save collabTitle error:', e);
                                    }}
                                }} catch (e) {{
                                    console.log('im.dialog.get parse error:', e);
                                    renderTitle();
                                }}
                            }});
                        }} catch (e) {{
                            console.log('im.dialog.get call error:', e);
                            renderTitle();
                        }}
                    }});
                }} catch (e) {{
                    console.log('BX24 init for title skipped:', e);
                    renderTitle();
                }}
            }}
            function calculateProgress() {{
                if (!progressValueEl || !progressBarEl) {{
                    return;
                }}
                if (currentChecklistKey === 'concept') {{
                    const activeItems = items.filter(item => String(item.status || '').trim() !== 'Не требуется');
                    const completedItems = activeItems.filter(item => {{
                        const status = String(item.status || '').trim();
                        const kind = String(item.statusKind || '').trim() || 'bool';

                        if (kind === 'bool') {{
                            return status === 'Да';
                        }}

                        return !!status;
                    }});
                    const activeCount = activeItems.length;
                    const completedCount = completedItems.length;
                    const percent = activeCount ? Math.round((completedCount / activeCount) * 100) : 0;
                    progressValueEl.textContent = percent + '%';
                    progressBarEl.style.width = percent + '%';
                    return;
                }}
                const activeItems = items.filter(x => normalizeStatus(x.status) !== 'Не требуется');
                const completedItems = activeItems.filter(x => normalizeStatus(x.status) === 'Есть');
                const activeCount = activeItems.length;
                const completedCount = completedItems.length;
                const percent = activeCount ? Math.round((completedCount / activeCount) * 100) : 0;
                progressValueEl.textContent = percent + '%';
                progressBarEl.style.width = percent + '%';
            }}
            async function updateItem(itemId, field, value, checklistKey = currentChecklistKey) {{
                setSaveState('saving', 'Сохраняем...');
                const response = await fetch(appUrl('api/checklist/update-item'), {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{ dialogId, checklistKey, itemId, field, value }})
                }});
                const result = await response.json();
                if (!response.ok || !result.ok) throw new Error(result.error || 'save failed');
                setSaveState('', 'Сохранено');
                return result;
            }}
            async function addItem(groupId, name, checklistKey = currentChecklistKey) {{
                setSaveState('saving', 'Сохраняем...');
                const response = await fetch(appUrl('api/checklist/add-item'), {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{ dialogId, checklistKey, groupId, name }})
                }});
                const result = await response.json();
                if (!response.ok || !result.ok) throw new Error(result.error || 'add item failed');
                setSaveState('', 'Сохранено');
                return result;
            }}

            async function removeDocument(itemId, documentId = '') {{
                setSaveState('saving', 'Сохраняем...');
                const response = await fetch(appUrl('api/checklist/remove-document'), {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{
                        dialogId,
                        checklistKey: currentChecklistKey,
                        itemId,
                        documentId
                    }})
                }});
                const result = await response.json();
                if (!response.ok || !result.ok) throw new Error(result.error || 'remove document failed');
                setSaveState('', 'Сохранено');
                return result;
            }}

            async function uploadDocument(itemId, file) {{
                setSaveState('saving', 'Сохраняем...');
                const item = items.find(x => x.id === itemId);
                const formData = new FormData();
                formData.append('dialogId', dialogId);
                formData.append('itemId', itemId);
                formData.append('file', file);
                formData.append('checklistKey', currentChecklistKey);
                formData.append('itemGroup', String(item && item.group ? item.group : ''));
                const response = await fetch(appUrl('api/checklist/upload-document'), {{ method: 'POST', body: formData }})
                const result = await response.json();
                if (!response.ok || !result.ok) throw new Error(result.error || 'upload document failed');
                setSaveState('', 'Сохранено');
                return result;
            }}
            function getItemsByGroup(groupId) {{
                return items
                    .filter(item => Number(item.group) === Number(groupId))
                    .sort((a, b) => Number(a.order || 0) - Number(b.order || 0));
            }}
            function hasItemsInGroup(groupId) {{
                return getItemsByGroup(groupId).length > 0;
            }}
            function renderProjectChecklistList() {{
                if (!projectChecklistListEl) {{
                    return;
                }}

                if (!Array.isArray(projectChecklists) || !projectChecklists.length) {{
                    projectChecklistListEl.innerHTML = '';
                    return;
                }}

                projectChecklistListEl.innerHTML = projectChecklists.map(item => {{
                    const active = item.key === currentChecklistKey ? 'side-link active' : 'side-link';
                    return `<button type="button" class="${{active}}" data-checklist-key="${{esc(item.key)}}">${{esc(item.title)}}</button>`;
                }}).join('');
                projectChecklistListEl.querySelectorAll('[data-checklist-key]').forEach(btn => {{
                    btn.addEventListener('click', async function () {{
                        const key = this.dataset.checklistKey;
                        await loadChecklistByKey(key);
                    }});
                }});
            }}
            function buildDocumentCell(item) {{
                if (normalizeStatus(item && item.status) === 'Не требуется') {{
                    return '';
                }}

                const documents = getItemDocuments(item);
                const itemId = String(item && item.id || '');
                const folderViewUrl = String(item.folderUrl || '').trim() || (documents.length ? (
                    appUrl('api/checklist/folder') +
                    '?dialogId=' + encodeURIComponent(dialogId) +
                    '&checklistKey=' + encodeURIComponent(currentChecklistKey) +
                    '&itemId=' + encodeURIComponent(itemId)
                ) : '');
                const showViewFolder = documents.length > 0 && !!folderViewUrl;

                const filesHtml = documents.map(doc => {{
                    const docId = String(doc.id || '');
                    const docName = String(doc.name || 'Файл');
                    const openUrl = appUrl('api/checklist/file') +
                        '?dialogId=' + encodeURIComponent(dialogId) +
                        '&checklistKey=' + encodeURIComponent(currentChecklistKey) +
                        '&itemId=' + encodeURIComponent(itemId) +
                        '&documentId=' + encodeURIComponent(docId);
                    const sizeText = formatFileSize(doc.size || 0);

                    return `
                        <div class="doc-file-row">
                            <a
                                href="javascript:void(0)"
                                class="doc-file-link"
                                data-role="view-file"
                                data-item-id="${{esc(itemId)}}"
                                data-document-id="${{esc(docId)}}"
                                data-open-url="${{esc(openUrl)}}"
                                title="${{esc(docName)}}"
                            >
                                ${{esc(docName)}}
                            </a>
                            <button
                                type="button"
                                class="doc-file-remove"
                                data-role="remove-file"
                                data-item-id="${{esc(itemId)}}"
                                data-document-id="${{esc(docId)}}"
                                data-document-name="${{esc(docName)}}"
                                title="Удалить файл"
                                ${{typeof disabledAttr === 'function' ? disabledAttr() : ''}}
                            >
                                ×
                            </button>
                            ${{sizeText ? `<span class="doc-file-meta">${{esc(sizeText)}}</span>` : ''}}
                        </div>
                    `;
                }}).join('');

                return `
                    <div class="doc-cell">
                        <div class="doc-actions">
                            <button
                                class="upload-btn"
                                type="button"
                                data-role="upload"
                                data-item-id="${{esc(itemId)}}"
                                ${{typeof disabledAttr === 'function' ? disabledAttr() : ''}}
                            >
                                Загрузить
                            </button>

                            ${{showViewFolder ? `
                                <button
                                    class="doc-btn"
                                    type="button"
                                    data-role="view-folder"
                                    data-item-id="${{esc(itemId)}}"
                                    data-folder-url="${{esc(folderViewUrl)}}"
                                >
                                    Посмотреть
                                </button>
                            ` : ''}}
                        </div>

                        ${{documents.length ? `
                            <div class="doc-files">
                                ${{filesHtml}}
                            </div>
                        ` : ''}}

                        <input
                            type="file"
                            data-role="file-input"
                            data-item-id="${{esc(itemId)}}"
                            style="display:none;"
                            multiple
                            ${{typeof disabledAttr === 'function' ? disabledAttr() : ''}}
                        >
                    </div>
                `;
            }}
            function pushSessionChange(itemId, itemName, field, oldValue, newValue) {{
                if (String(oldValue || '') === String(newValue || '')) {{
                    return;
                }}

                sessionChanges.push({{
                    field,
                    itemId: itemId || '',
                    itemName: itemName || '',
                    oldValue: oldValue || '',
                    newValue: newValue || ''
                }});
                sessionDirty = true;
                closeSummarySent = false;
                setSaveState('saving', 'Есть несохраненные изменения');
            }}
            function autoGrowTextarea(el) {{
                if (!el) return;

                const baseHeight = Number(el.dataset.baseHeight || 0) || 32;
                const initialValue = String(el.dataset.initialValue || '');
                el.dataset.baseHeight = String(baseHeight);
                el.style.height = baseHeight + 'px';

                if (String(el.value || '').length <= initialValue.length) {{
                    return;
                }}

                let nextHeight = baseHeight;
                while (el.scrollHeight > el.clientHeight && nextHeight < 1600) {{
                    nextHeight *= 2;
                    el.style.height = nextHeight + 'px';
                }}
            }}
            function extractConceptUnit(placeholder) {{
                const raw = String(placeholder || '').trim();
                const match = raw.match(new RegExp('^_+\\s*(.+)$'));
                return match ? match[1].trim() : '';
            }}
            function formatConceptTextStatus(value, placeholder) {{
                const rawValue = String(value || '').trim();
                if (!rawValue) {{
                    return '';
                }}

                if (rawValue === 'Не требуется') {{
                    return rawValue;
                }}

                const unit = extractConceptUnit(placeholder);
                if (!unit) {{
                    return rawValue;
                }}

                if (rawValue.toLowerCase().endsWith(unit.toLowerCase())) {{
                    return rawValue;
                }}

                if (!/^[0-9]+([.,][0-9]+)?$/.test(rawValue)) {{
                    return rawValue;
                }}

                return unit.startsWith('%') ? rawValue + unit : rawValue + ' ' + unit;
            }}
            function buildClientItemId(prefix) {{
                return prefix + '_' + Date.now() + '_' + Math.random().toString(36).slice(2, 8);
            }}
            function createLocalItem(groupId, name, checklistKey) {{
                const groupItems = getItemsByGroup(groupId);
                const nextOrder = groupItems.length + 1;

                if (checklistKey === 'concept') {{
                    return {{
                        id: buildClientItemId('concept_g' + groupId + '_custom'),
                        group: groupId,
                        order: nextOrder,
                        name,
                        source: '',
                        statusKind: 'text',
                        statusOptions: [],
                        statusPlaceholder: '',
                        status: '',
                        extraInfo: '',
                        extraInfoPlaceholder: '',
                        documentUrl: '',
                        documentName: '',
                        isCustom: true,
                        priority: 'white'
                    }};
                }}

                return {{
                    id: buildClientItemId('item_g' + groupId + '_custom'),
                    group: groupId,
                    order: nextOrder,
                    name,
                    priority: 'white',
                    status: '',
                    plan: '',
                    fact: '',
                    documentUrl: '',
                    documentName: '',
                    isCustom: true
                }};
            }}
            function resolveConceptGroupId(item) {{
                const itemId = String(item && item.id || '');
                const name = String(item && item.name || '').trim();

                if (itemId.startsWith('concept_g')) {{
                    const match = itemId.match(new RegExp('^concept_g(\\d+)_'));
                    if (match) {{
                        const groupId = Number(match[1]);
                        if (groupId && groupId !== 10) {{
                            return groupId;
                        }}
                    }}
                }}

                const byName = groups.find(group => Number(group.id) !== 10 && Array.isArray(items) && items.some(existing =>
                    existing !== item &&
                    Number(existing.group) === Number(group.id) &&
                    String(existing.name || '').trim() === name
                ));
                if (byName) {{
                    return Number(byName.id);
                }}

                return 1;
            }}
            function conceptIndicatorClass(item) {{
                const status = String(item.status || '').trim();
                const kind = String(item.statusKind || '').trim();

                if (status === 'Не требуется') {{
                    return 'status-indicator gray';
                }}

                if (kind === 'bool') {{
                    if (status === 'Да') return 'status-indicator green';
                    if (status === 'Нет') return 'status-indicator gray';
                    return 'status-indicator';
                }}

                return status ? 'status-indicator green' : 'status-indicator';
            }}
            function buildConceptStatusCell(item) {{
                if (item.statusKind === 'bool') {{
                    return `
                        <select class="status-select" data-role="concept-status" data-item-id="${{esc(item.id)}}">
                            <option value="" ${{item.status === '' ? 'selected' : ''}}></option>
                            <option value="Да" ${{item.status === 'Да' ? 'selected' : ''}}>Да</option>
                            <option value="Нет" ${{item.status === 'Нет' ? 'selected' : ''}}>Нет</option>
                            <option value="Не требуется" ${{item.status === 'Не требуется' ? 'selected' : ''}}>Не требуется</option>
                        </select>
                    `;
                }}
                if (item.statusKind === 'select') {{
                    const options = [''].concat(item.statusOptions || [], ['Не требуется']);
                    return `
                        <select class="status-select" data-role="concept-status" data-item-id="${{esc(item.id)}}">
                            ${{options.map(option => `<option value="${{esc(option)}}" ${{item.status === option ? 'selected' : ''}}>${{esc(option)}}</option>`).join('')}}
                        </select>
                    `;
                }}
                return `<input class="status-select" type="text" data-role="concept-status" data-item-id="${{esc(item.id)}}" placeholder="${{esc(item.statusPlaceholder || '')}}" value="${{esc(item.status || '')}}">`;
            }}
            function buildConceptExtraCell(item) {{
                return `<textarea class="concept-extra-textarea" data-role="concept-extra" data-item-id="${{esc(item.id)}}" placeholder="${{esc(item.extraInfoPlaceholder || '')}}">${{esc(item.extraInfo || '')}}</textarea>`;
            }}
            function renderConceptGroup(group) {{
                const groupItems = getItemsByGroup(group.id);
                const allowAdd = Number(group.id) !== 10;
                const rows = groupItems.map(item => `
                    <div class="row" style="grid-template-columns: 1.05fr 0.66fr 190px 150px 1.15fr;" data-item-id="${{esc(item.id)}}">
                        <div class="td">
                            <div class="cell-name">
                                <div class="${{conceptIndicatorClass(item)}}"></div>
                                <div class="item-name" style="${{item.status === 'Не требуется' ? 'text-decoration:line-through;color:#98a2b3;' : ''}}">
                                    ${{esc(item.name)}}
                                </div>
                            </div>
                        </div>
                        <div class="td">${{esc(item.source || '')}}</div>
                        <div class="td">${{buildDocumentCell(item)}}</div>
                        <div class="td">${{buildConceptStatusCell(item)}}</div>
                        <div class="td">${{buildConceptExtraCell(item)}}</div>
                    </div>
                `).join('');

                const addBlock = allowAdd ? `
                    <div class="add-item-row">
                        <input class="add-item-input" id="conceptAddItemInput_${{group.id}}" type="text" placeholder="Новый пункт">
                        <button class="add-item-btn" type="button" data-role="concept-add-item" data-group-id="${{group.id}}">Добавить пункт</button>
                    </div>
                ` : '';

                return `<div class="group-block"><div class="group-title">${{esc(group.title)}}</div>${{rows}}${{addBlock}}</div>`;
            }}
            function buildConceptTableHtml(conceptGroups) {{
                return `
                    <div class="thead">
                        <div class="thead-top" style="grid-template-columns: 1.05fr 0.66fr 190px 150px 1.15fr;">
                            <div class="th">Пункт</div>
                            <div class="th">Нормативы</div>
                            <div class="th">Документ</div>
                            <div class="th">Статус</div>
                            <div class="th">Доп информация</div>
                        </div>
                    </div>
                    <div>${{conceptGroups.map(renderConceptGroup).join('')}}</div>
                `;
            }}
            function renderConceptTable() {{
                if (!leftTableEl || !rightTableEl || !tablesGridEl) {{
                    throw new Error('concept table containers not found');
                }}

                tablesGridEl.style.gridTemplateColumns = '1fr 1fr';
                tablesGridEl.style.justifyContent = '';

                if (tablePanels[0]) {{
                    tablePanels[0].style.display = '';
                    tablePanels[0].style.flex = '';
                    tablePanels[0].style.maxWidth = '';
                }}
                if (tablePanels[1]) {{
                    tablePanels[1].style.display = '';
                    tablePanels[1].style.flex = '';
                    tablePanels[1].style.maxWidth = '';
                }}
                if (tablePanels[2]) {{
                    tablePanels[2].style.display = 'none';
                    tablePanels[2].style.flex = '';
                    tablePanels[2].style.maxWidth = '';
                }}

                leftTableEl.style.width = '';
                leftTableEl.style.maxWidth = '';
                rightTableEl.style.width = '';
                rightTableEl.style.maxWidth = '';

                const visibleGroups = groups.filter(group => {{
                    if (Number(group.id) !== 10) return true;
                    return items.some(x => Number(x.group) === 10);
                }});

                const leftGroups = visibleGroups.filter(group => [1, 3, 5, 7, 9].includes(Number(group.id)));
                const rightGroups = visibleGroups.filter(group => [2, 4, 6, 8, 10].includes(Number(group.id)));

                leftTableEl.innerHTML = buildConceptTableHtml(leftGroups);
                rightTableEl.innerHTML = buildConceptTableHtml(rightGroups);
            }}

            function isIdChecklist() {{
                return currentChecklistKey === 'id';
            }}

            function isIdDatesVisible(groupId) {{
                return !!idDateVisibility[Number(groupId)];
            }}

            function getIdGridClass(showDates) {{
                return showDates ? 'id-grid id-grid-expanded' : 'id-grid id-grid-compact';
            }}

            function buildIdHeader(group, showDates) {{
                const groupId = Number(group.id);
                const toggleTitle = showDates ? 'Скрыть даты' : 'Показать даты';
                const toggleIcon = '📅';

                return `
                    <div class="thead-top ${{getIdGridClass(showDates)}}">
                        <div class="th">${{esc(group.title)}}</div>
                        <div class="th">Документ</div>
                        <div class="th th-status-with-toggle">
                            <span>Статус</span>
                            <button
                                type="button"
                                class="id-dates-toggle"
                                data-role="toggle-id-dates"
                                data-group-id="${{esc(groupId)}}"
                                title="${{esc(toggleTitle)}}"
                                aria-label="${{esc(toggleTitle)}}"
                            >
                                ${{toggleIcon}}
                            </button>
                        </div>
                        ${{showDates ? `<div class="th center" style="grid-column: 4 / span 2;">Дата получения</div>` : ''}}
                    </div>
                    ${{showDates ? `
                        <div class="thead-bottom ${{getIdGridClass(showDates)}}">
                            <div class="th"></div>
                            <div class="th"></div>
                            <div class="th"></div>
                            <div class="th">План</div>
                            <div class="th">Факт</div>
                        </div>
                    ` : ''}}
                `;
            }}

            function renderIdGroup(group, showDates) {{
                const groupItems = getItemsByGroup(group.id);
                const allowAdd = Number(group.id) !== 4;
                const gridClass = getIdGridClass(showDates);

                const rows = groupItems.map(item => {{
                    const rowClass = normalizeStatus(item.status) === 'Не требуется' ? 'row not-required' : 'row';

                    return `
                        <div class="${{rowClass}} ${{gridClass}}" data-item-id="${{esc(item.id)}}">
                            <div class="td">
                                <div class="cell-name">
                                    <div class="${{indicatorClass(item.status)}}"></div>
                                    <div class="item-name">${{esc(item.name)}}</div>
                                </div>
                            </div>
                            <div class="td">${{buildDocumentCell(item)}}</div>
                            <div class="td">
                                <select class="status-select" data-role="status" data-item-id="${{esc(item.id)}}" ${{disabledAttr()}}>
                                    <option value="" ${{normalizeStatus(item.status) === '' ? 'selected' : ''}}></option>
                                    <option value="Есть" ${{normalizeStatus(item.status) === 'Есть' ? 'selected' : ''}}>Есть</option>
                                    <option value="Нет" ${{normalizeStatus(item.status) === 'Нет' ? 'selected' : ''}}>Нет</option>
                                    <option value="Не требуется" ${{normalizeStatus(item.status) === 'Не требуется' ? 'selected' : ''}}>Не требуется</option>
                                </select>
                            </div>
                            ${{showDates ? `
                                <div class="td">
                                    <input class="date-input" type="date" data-role="plan" data-item-id="${{esc(item.id)}}" value="${{esc(toInputDate(item.plan))}}" ${{disabledAttr()}}>
                                </div>
                                <div class="td">
                                    <input class="date-input" type="date" data-role="fact" data-item-id="${{esc(item.id)}}" value="${{esc(toInputDate(item.fact))}}" ${{disabledAttr()}}>
                                </div>
                            ` : ''}}
                        </div>
                    `;
                }}).join('');

                const addBlock = allowAdd ? `
                    <div class="add-item-row">
                        <input class="add-item-input" id="addItemInput_${{group.id}}" type="text" placeholder="Новый пункт" ${{disabledAttr()}}>
                        <button class="add-item-btn" type="button" data-role="add-item" data-group-id="${{group.id}}" ${{disabledAttr()}}>Добавить пункт</button>
                    </div>
                ` : '';

                return `<div class="group-block"><div class="group-title">${{esc(group.title)}}</div>${{rows}}${{addBlock}}</div>`;
            }}

            function renderIdPanel(mainGroup, appendNotRequired = false) {{
                const showDates = isIdDatesVisible(mainGroup.id);
                const notRequiredGroup = appendNotRequired
                    ? groups.find(g => Number(g.id) === 4)
                    : null;

                const panelHtml = `
                    <div class="table id-table">
                        <div class="thead">
                            ${{buildIdHeader(mainGroup, showDates)}}
                        </div>
                        <div>
                            ${{renderIdGroup(mainGroup, showDates)}}
                            ${{appendNotRequired && notRequiredGroup && hasItemsInGroup(4) ? renderIdGroup(notRequiredGroup, false) : ''}}
                        </div>
                    </div>
                `;

                return panelHtml;
            }}

            function renderIdTables() {{
                if (!leftTableEl || !middleTableEl || !rightTableEl || !tablesGridEl) {{
                    throw new Error('id table containers not found');
                }}

                const idGroup = groups.find(g => Number(g.id) === 1) || {{ id: 1, title: 'ИД' }};
                const tuGroup = groups.find(g => Number(g.id) === 2) || {{ id: 2, title: 'ТУ' }};
                const otherGroup = groups.find(g => Number(g.id) === 3) || {{ id: 3, title: 'Прочее' }};

                tablesGridEl.classList.add('id-three-cols');
                tablesGridEl.style.gridTemplateColumns = 'minmax(0, 1fr) minmax(0, 1fr) minmax(0, 1fr)';
                tablesGridEl.style.justifyContent = '';

                if (tablePanels[0]) {{
                    tablePanels[0].style.display = '';
                    tablePanels[0].style.flex = '';
                    tablePanels[0].style.maxWidth = '';
                }}
                if (tablePanels[1]) {{
                    tablePanels[1].style.display = '';
                    tablePanels[1].style.flex = '';
                    tablePanels[1].style.maxWidth = '';
                }}
                if (tablePanels[2]) {{
                    tablePanels[2].style.display = '';
                    tablePanels[2].style.flex = '';
                    tablePanels[2].style.maxWidth = '';
                }}

                leftTableEl.style.width = '';
                leftTableEl.style.maxWidth = '';
                middleTableEl.style.width = '';
                middleTableEl.style.maxWidth = '';
                rightTableEl.style.width = '';
                rightTableEl.style.maxWidth = '';

                leftTableEl.innerHTML = renderIdPanel(idGroup, false);
                middleTableEl.innerHTML = renderIdPanel(tuGroup, false);
                rightTableEl.innerHTML = renderIdPanel(otherGroup, true);
            }}

            function renderGroup(group) {{
                const groupItems = getItemsByGroup(group.id);
                const allowAdd = group.id !== 4;
                const rows = groupItems.map(item => {{
                    const rowClass = normalizeStatus(item.status) === 'Не требуется' ? 'row not-required' : 'row';
                    return `
                        <div class="${{rowClass}}" data-item-id="${{esc(item.id)}}">
                            <div class="td"><div class="cell-name"><div class="${{indicatorClass(item.status)}}"></div><div class="item-name">${{esc(item.name)}}</div></div></div>
                            <div class="td">${{buildDocumentCell(item)}}</div>
                            <div class="td">
                                <select class="status-select" data-role="status" data-item-id="${{esc(item.id)}}" ${{disabledAttr()}}>
                                    <option value="" ${{normalizeStatus(item.status) === '' ? 'selected' : ''}}></option>
                                    <option value="Есть" ${{normalizeStatus(item.status) === 'Есть' ? 'selected' : ''}}>Есть</option>
                                    <option value="Нет" ${{normalizeStatus(item.status) === 'Нет' ? 'selected' : ''}}>Нет</option>
                                    <option value="Не требуется" ${{normalizeStatus(item.status) === 'Не требуется' ? 'selected' : ''}}>Не требуется</option>
                                </select>
                            </div>
                            <div class="td"><input class="date-input" type="date" data-role="plan" data-item-id="${{esc(item.id)}}" value="${{esc(toInputDate(item.plan))}}" ${{disabledAttr()}}></div>
                            <div class="td"><input class="date-input" type="date" data-role="fact" data-item-id="${{esc(item.id)}}" value="${{esc(toInputDate(item.fact))}}" ${{disabledAttr()}}></div>
                        </div>
                    `;
                }}).join('');
                const addBlock = allowAdd ? `
                    <div class="add-item-row">
                        <input class="add-item-input" id="addItemInput_${{group.id}}" type="text" placeholder="Новый пункт" ${{disabledAttr()}}>
                        <button class="add-item-btn" type="button" data-role="add-item" data-group-id="${{group.id}}" ${{disabledAttr()}}>Добавить пункт</button>
                    </div>` : '';
                return `<div class="group-block"><div class="group-title">${{esc(group.title)}}</div>${{rows}}${{addBlock}}</div>`;
            }}
            function renderTables() {{
                if (currentChecklistKey === 'concept') {{
                    renderConceptTable();
                    return;
                }}

                if (currentChecklistKey === 'id') {{
                    renderIdTables();
                    return;
                }}

                if (currentChecklistKey === 'opr') {{
                    renderOprTables();
                    return;
                }}

                if (tablesGridEl) {{
                    tablesGridEl.classList.remove('id-three-cols');
                    tablesGridEl.style.gridTemplateColumns = '1fr 1fr';
                }}

                if (tablePanels[0]) tablePanels[0].style.display = '';
                if (tablePanels[1]) tablePanels[1].style.display = '';
                if (tablePanels[2]) tablePanels[2].style.display = 'none';

                if (leftTableEl) leftTableEl.innerHTML = idTableShellHtml;
                if (middleTableEl) middleTableEl.innerHTML = idTableShellHtml;

                const leftBody = document.getElementById('leftTableBody');
                const middleBody = document.getElementById('middleTableBody');

                if (!leftBody || !middleBody) {{
                    throw new Error('leftTableBody or middleTableBody not found');
                }}

                const leftGroups = groups.filter(g => Number(g.id) === 1 || Number(g.id) === 3);
                const rightGroups = groups.filter(g => Number(g.id) === 2);

                if (hasItemsInGroup(4)) {{
                    const notRequiredGroup = groups.find(g => Number(g.id) === 4);
                    if (notRequiredGroup) {{
                        rightGroups.push(notRequiredGroup);
                    }}
                }}

                leftBody.innerHTML = leftGroups.map(renderGroup).join('');
                middleBody.innerHTML = rightGroups.map(renderGroup).join('');
            }}
            function renderAll() {{
                renderTables();
                bindEvents();
                calculateProgress();
                renderTitle();
                renderProjectChecklistList();

                if (progressBoxEl) {{
                    progressBoxEl.classList.toggle('id-accent', currentChecklistKey === 'id' || currentChecklistKey === 'opr');
                }}

                updateDebugPanelAccess();
                syncChecklistCache();
            }}
            function replaceItem(updatedItem) {{
                if (!updatedItem) return;

                const normalizedItem = Object.assign({{
                    folderKey: '',
                    folderPath: '',
                    folderUrl: '',
                    documents: [],
                    documentUrl: '',
                    documentName: ''
                }}, updatedItem || {{}});

                normalizedItem.documents = Array.isArray(normalizedItem.documents) ? normalizedItem.documents : [];

                if (!Object.prototype.hasOwnProperty.call(normalizedItem, 'documentUrl')) {{
                    normalizedItem.documentUrl = '';
                }}
                if (!Object.prototype.hasOwnProperty.call(normalizedItem, 'documentName')) {{
                    normalizedItem.documentName = '';
                }}

                const idx = items.findIndex(x => x.id === normalizedItem.id);
                if (idx >= 0) {{
                    items[idx] = Object.assign({{}}, items[idx], normalizedItem, {{
                        folderKey: normalizedItem.folderKey || '',
                        folderPath: normalizedItem.folderPath || '',
                        folderUrl: normalizedItem.folderUrl || '',
                        documents: normalizedItem.documents || [],
                        documentUrl: normalizedItem.documentUrl || '',
                        documentName: normalizedItem.documentName || ''
                    }});
                }} else {{
                    items.push(normalizedItem);
                }}
            }}
            function bindEvents() {{
                document.querySelectorAll('[data-role="concept-status"]').forEach(el => {{
                    const commitConceptStatus = function () {{
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;

                        const oldValue = item.status || '';
                        let newValue = this.value;

                        if (String(item.statusKind || '') === 'text') {{
                            newValue = formatConceptTextStatus(newValue, item.statusPlaceholder || '');
                            this.value = newValue;
                        }}

                        if (newValue === 'Не требуется') {{
                            item.group = 10;
                        }} else if (Number(item.group) === 10) {{
                            item.group = resolveConceptGroupId(item);
                        }}

                        item.status = newValue;
                        pushSessionChange(item.id, item.name, 'status', oldValue, newValue);
                        renderAll();
                    }};

                    const item = items.find(x => x.id === el.dataset.itemId);
                    if (!item) return;

                    if (String(item.statusKind || '') === 'text') {{
                        el.addEventListener('blur', commitConceptStatus);
                        el.addEventListener('keydown', function (e) {{
                            if (e.key === 'Enter') {{
                                e.preventDefault();
                                commitConceptStatus.call(this);
                            }}
                        }});
                    }} else {{
                        el.addEventListener('change', commitConceptStatus);
                    }}
                }});

                document.querySelectorAll('[data-role="concept-extra"]').forEach(el => {{
                    el.dataset.initialValue = String(el.value || '');
                    el.dataset.baseHeight = '32';
                    el.style.height = '32px';
                    el.addEventListener('input', function () {{
                        autoGrowTextarea(this);
                    }});

                    const handler = function () {{
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;

                        const oldValue = item.extraInfo || '';
                        const newValue = this.value;
                        item.extraInfo = newValue;
                        pushSessionChange(item.id, item.name, 'extraInfo', oldValue, newValue);
                        renderAll();
                    }};

                    el.addEventListener('change', handler);
                    el.addEventListener('blur', handler);
                }});
                document.querySelectorAll('[data-role="status"]').forEach(el => {{
                    el.addEventListener('change', async function() {{
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;

                        const oldItem = JSON.parse(JSON.stringify(item));
                        const newValue = this.value;
                        const oldDocuments = getItemDocuments(oldItem);

                        pushSessionChange(item.id, item.name, 'status', oldItem.status || '', newValue || '');
                        debugLog('status_changed', {{
                            itemId: item.id,
                            itemName: item.name,
                            oldValue: oldItem.status || '',
                            newValue: newValue || ''
                        }});

                        if (newValue === 'Нет') {{
                            if (oldDocuments.length) {{
                                const removedNames = oldDocuments.map(x => x.name || 'uploaded').join(', ');

                                pushSessionChange(item.id, item.name, 'document', removedNames, 'removed');
                                debugLog('document_removed_by_status', {{
                                    itemId: item.id,
                                    itemName: item.name,
                                    oldValue: removedNames,
                                    newValue: 'removed'
                                }});
                            }}

                            try {{
                                const result = await updateItem(item.id, 'status', newValue);
                                if (result && result.item) {{
                                    replaceItem(result.item);
                                }}
                            }} catch (e) {{
                                console.log('status save error:', e);
                                setSaveState('error', 'Ошибка сохранения статуса');
                                return;
                            }}

                            renderAll();
                            return;
                        }}

                        item.status = newValue;
                        renderAll();
                    }});
                }});

                document.querySelectorAll('[data-role="plan"]').forEach(el => {{
                    el.addEventListener('change', function() {{
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;
                        const oldItem = JSON.parse(JSON.stringify(item));
                        const newValue = fromInputDate(this.value);

                        item.plan = newValue;
                        pushSessionChange(item.id, item.name, 'plan', oldItem.plan || '', newValue || '');
                        debugLog('plan_changed', {{
                            itemId: item.id,
                            itemName: item.name,
                            oldValue: oldItem.plan || '',
                            newValue: newValue || ''
                        }});

                        renderAll();
                    }});
                }});

                document.querySelectorAll('[data-role="fact"]').forEach(el => {{
                    el.addEventListener('change', function() {{
                        const item = items.find(x => x.id === this.dataset.itemId);
                        if (!item) return;
                        const oldItem = JSON.parse(JSON.stringify(item));
                        const newValue = fromInputDate(this.value);

                        item.fact = newValue;
                        pushSessionChange(item.id, item.name, 'fact', oldItem.fact || '', newValue || '');
                        debugLog('fact_changed', {{
                            itemId: item.id,
                            itemName: item.name,
                            oldValue: oldItem.fact || '',
                            newValue: newValue || ''
                        }});

                        renderAll();
                    }});
                }});

                document.querySelectorAll('[data-role="add-item"]').forEach(btn => {{
                    btn.addEventListener('click', async function() {{
                        const groupId = Number(this.dataset.groupId);
                        const input = document.getElementById('addItemInput_' + groupId);
                        if (!input) return;

                        const name = (input.value || '').trim();
                        if (!name) return;

                        this.disabled = true;

                        try {{
                            const result = await addItem(groupId, name, 'id');
                            if (!result || !result.item) {{
                                throw new Error('add item failed');
                            }}

                            replaceItem(result.item);
                            pushSessionChange(result.item.id, result.item.name, 'add-item', '', result.item.name);

                            debugLog('item_added', {{
                                itemId: result.item.id,
                                itemName: result.item.name,
                                groupId: groupId
                            }});

                            input.value = '';
                            renderAll();
                        }} catch (e) {{
                            console.log('add item error:', e);
                            setSaveState('error', 'Ошибка добавления пункта');
                        }} finally {{
                            this.disabled = false;
                        }}
                    }});
                }});

                document.querySelectorAll('[data-role="toggle-id-dates"]').forEach(btn => {{
                    btn.addEventListener('click', function () {{
                        const groupId = Number(this.dataset.groupId || 0);
                        if (![1, 2, 3].includes(groupId)) return;

                        idDateVisibility[groupId] = !idDateVisibility[groupId];
                        renderAll();
                    }});
                }});

                document.querySelectorAll('[data-role="concept-add-item"]').forEach(btn => {{
                    btn.addEventListener('click', async function() {{
                        const groupId = Number(this.dataset.groupId);
                        const input = document.getElementById('conceptAddItemInput_' + groupId);
                        if (!input) return;

                        const name = (input.value || '').trim();
                        if (!name) return;

                        this.disabled = true;

                        try {{
                            const result = await addItem(groupId, name, 'concept');
                            if (!result || !result.item) {{
                                throw new Error('add concept item failed');
                            }}

                            replaceItem(result.item);
                            pushSessionChange(result.item.id, result.item.name, 'add-item', '', result.item.name);

                            debugLog('item_added', {{
                                itemId: result.item.id,
                                itemName: result.item.name,
                                groupId: groupId
                            }});

                            input.value = '';
                            renderAll();
                        }} catch (e) {{
                            console.log('concept add item error:', e);
                            setSaveState('error', 'Ошибка добавления пункта');
                        }} finally {{
                            this.disabled = false;
                        }}
                    }});
                }});

                document.querySelectorAll('[data-role="upload"]').forEach(btn => {{
                    btn.addEventListener('click', function() {{
                        const input = document.querySelector('[data-role="file-input"][data-item-id="' + this.dataset.itemId + '"]');
                        if (input) input.click();
                    }});
                }});

                document.querySelectorAll('[data-role="view-folder"]').forEach(btn => {{
                    btn.addEventListener('click', function () {{
                        const folderUrl = this.dataset.folderUrl || '';
                        if (!folderUrl) return;

                        try {{
                            window.open(folderUrl, '_blank');
                        }} catch (e) {{
                            console.log('open folder error:', e);
                            setSaveState('error', 'Ошибка открытия папки');
                        }}
                    }});
                }});

                document.querySelectorAll('[data-role="view-file"]').forEach(link => {{
                    link.addEventListener('click', function (event) {{
                        event.preventDefault();
                        event.stopPropagation();

                        const openUrl = this.dataset.openUrl || '';
                        if (!openUrl) return;

                        try {{
                            const absoluteUrl = new URL(openUrl, window.location.href).href;
                            window.open(absoluteUrl, '_blank', 'noopener,noreferrer');
                        }} catch (e) {{
                            console.log('open file error:', e);
                            setSaveState('error', 'Ошибка открытия файла');
                        }}
                    }});
                }});

                document.querySelectorAll('[data-role="remove-file"]').forEach(btn => {{
                    btn.onclick = async function (event) {{
                        event.preventDefault();
                        event.stopPropagation();

                        if (this.disabled) return;

                        const itemId = this.dataset.itemId || '';
                        const documentId = this.dataset.documentId || '';
                        const documentName = this.dataset.documentName || 'файл';

                        if (!confirmFileRemoval(documentName)) {{
                            return;
                        }}

                        const item = items.find(x => x.id === itemId);
                        const documents = getItemDocuments(item);
                        const doc = documents.find(x => String(x.id || '') === String(documentId || ''));

                        this.disabled = true;

                        try {{
                            const removeResult = await removeDocument(itemId, documentId);

                            if (removeResult && removeResult.item) {{
                                replaceItem(removeResult.item);
                            }}

                            await reloadCurrentChecklistFromServer();

                            pushSessionChange(
                                itemId,
                                item ? item.name : '',
                                'document',
                                doc ? (doc.name || documentName) : documentName,
                                'removed'
                            );

                            renderAll();
                        }} catch (e) {{
                            console.log('remove file error:', e);
                            setSaveState('error', 'Ошибка удаления файла');
                        }} finally {{
                            this.disabled = false;
                        }}
                    }};
                }});

                document.querySelectorAll('[data-role="file-input"]').forEach(input => {{
                    input.addEventListener('change', async function() {{
                        const itemId = this.dataset.itemId;
                        const files = Array.from(this.files || []);
                        if (!files.length) return;

                        const item = items.find(x => x.id === itemId);
                        const initialStatus = item ? normalizeStatus(item.status) : '';
                        let currentStatus = initialStatus;

                        try {{
                            for (const file of files) {{
                                const result = await uploadDocument(itemId, file);
                                replaceItem(result.item);

                                const uploadedDocs = getItemDocuments(result.item);
                                let uploadedDoc = uploadedDocs.find(x => String(x.name || '') === String(file.name || ''));

                                if (!uploadedDoc && uploadedDocs.length) {{
                                    uploadedDoc = uploadedDocs[uploadedDocs.length - 1];
                                }}

                                sessionChanges.push({{
                                    field: 'document',
                                    itemId: result.item ? result.item.id : itemId,
                                    itemName: result.item ? result.item.name : (item ? item.name : ''),
                                    oldValue: '',
                                    newValue: uploadedDoc ? (uploadedDoc.name || file.name || 'uploaded') : (file.name || 'uploaded')
                                }});
                                sessionDirty = true;

                                const newStatus = result.item ? normalizeStatus(result.item.status) : currentStatus;
                                if (newStatus !== currentStatus) {{
                                    sessionChanges.push({{
                                        field: 'status',
                                        itemId: result.item ? result.item.id : itemId,
                                        itemName: result.item ? result.item.name : (item ? item.name : ''),
                                        oldValue: currentStatus || '',
                                        newValue: newStatus || ''
                                    }});
                                    sessionDirty = true;
                                    currentStatus = newStatus;
                                }}
                            }}

                            renderAll();
                        }} catch (e) {{
                            console.log(e);
                            setSaveState('error', 'Ошибка загрузки файлов');
                        }} finally {{
                            this.value = '';
                        }}
                    }});
                }});
            }}

            const baseLoadChecklistByKey = loadChecklistByKey;
            loadChecklistByKey = async function (checklistKey) {{
                const targetKey = String(checklistKey || '').trim() || 'id';
                if (targetKey !== 'opr') {{
                    return baseLoadChecklistByKey(targetKey);
                }}
                if (targetKey === currentChecklistKey) {{
                    return;
                }}

                await flushCurrentChecklistSummary();
                setSaveState('saving', 'Загружаем...');

                try {{
                    const cachedData = checklistCache[targetKey];
                    if (cachedData) {{
                        applyChecklistData(deepClone(cachedData));
                        renderAll();
                        debugLog('checklist_switched_cached', {{
                            checklistKey: targetKey
                        }});
                        setSaveState('', 'Сохранено');
                        return;
                    }}

                    const response = await fetch(
                        appUrl('api/checklist') +
                        '?dialogId=' + encodeURIComponent(dialogId) +
                        '&checklistKey=' + encodeURIComponent(targetKey)
                    );
                    const result = await response.json();

                    if (!response.ok) {{
                        throw new Error(result.error || 'load checklist failed');
                    }}

                    applyChecklistData(result);
                    renderAll();
                    debugLog('checklist_switched', {{
                        checklistKey: targetKey
                    }});
                    setSaveState('', 'Сохранено');
                }} catch (e) {{
                    console.log('loadChecklistByKey error:', e);
                    setSaveState('error', 'Ошибка загрузки чек-листа');
                }}
            }};
            {popup_session_enhancements_js}

            function safeInitBx24ForPopup() {{
                function applyPopupWindowSize() {{
                    try {{
                        if (typeof window.BX24.resizeWindow === 'function') {{
                            window.BX24.resizeWindow(1180, 720);
                        }}
                        if (typeof window.BX24.fitWindow === 'function') {{
                            window.BX24.fitWindow();
                        }}
                    }} catch (e) {{
                        console.log('BX24 popup sizing error:', e);
                    }}
                }}

                try {{
                    if (window.BX24 && typeof window.BX24.init === 'function') {{
                        window.BX24.init(function () {{
                            applyPopupWindowSize();
                            setTimeout(applyPopupWindowSize, 80);
                            setTimeout(applyPopupWindowSize, 220);
                        }});
                    }}
                }} catch (e) {{
                    console.log('BX24.init skipped:', e);
                }}
            }}
            try {{
                logRenderState('before_renderAll');
                renderAll();
                logRenderState('after_renderAll');

                debugLog('popup_loaded', {{
                    href: window.location.href,
                    hasDialogId: !!dialogId
                }});
            }} catch (e) {{
                logRenderError('renderAll', e);
            }}

            try {{
                fetchChatTitleIfMissing();
            }} catch (e) {{
                logRenderError('fetchChatTitleIfMissing', e);
            }}

            try {{
                fetchCurrentUserIfPossible();
            }} catch (e) {{
                logRenderError('fetchCurrentUserIfPossible', e);
            }}

            try {{
                safeInitBx24ForPopup();
            }} catch (e) {{
                logRenderError('safeInitBx24ForPopup', e);
            }}
        </script>
    </body>
    </html>
    """

@app.get("/api/checklist")
def api_checklist(dialogId: str = "", checklistKey: str = "id"):
    dialogId = normalize_dialog_id(dialogId)
    checklistKey = normalize_checklist_key(checklistKey)
    return JSONResponse(get_checklist(dialogId, checklistKey))


@app.post("/api/checklist/update-item")
async def api_checklist_update_item(request: Request):
    payload = await request.json()

    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    checklist_key = normalize_checklist_key(payload.get("checklistKey"))
    item_id = str(payload.get("itemId") or "").strip()
    field = str(payload.get("field") or "").strip()
    value = payload.get("value")

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    if not item_id:
        return JSONResponse({"ok": False, "error": "itemId is required"}, status_code=400)

    if checklist_key == "concept":
        allowed_fields = {"priority", "status", "extraInfo"}
    elif checklist_key == "opr":
        allowed_fields = {"priority", "status", "plan", "fact"}
    else:
        allowed_fields = {"priority", "status", "plan", "fact"}

    if field not in allowed_fields:
        return JSONResponse({"ok": False, "error": "invalid field"}, status_code=400)

    data = get_checklist(dialog_id, checklist_key)
    items = data.get("items", [])

    target_item = None

    for index, item in enumerate(items):
        if str(item.get("id") or "") == item_id:
            target_item = migrate_legacy_document_fields(item)
            items[index] = target_item
            break

    if not target_item:
        return JSONResponse({"ok": False, "error": "item not found"}, status_code=404)

    if field == "priority":
        target_item["priority"] = normalize_priority(value)
    elif field == "status":
        new_value = clean_cell_value(value)

        if checklist_key == "concept":
            target_item["status"] = new_value

            status_kind = str(target_item.get("statusKind") or "").strip()

            if new_value == "Не требуется":
                target_item["group"] = 10
                target_item["priority"] = "gray"
            else:
                if target_item.get("group") == 10:
                    target_item["group"] = resolve_concept_group_id_by_item_id_or_name(target_item)

                if status_kind == "bool":
                    if new_value == "Да":
                        target_item["priority"] = "green"
                    elif new_value == "Нет":
                        target_item["priority"] = "gray"
                    else:
                        target_item["priority"] = "white"
                else:
                    target_item["priority"] = "green" if new_value else "white"
        elif checklist_key == "opr":
            new_status = normalize_status(value)
            target_item["status"] = new_status
            target_item["priority"] = derive_indicator_from_status(new_status)

            if new_status == "Не требуется":
                target_item["group"] = 2
            elif int(target_item.get("group") or 0) == 2:
                target_item["group"] = resolve_opr_group_id_by_item_id_or_name(target_item)
        else:
            new_status = normalize_status(value)
            target_item["status"] = new_status
            target_item["priority"] = derive_indicator_from_status(new_status)

            if new_status == "Нет":
                migrated_item = migrate_legacy_document_fields(target_item)
                target_item.clear()
                target_item.update(migrated_item)

                existing_documents = normalize_documents_list(target_item.get("documents"))

                for doc in existing_documents:
                    remove_item_document_file({
                        "documentUrl": clean_cell_value(doc.get("fileUrl"))
                            or clean_cell_value(doc.get("previewUrl"))
                            or clean_cell_value(doc.get("path"))
                    })

                target_item["documents"] = []
                target_item["documentUrl"] = ""
                target_item["documentName"] = ""
                target_item["folderPath"] = ""
                target_item["folderUrl"] = ""
    elif field == "plan":
        target_item["plan"] = normalize_date_string(value)
    elif field == "fact":
        target_item["fact"] = normalize_date_string(value)
    elif field == "extraInfo":
        target_item["extraInfo"] = clean_cell_value(value)

    data["items"] = items
    data = normalize_checklist_data(data, checklist_key)
    save_checklist(dialog_id, data, checklist_key)
    updated_item = None
    for item in data.get("items", []):
        if str(item.get("id")) == item_id:
            updated_item = item
            break

    return JSONResponse({
        "ok": True,
        "item": updated_item,
        "dialogId": dialog_id,
        "checklistKey": checklist_key,
        "progressPercent": data.get("progressPercent", 0),
    })

@app.post("/api/integrations/n8n/project-storage-context")
async def api_project_storage_context(request: Request):
    expected_token = N8N_SHARED_TOKEN
    provided_token = (request.headers.get("X-N8N-Token") or "").strip()

    if expected_token and provided_token != expected_token:
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)

    try:
        payload = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "invalid json"}, status_code=400)

    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    project_name = clean_cell_value(payload.get("projectName"))
    yandex_disk = payload.get("yandexDisk") or {}
    item_mappings = payload.get("itemMappings") or []

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    if not project_name:
        return JSONResponse({"ok": False, "error": "projectName is required"}, status_code=400)

    if not isinstance(yandex_disk, dict) or not yandex_disk:
        return JSONResponse({"ok": False, "error": "yandexDisk is required"}, status_code=400)

    if not isinstance(item_mappings, list):
        return JSONResponse({"ok": False, "error": "itemMappings must be a list"}, status_code=400)

    try:
        save_project_storage_context(dialog_id, payload)
    except Exception as e:
        return JSONResponse({
            "ok": False,
            "error": "failed to save storage context",
            "details": str(e),
        }, status_code=500)

    return JSONResponse({
        "ok": True,
        "dialogId": dialog_id,
        "projectId": str(payload.get("projectId") or "").strip(),
        "projectName": project_name,
        "provider": str(yandex_disk.get("provider") or "yandex_disk").strip(),
        "stored": True,
        "mappingCount": len(item_mappings),
    })

@app.get("/api/integrations/n8n/project-storage-context")
def api_get_project_storage_context(dialogId: str = ""):
    dialog_id = normalize_dialog_id(dialogId)
    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    context = get_project_storage_context(dialog_id)
    if not context:
        return JSONResponse({"ok": False, "error": "not found"}, status_code=404)

    return JSONResponse({
        "ok": True,
        "context": context
    })

@app.post("/api/checklist/update-meta")
async def api_checklist_update_meta(request: Request):
    payload = await request.json()

    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    checklist_key = normalize_checklist_key(payload.get("checklistKey"))
    field = str(payload.get("field") or "").strip()
    value = payload.get("value")

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    if field != "collabTitle":
        return JSONResponse({"ok": False, "error": "only collabTitle is supported now"}, status_code=400)

    data = get_checklist(dialog_id, checklist_key)
    data["collabTitle"] = clean_cell_value(value)

    data = normalize_checklist_data(data, checklist_key)
    save_checklist(dialog_id, data, checklist_key)

    return JSONResponse({
        "ok": True,
        "dialogId": dialog_id,
        "field": field,
        "value": data.get("collabTitle", ""),
        "progressPercent": data.get("progressPercent", 0),
    })


@app.post("/api/checklist/add-item")
async def api_checklist_add_item(request: Request):
    payload = await request.json()

    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    checklist_key = normalize_checklist_key(payload.get("checklistKey"))
    group_id = int(payload.get("groupId") or 0)
    name = clean_cell_value(payload.get("name"))

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    if checklist_key == "id" and group_id not in [1, 2, 3]:
        return JSONResponse({"ok": False, "error": "invalid groupId"}, status_code=400)

    if checklist_key == "concept" and group_id not in [group["id"] for group in CONCEPT_GROUPS if group["id"] != 10]:
        return JSONResponse({"ok": False, "error": "invalid groupId"}, status_code=400)

    if checklist_key == "opr" and group_id not in [group["id"] for group in OPR_GROUPS if group["id"] != 2]:
        return JSONResponse({"ok": False, "error": "invalid groupId"}, status_code=400)

    if not name:
        return JSONResponse({"ok": False, "error": "name is required"}, status_code=400)

    data = get_checklist(dialog_id, checklist_key)
    items = data.get("items", [])

    group_items = [x for x in items if x.get("group") == group_id]
    next_order = len(group_items) + 1

    if checklist_key == "concept":
        new_item_id = f"concept_g{group_id}_custom_{uuid.uuid4().hex[:8]}"
        new_item = {
            "id": new_item_id,
            "group": group_id,
            "order": next_order,
            "name": name,
            "source": "",
            "statusKind": "text",
            "statusOptions": [],
            "statusPlaceholder": "",
            "status": "",
            "extraInfo": "",
            "extraInfoPlaceholder": "",
            "folderKey": build_folder_key("concept", name, new_item_id),
            "folderPath": "",
            "folderUrl": "",
            "documents": [],
            "documentUrl": "",
            "documentName": "",
            "isCustom": True,
            "priority": "white",
        }
    elif checklist_key == "opr":
        new_item_id = f"opr_g{group_id}_custom_{uuid.uuid4().hex[:8]}"
        new_item = {
            "id": new_item_id,
            "group": group_id,
            "order": next_order,
            "name": name,
            "priority": "white",
            "status": "",
            "plan": "",
            "fact": "",
            "folderKey": build_folder_key("opr", name, new_item_id),
            "folderPath": "",
            "folderUrl": "",
            "documents": [],
            "documentUrl": "",
            "documentName": "",
            "isCustom": True,
        }

        try:
            ensure_yandex_folder_for_custom_opr_item(
                dialog_id=dialog_id,
                checklist_key=checklist_key,
                group_id=group_id,
                item_name=name,
                item_id=new_item_id,
            )
        except Exception as e:
            return JSONResponse({
                "ok": False,
                "error": "failed to create yandex folder for custom OPR item",
                "details": str(e),
            }, status_code=500)

    items.append(new_item)
    data["items"] = items
    data = normalize_checklist_data(data, checklist_key)
    save_checklist(dialog_id, data, checklist_key)

    created_item = None
    for item in data.get("items", []):
        if str(item.get("id")) == new_item["id"]:
            created_item = item
            break

    return JSONResponse({
        "ok": True,
        "dialogId": dialog_id,
        "item": created_item or new_item,
        "progressPercent": data.get("progressPercent", 0),
    })

@app.post("/api/checklist/upload-document")
async def api_checklist_upload_document(
    dialogId: str = Form(...),
    itemId: str = Form(...),
    file: UploadFile = File(...),
    checklistKey: str = Form("id"),
    itemGroup: str = Form("")
):
    dialog_id = normalize_dialog_id(dialogId)
    checklist_key = normalize_checklist_key(checklistKey)
    item_id = str(itemId or "").strip()
    item_group = int(str(itemGroup or "0").strip() or 0)

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    if not item_id:
        return JSONResponse({"ok": False, "error": "itemId is required"}, status_code=400)

    data = get_checklist(dialog_id, checklist_key)
    items = data.get("items", []) or []

    target_item = None
    for index, item in enumerate(items):
        if str(item.get("id") or "") == item_id:
            target_item = migrate_legacy_document_fields(item)
            items[index] = target_item
            break

    if not target_item:
        return JSONResponse({"ok": False, "error": "item not found"}, status_code=404)

    rel_path = build_upload_rel_path(dialog_id, item_id, file.filename or "file.bin")
    abs_path = UPLOAD_ROOT / rel_path
    abs_path.parent.mkdir(parents=True, exist_ok=True)

    file_bytes = await file.read()
    with open(abs_path, "wb") as f:
        f.write(file_bytes)

    file_url = "/uploads/" + rel_path.replace("\\", "/")
    document_id = uuid.uuid4().hex
    document_view_url = build_document_view_url(dialog_id, checklist_key, item_id, document_id)
    folder_view_url = build_folder_view_url(dialog_id, checklist_key, item_id)

    try:
        folder_path = "/" + str(abs_path.parent.relative_to(BASE_DIR)).replace("\\", "/")
    except Exception:
        folder_path = file_url.rsplit("/", 1)[0]

    uploaded_name = Path(file.filename or "file.bin").name

    document_record = normalize_document_record({
        "id": document_id,
        "name": uploaded_name,
        "path": file_url,
        "fileUrl": file_url,
        "previewUrl": document_view_url,
        "size": len(file_bytes),
        "modifiedAt": datetime.now().isoformat(timespec="seconds"),
        "source": "local",

        "mirrorStatus": "",
        "mirrorError": "",
        "yandexPath": "",
        "yandexFileUrl": "",
        "yandexFolderAlias": "",
    })

    try:
        mirror_result = mirror_document_to_yandex(
            dialog_id=dialog_id,
            checklist_key=checklist_key,
            item_name=clean_cell_value(target_item.get("name")),
            filename=uploaded_name,
            file_bytes=file_bytes,
            item_id=str(target_item.get("id") or ""),
            item_group=int(target_item.get("group") or 0),
            is_custom=bool(target_item.get("isCustom", False)),
        )

        if mirror_result.get("ok"):
            document_record["mirrorStatus"] = "synced"
            document_record["mirrorError"] = ""
            document_record["yandexPath"] = clean_cell_value(mirror_result.get("filePath"))
            document_record["yandexFileUrl"] = clean_cell_value(mirror_result.get("folderUrl"))
            document_record["yandexFolderAlias"] = clean_cell_value(mirror_result.get("folderAlias"))
        else:
            document_record["mirrorStatus"] = "error"
            document_record["mirrorError"] = clean_cell_value(mirror_result.get("reason")) or "mirror failed"
    except Exception as e:
        document_record["mirrorStatus"] = "error"
        document_record["mirrorError"] = str(e)

    existing_documents = normalize_documents_list(target_item.get("documents"))
    existing_documents.append(document_record)
    normalized_documents = normalize_documents_list(existing_documents)

    target_item["documents"] = normalized_documents
    target_item["folderPath"] = folder_path
    target_item["folderUrl"] = folder_view_url if normalized_documents else ""

    first_doc = normalized_documents[0] if normalized_documents else {}
    target_item["documentUrl"] = clean_cell_value(first_doc.get("fileUrl"))
    target_item["documentName"] = clean_cell_value(first_doc.get("name"))

    actual_group = int(target_item.get("group") or item_group or 0)
    if checklist_key == "id" and actual_group != 4:
        target_item["status"] = "Есть"
        target_item["priority"] = derive_indicator_from_status("Есть")
    elif checklist_key == "opr" and actual_group != 2:
        target_item["status"] = "Есть"
        target_item["priority"] = derive_indicator_from_status("Есть")

    data["items"] = items
    data = normalize_checklist_data(data, checklist_key)
    save_checklist(dialog_id, data, checklist_key)

    updated_item = None
    for item in data.get("items", []):
        if str(item.get("id") or "") == item_id:
            updated_item = item
            break

    if not updated_item:
        return JSONResponse({"ok": False, "error": "updated item not found"}, status_code=500)

    return JSONResponse({
        "ok": True,
        "dialogId": dialog_id,
        "checklistKey": checklist_key,
        "item": updated_item,
        "progressPercent": data.get("progressPercent", 0),
    })


@app.post("/api/checklist/remove-document")
async def api_checklist_remove_document(request: Request):
    payload = await request.json()

    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    checklist_key = normalize_checklist_key(payload.get("checklistKey"))
    item_id = str(payload.get("itemId") or "").strip()
    document_id = clean_cell_value(payload.get("documentId"))
    document_url = clean_cell_value(payload.get("documentUrl"))
    preserve_status = bool(payload.get("preserveStatus"))

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    if not item_id:
        return JSONResponse({"ok": False, "error": "itemId is required"}, status_code=400)

    data = get_checklist(dialog_id, checklist_key)
    items = data.get("items", []) or []

    target_item = None
    for index, item in enumerate(items):
        if str(item.get("id") or "") == item_id:
            target_item = migrate_legacy_document_fields(item)
            items[index] = target_item
            break

    if not target_item:
        return JSONResponse({"ok": False, "error": "item not found"}, status_code=404)

    documents = normalize_documents_list(target_item.get("documents"))

    doc_to_remove = None

    if document_id:
        for doc in documents:
            if str(doc.get("id") or "") == document_id:
                doc_to_remove = doc
                break

    if not doc_to_remove and document_url:
        for doc in documents:
            doc_file_url = clean_cell_value(doc.get("fileUrl"))
            doc_preview_url = clean_cell_value(doc.get("previewUrl"))
            doc_path = clean_cell_value(doc.get("path"))
            if document_url in {doc_file_url, doc_preview_url, doc_path}:
                doc_to_remove = doc
                break

    if not doc_to_remove and documents:
        doc_to_remove = documents[0]

    if doc_to_remove:
        local_document_url = (
            clean_cell_value(doc_to_remove.get("fileUrl"))
            or clean_cell_value(doc_to_remove.get("previewUrl"))
            or clean_cell_value(doc_to_remove.get("path"))
        )

        remove_item_document_file({
            "documentUrl": local_document_url
        })

        yandex_path = clean_cell_value(doc_to_remove.get("yandexPath"))
        if yandex_path and is_yandex_disk_enabled():
            try:
                yandex_disk_delete_path(yandex_path, permanently=True)
            except Exception as e:
                write_debug_log("yandex_mirror_delete_error", {
                    "dialogId": dialog_id,
                    "checklistKey": checklist_key,
                    "itemId": item_id,
                    "documentId": clean_cell_value(doc_to_remove.get("id")),
                    "yandexPath": yandex_path,
                    "error": str(e),
                })

    remaining_documents = []
    removed = False

    for doc in documents:
        same_id = document_id and str(doc.get("id") or "") == document_id
        same_url = document_url and document_url in {
            clean_cell_value(doc.get("fileUrl")),
            clean_cell_value(doc.get("previewUrl")),
            clean_cell_value(doc.get("path")),
        }

        if not removed and (same_id or same_url or (doc_to_remove and str(doc.get("id") or "") == str(doc_to_remove.get("id") or ""))):
            removed = True
            continue

        remaining_documents.append(doc)

    normalized_documents = normalize_documents_list(remaining_documents)
    target_item["documents"] = normalized_documents

    first_doc = normalized_documents[0] if normalized_documents else {}
    target_item["documentUrl"] = clean_cell_value(first_doc.get("fileUrl"))
    target_item["documentName"] = clean_cell_value(first_doc.get("name"))

    if normalized_documents:
        first_file_url = clean_cell_value(first_doc.get("fileUrl"))
        target_item["folderPath"] = first_file_url.rsplit("/", 1)[0] if first_file_url.startswith("/") else ""
        target_item["folderUrl"] = build_folder_view_url(dialog_id, checklist_key, item_id)
    else:
        target_item["folderPath"] = ""
        target_item["folderUrl"] = ""
        target_item["documentUrl"] = ""
        target_item["documentName"] = ""

        if checklist_key == "id" and int(target_item.get("group") or 0) != 4 and not preserve_status:
            target_item["status"] = ""
            target_item["priority"] = "white"
        elif checklist_key == "opr" and int(target_item.get("group") or 0) != 2 and not preserve_status:
            target_item["status"] = ""
            target_item["priority"] = "white"

    data["items"] = items
    data = normalize_checklist_data(data, checklist_key)
    save_checklist(dialog_id, data, checklist_key)

    updated_item = None
    for item in data.get("items", []):
        if str(item.get("id") or "") == item_id:
            updated_item = item
            break

    if not updated_item:
        return JSONResponse({"ok": False, "error": "updated item not found"}, status_code=500)

    return JSONResponse({
        "ok": True,
        "dialogId": dialog_id,
        "checklistKey": checklist_key,
        "item": updated_item,
        "progressPercent": data.get("progressPercent", 0),
    })

@app.get("/api/checklist/folder", response_class=HTMLResponse)
def api_checklist_folder(dialogId: str = "", itemId: str = "", checklistKey: str = "id"):
    dialog_id = normalize_dialog_id(dialogId)
    checklist_key = normalize_checklist_key(checklistKey)
    item_id = str(itemId or "").strip()

    if not dialog_id or not item_id:
        return HTMLResponse("<h3>Не переданы dialogId или itemId</h3>", status_code=400)

    data = get_checklist(dialog_id, checklist_key)
    items = data.get("items", []) or []

    target_item = None

    for index, item in enumerate(items):
        if str(item.get("id") or "") == item_id:
            target_item = migrate_legacy_document_fields(item)
            items[index] = target_item
            break

    if not target_item:
        return JSONResponse({"ok": False, "error": "item not found"}, status_code=404)
    documents = normalize_documents_list(target_item.get("documents"))
    yandex_folder_data = get_item_yandex_folder(
        dialog_id,
        checklist_key,
        clean_cell_value(target_item.get("name"))
    )
    yandex_folder = (yandex_folder_data or {}).get("folder") or {}
    yandex_folder_url = clean_cell_value(yandex_folder.get("url"))
    yandex_folder_path = clean_cell_value(yandex_folder.get("path"))
    rows = []
    for doc in documents:
        doc_id = str(doc.get("id") or "")
        doc_name = html.escape(str(doc.get("name") or "Файл"))
        doc_size = html.escape(format_file_size(doc.get("size") or 0))
        open_url = build_document_view_url(dialog_id, checklist_key, item_id, doc_id)
        download_url = open_url + "&download=1"

        rows.append(f"""
            <tr>
                <td style="padding:10px 12px;border-bottom:1px solid #edf0f2;">{doc_name}</td>
                <td style="padding:10px 12px;border-bottom:1px solid #edf0f2;white-space:nowrap;">{doc_size}</td>
                <td style="padding:10px 12px;border-bottom:1px solid #edf0f2;white-space:nowrap;">
                    <a href="{html.escape(open_url)}" target="_blank">Открыть</a>
                    &nbsp;|&nbsp;
                    <a href="{html.escape(download_url)}" target="_blank">Скачать</a>
                    &nbsp;|&nbsp;
                    <button
                        type="button"
                        data-role="folder-remove-file"
                        data-dialog-id="{html.escape(dialog_id)}"
                        data-checklist-key="{html.escape(checklist_key)}"
                        data-item-id="{html.escape(item_id)}"
                        data-document-id="{html.escape(doc_id)}"
                        data-document-name="{doc_name}"
                        style="border:none;background:transparent;color:#b42318;cursor:pointer;font-size:16px;line-height:1;padding:0 2px;"
                        title="Удалить файл"
                    >
                        ×
                    </button>
                </td>
            </tr>
        """)

    table_html = "".join(rows) if rows else """
        <tr>
            <td colspan="3" style="padding:14px 12px;color:#667085;">В папке пока нет файлов</td>
        </tr>
    """

    title = html.escape(str(target_item.get("name") or "Папка"))
    checklist_title = html.escape(str(data.get("title") or "Чек-лист"))
    remove_api_url = html.escape(f"{normalize_base_path(APP_BASE_PATH)}/api/checklist/remove-document")

    yandex_button_html = ""
    if yandex_folder_url:
        yandex_button_html = f'''
            <div style="margin-top:12px;">
                <a
                    href="{html.escape(yandex_folder_url)}"
                    target="_blank"
                    style="display:inline-block;padding:8px 12px;border:1px solid #d0d7de;border-radius:8px;background:#f8fafc;color:#1f2328;text-decoration:none;"
                >
                    Открыть папку на Яндекс Диске
                </a>
            </div>
        '''
    elif yandex_folder_path:
        yandex_button_html = f'''
            <div style="margin-top:12px;font-size:12px;color:#667085;">
                Папка Яндекс Диска: {html.escape(yandex_folder_path)}
            </div>
        '''

    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>{title}</title>
    </head>
    <body style="font-family:Arial,sans-serif;background:#f8fafc;margin:0;padding:24px;color:#1f2328;">
        <div style="max-width:960px;margin:0 auto;background:#fff;border:1px solid #e5e7eb;border-radius:14px;overflow:hidden;">
            <div style="padding:16px 18px;border-bottom:1px solid #edf0f2;background:#fafbfc;">
                <div style="font-size:13px;color:#667085;margin-bottom:4px;">{checklist_title}</div>
                <div style="font-size:22px;font-weight:700;">{title}</div>
                {yandex_button_html}
            </div>
            <div style="padding:18px;">
                <table style="width:100%;border-collapse:collapse;">
                    <thead>
                        <tr>
                            <th style="text-align:left;padding:10px 12px;background:#f8fafc;border-bottom:1px solid #e5e7eb;">Файл</th>
                            <th style="text-align:left;padding:10px 12px;background:#f8fafc;border-bottom:1px solid #e5e7eb;">Размер</th>
                            <th style="text-align:left;padding:10px 12px;background:#f8fafc;border-bottom:1px solid #e5e7eb;">Действия</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_html}
                    </tbody>
                </table>
            </div>
        </div>
        <script>
            const folderRemoveApiUrl = "{remove_api_url}";

            document.querySelectorAll('[data-role="folder-remove-file"]').forEach(btn => {{
                btn.addEventListener('click', async function () {{
                    const documentName = this.dataset.documentName || 'файл';
                    if (!window.confirm('Удалить файл "' + documentName + '"?')) {{
                        return;
                    }}

                    this.disabled = true;

                    try {{
                        const response = await fetch(folderRemoveApiUrl, {{
                            method: 'POST',
                            headers: {{ 'Content-Type': 'application/json' }},
                            body: JSON.stringify({{
                                dialogId: this.dataset.dialogId,
                                checklistKey: this.dataset.checklistKey,
                                itemId: this.dataset.itemId,
                                documentId: this.dataset.documentId
                            }})
                        }});

                        const result = await response.json();
                        if (!response.ok || !result.ok) {{
                            throw new Error(result.error || 'remove document failed');
                        }}

                        try {{
                            if (window.opener && typeof window.opener.postMessage === 'function') {{
                                window.opener.postMessage({{
                                    type: 'checklist-document-removed',
                                    dialogId: this.dataset.dialogId,
                                    checklistKey: this.dataset.checklistKey,
                                    itemId: this.dataset.itemId
                                }}, '*');
                            }}
                        }} catch (e) {{
                            console.log('opener sync error:', e);
                        }}

                        window.location.reload();
                    }} catch (e) {{
                        console.log('folder remove error:', e);
                        alert('Ошибка удаления файла');
                    }} finally {{
                        this.disabled = false;
                    }}
                }});
            }});
        </script>
    </body>
    </html>
    """

@app.get("/api/checklist/file")
def api_checklist_file(
    dialogId: str = "",
    itemId: str = "",
    documentId: str = "",
    checklistKey: str = "id",
    download: int = 0
):
    dialog_id = normalize_dialog_id(dialogId)
    checklist_key = normalize_checklist_key(checklistKey)
    item_id = str(itemId or "").strip()
    document_id = str(documentId or "").strip()

    if not dialog_id or not item_id or not document_id:
        return JSONResponse({"ok": False, "error": "dialogId, itemId and documentId are required"}, status_code=400)

    data = get_checklist(dialog_id, checklist_key)
    items = data.get("items", []) or []

    target_item = None
    for item in items:
        if str(item.get("id") or "") == item_id:
            target_item = item
            break

    if not target_item:
        return JSONResponse({"ok": False, "error": "item not found"}, status_code=404)

    target_item = migrate_legacy_document_fields(target_item)
    documents = normalize_documents_list(target_item.get("documents"))

    target_doc = None
    for doc in documents:
        if str(doc.get("id") or "") == document_id:
            target_doc = doc
            break

    if not target_doc:
        return JSONResponse({"ok": False, "error": "document not found"}, status_code=404)

    file_url = clean_cell_value(target_doc.get("fileUrl")) or clean_cell_value(target_doc.get("path"))
    file_path = get_upload_file_path_from_url(file_url)

    if not file_path or not file_path.exists():
        return JSONResponse({"ok": False, "error": "file not found on disk"}, status_code=404)

    filename = clean_cell_value(target_doc.get("name")) or file_path.name
    media_type, _ = mimetypes.guess_type(str(file_path))
    media_type = media_type or "application/octet-stream"

    inline_allowed = can_preview_in_browser(filename, media_type)
    disposition = "attachment" if download else ("inline" if inline_allowed else "attachment")

    response = FileResponse(
        path=str(file_path),
        media_type=media_type
    )
    response.headers["Content-Disposition"] = f"{disposition}; filename*=UTF-8''{quote(filename)}"
    return response

@app.post("/api/debug/event")
async def api_debug_event(request: Request):
    try:
        payload = await request.json()
    except Exception:
        raw = await request.body()
        try:
            payload = json.loads(raw.decode("utf-8"))
        except Exception:
            payload = {"raw": raw.decode("utf-8", errors="ignore")}

    event = str(payload.get("event") or "unknown").strip()
    write_debug_log(event, payload)

    return JSONResponse({"ok": True})


@app.post("/api/checklist/lock/acquire")
async def api_checklist_lock_acquire(request: Request):
    payload = await request.json()
    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    checklist_key = normalize_checklist_key(payload.get("checklistKey"))
    user_id = str(payload.get("userId") or "").strip()
    user_name = str(payload.get("userName") or "").strip()
    lock_id = str(payload.get("lockId") or "").strip()

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    result = acquire_checklist_lock(dialog_id, checklist_key, user_id, user_name, lock_id)
    write_debug_log("lock_acquire", result)
    return JSONResponse(result)


@app.post("/api/checklist/lock/heartbeat")
async def api_checklist_lock_heartbeat(request: Request):
    payload = await request.json()
    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    checklist_key = normalize_checklist_key(payload.get("checklistKey"))
    user_id = str(payload.get("userId") or "").strip()
    user_name = str(payload.get("userName") or "").strip()
    lock_id = str(payload.get("lockId") or "").strip()

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    result = heartbeat_checklist_lock(dialog_id, checklist_key, user_id, user_name, lock_id)
    return JSONResponse(result)


@app.post("/api/checklist/lock/release")
async def api_checklist_lock_release(request: Request):
    try:
        payload = await request.json()
    except Exception:
        raw = await request.body()
        try:
            payload = json.loads(raw.decode("utf-8") or "{}")
        except Exception:
            payload = {"raw": raw.decode("utf-8", errors="ignore")}

    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    checklist_key = normalize_checklist_key(payload.get("checklistKey"))
    lock_id = str(payload.get("lockId") or "").strip()
    user_id = str(payload.get("userId") or "").strip()

    if not dialog_id:
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    result = release_checklist_lock(dialog_id, checklist_key, lock_id, user_id)
    write_debug_log("lock_release", result)
    return JSONResponse(result)


@app.post("/api/checklist/close-session")
async def api_checklist_close_session(request: Request):
    try:
        payload = await request.json()
    except Exception:
        raw = await request.body()
        try:
            payload = json.loads(raw.decode("utf-8") or "{}")
        except Exception:
            payload = {"raw": raw.decode("utf-8", errors="ignore")}

    write_debug_log("close_session_received", payload)

    dialog_id = normalize_dialog_id(payload.get("dialogId"))
    editor = payload.get("editor") or {}
    raw_sessions = payload.get("sessions") or []

    def build_message_failure_response(base_dialog_id: str, extra: dict | None = None):
        response = {
            "ok": True,
            "saved": True,
            "messageOk": False,
            "dialogId": base_dialog_id,
        }
        if extra:
            response.update(extra)
        return JSONResponse(response)

    if raw_sessions:
        sessions = []
        for raw_session in raw_sessions:
            checklist_key = normalize_checklist_key(raw_session.get("checklistKey"))
            session_dialog_id = normalize_dialog_id(raw_session.get("dialogId") or dialog_id)
            changes = raw_session.get("changes") or []
            session_data = raw_session.get("data") or {}

            if not session_dialog_id:
                continue

            if changes and session_data:
                session_data = dict(session_data)
                session_data["checklistKey"] = checklist_key
                session_data["resolvedDialogId"] = session_dialog_id
                data = normalize_checklist_data(session_data, checklist_key)
                data["resolvedDialogId"] = session_dialog_id
                save_checklist(session_dialog_id, data, checklist_key)
            else:
                data = get_checklist(session_dialog_id, checklist_key)

            sessions.append({
                "dialogId": session_dialog_id,
                "checklistKey": checklist_key,
                "changes": changes,
                "data": data,
            })

        if not sessions:
            write_debug_log("close_session_skipped", {
                "dialogId": dialog_id,
                "reason": "no sessions"
            })
            return JSONResponse({"ok": True, "skipped": True, "reason": "no sessions"})

        visible_sessions = [session for session in sessions if build_recent_changes_sections(session["changes"], session["checklistKey"])]
        if not visible_sessions:
            write_debug_log("close_session_message_skipped", {
                "dialogId": dialog_id or sessions[0]["dialogId"],
                "reason": "no visible message changes",
                "sessions": [
                    {
                        "checklistKey": session["checklistKey"],
                        "changesCount": len(session["changes"]),
                    }
                    for session in sessions
                ]
            })
            return JSONResponse({
                "ok": True,
                "dialogId": dialog_id or sessions[0]["dialogId"],
                "saved": True,
                "messageSkipped": True,
            })

        target_dialog_id = dialog_id or sessions[0]["dialogId"]

        try:
            message = build_multi_checklist_chat_message(visible_sessions, editor)
            result = bitrix_webhook_call("im.message.add", {
                "DIALOG_ID": target_dialog_id,
                "MESSAGE": message,
            })
        except Exception as exc:
            write_debug_log("close_session_message_exception", {
                "dialogId": target_dialog_id,
                "editor": editor,
                "checklistKeys": [session["checklistKey"] for session in visible_sessions],
                "error": str(exc),
            })
            return build_message_failure_response(target_dialog_id, {
                "checklistKeys": [session["checklistKey"] for session in sessions],
                "messageError": str(exc),
            })

        write_debug_log("close_session_im_message_add_result", {
            "dialogId": target_dialog_id,
            "changesCount": sum(len(session["changes"]) for session in visible_sessions),
            "editor": editor,
            "result": result,
            "checklistKeys": [session["checklistKey"] for session in visible_sessions],
        })

        if "error" in result:
            return build_message_failure_response(target_dialog_id, {
                "checklistKeys": [session["checklistKey"] for session in sessions],
                "messageError": result.get("error_description") or result.get("error") or "message send failed",
                "result": result,
            })

        return JSONResponse({
            "ok": True,
            "saved": True,
            "messageOk": True,
            "dialogId": target_dialog_id,
            "result": result,
            "checklistKeys": [session["checklistKey"] for session in sessions],
        })

    checklist_key = normalize_checklist_key(payload.get("checklistKey"))
    changes = payload.get("changes") or []
    session_data = payload.get("data") or {}

    if not dialog_id:
        write_debug_log("close_session_invalid", {
            "reason": "dialogId is required",
            "payload": payload
        })
        return JSONResponse({"ok": False, "error": "dialogId is required"}, status_code=400)

    if not changes:
        write_debug_log("close_session_skipped", {
            "dialogId": dialog_id,
            "checklistKey": checklist_key,
            "reason": "no changes"
        })
        return JSONResponse({"ok": True, "skipped": True, "reason": "no changes"})

    if session_data:
        session_data = dict(session_data)
        session_data["checklistKey"] = checklist_key
        session_data["resolvedDialogId"] = dialog_id
        data = normalize_checklist_data(session_data, checklist_key)
        data["resolvedDialogId"] = dialog_id
        save_checklist(dialog_id, data, checklist_key)
    else:
        data = get_checklist(dialog_id, checklist_key)

    if not build_recent_changes_sections(changes, checklist_key):
        write_debug_log("close_session_message_skipped", {
            "dialogId": dialog_id,
            "checklistKey": checklist_key,
            "reason": "no visible message changes",
            "changesCount": len(changes),
        })
        return JSONResponse({
            "ok": True,
            "dialogId": dialog_id,
            "checklistKey": checklist_key,
            "saved": True,
            "messageSkipped": True,
        })

    try:
        message = build_checklist_chat_message(data, changes, editor)
        result = bitrix_webhook_call("im.message.add", {
            "DIALOG_ID": dialog_id,
            "MESSAGE": message,
        })
    except Exception as exc:
        write_debug_log("close_session_message_exception", {
            "dialogId": dialog_id,
            "checklistKey": checklist_key,
            "editor": editor,
            "error": str(exc)
        })
        return build_message_failure_response(dialog_id, {
            "checklistKey": checklist_key,
            "messageError": str(exc),
        })

    write_debug_log("close_session_im_message_add_result", {
        "dialogId": dialog_id,
        "checklistKey": checklist_key,
        "changesCount": len(changes),
        "editor": editor,
        "result": result
    })

    if "error" in result:
        return build_message_failure_response(dialog_id, {
            "checklistKey": checklist_key,
            "messageError": result.get("error_description") or result.get("error") or "message send failed",
            "result": result,
        })

    return JSONResponse({
        "ok": True,
        "saved": True,
        "messageOk": True,
        "dialogId": dialog_id,
        "checklistKey": checklist_key,
        "result": result,
    })


@app.get("/debug/logs", response_class=HTMLResponse)
def debug_logs(userId: str = ""):
    allowed_debug_user_ids = {"138", "18"}
    normalized_user_id = str(userId or "").strip()

    if normalized_user_id not in allowed_debug_user_ids:
        return HTMLResponse(
            """
            <html>
            <head>
                <meta charset="utf-8">
                <title>Access denied</title>
            </head>
            <body style="font-family:Arial,sans-serif;padding:24px">
                <h1>Доступ запрещён</h1>
                <p>Эта страница доступна только техническим пользователям.</p>
            </body>
            </html>
            """,
            status_code=403
        )

    if not DEBUG_LOG_PATH.exists():
        content = "Логов пока нет"
    else:
        with open(DEBUG_LOG_PATH, "r", encoding="utf-8") as f:
            content = f.read() or "Логов пока нет"

    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>Debug Logs</title>
    </head>
    <body style="font-family:Arial,sans-serif;padding:24px">
        <h1>Debug logs</h1>
        <pre style="white-space:pre-wrap;word-break:break-word;">{html.escape(content)}</pre>
    </body>
    </html>
    """

@app.get("/admin", response_class=HTMLResponse)
def admin():
    conn = get_conn()
    rows = conn.execute("SELECT dialog_id, title FROM checklists ORDER BY dialog_id").fetchall()
    conn.close()

    items = "".join(
        f"<li><b>{html.escape(row['dialog_id'])}</b> — {html.escape(row['title'])}</li>"
        for row in rows
    ) or "<li>Пока ничего не загружено</li>"

    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>Загрузка чек-листа</title>
    </head>
    <body style="font-family:Arial,sans-serif;padding:40px;max-width:900px">
        <h1>Загрузка Excel для коллабы</h1>
        <p>Шаг 1: откройте коллабу и посмотрите значение <b>dialogId</b> в sidebar.</p>
        <p>Шаг 2: вставьте этот dialogId сюда и загрузите .xlsx файл.</p>

        <form action="/admin/upload" method="post" enctype="multipart/form-data" style="margin:30px 0">
            <div style="margin-bottom:16px">
                <label>dialogId</label><br>
                <input type="text" name="dialog_id" required style="width:100%;padding:10px">
            </div>

            <div style="margin-bottom:16px">
                <label>XLSX файл</label><br>
                <input type="file" name="file" accept=".xlsx" required>
            </div>

            <button type="submit" style="padding:10px 16px">Загрузить чек-лист</button>
        </form>

        <h2>Уже загруженные чек-листы</h2>
        <ul>{items}</ul>
    </body>
    </html>
    """


@app.post("/admin/upload", response_class=HTMLResponse)
async def admin_upload(dialog_id: str = Form(...), file: UploadFile = File(...)):
    dialog_id = normalize_dialog_id(dialog_id)
    file_bytes = await file.read()
    data = parse_xlsx_to_checklist(file_bytes)
    save_checklist(dialog_id, data)

    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>Готово</title>
    </head>
    <body style="font-family:Arial,sans-serif;padding:40px">
        <h1>Чек-лист сохранён</h1>
        <p><b>dialogId:</b> {html.escape(dialog_id)}</p>
        <p><b>Файл:</b> {html.escape(file.filename or '')}</p>
        <p>Теперь вернитесь в коллабу и обновите sidebar.</p>
        <p><a href="/admin">Назад в /admin</a></p>
    </body>
    </html>
    """
